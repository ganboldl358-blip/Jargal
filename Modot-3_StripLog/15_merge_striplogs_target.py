#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
15_merge_striplogs_target.py — Модот-3 (XV-020181) элемент-зорилтот цооногийн багана (T1–T4):
загвар тус бүрд НЭГ нэгтгэсэн PDF = нүүр + агуулгын хүснэгт + тайлбар хуудас + цооног бүрийн PDF (bookmark-тай).

Эх (12_make_striplog_MT_A3L_target.py-ийн гаралт):
    <Out_Target>/<HOLE>/<HOLE>_DrillLog_A3L_<TAG>_2026.pdf      TAG = T1_Zn-Ag · T2_Mo-W-Sn · T3_Ag-As-Sb · T4_GeoLog
Гаралт:
    <Out_Target>/_Merged/Modot3_StripLog_<TAG>_<YYYYMMDD>.pdf   (загвар бүрд нэг файл)
    <Out_Target>/_Merged/Merge_index_<YYYYMMDD>.md              (хуудас, хэмжээ, тулгалт)
Дата: MT_Drilling_Database.xlsx (Collar · Sig_Intervals · Assay · Lithology · PhotoLog · Codes · Assay_CaF2)
      hole_template_map.csv (Template + Second_template_if_any) — байвал тулгалтад ашиглана.

Ажиллуулах (Windows, D:/G: default замтай):
    py 15_merge_striplogs_target.py
    py 15_merge_striplogs_target.py --templates T1 T2 --date 20260903
    py 15_merge_striplogs_target.py --out-target "D:/.../Out_Target" --db "G:/.../MT_Drilling_Database.xlsx" --map "D:/.../hole_template_map.csv"
Шаардлага: py ≥3.9, openpyxl, pypdf, matplotlib.

14_merge_striplogs.py (v3.2 жонш) -ийн загварыг дагав: нүүр + агуулга + тайлбар + bookmark; цооногийн файлууд хэвээр.
"""
import argparse
import collections
import csv
import datetime as dt
import math
import os
import re
import sys
import tempfile
from pathlib import Path

# ----------------------------------------------------------------------------------------------
# Тохиргоо
# ----------------------------------------------------------------------------------------------
DEFAULT_OUT_TARGET = "D:/AZ9/_00_Work_Logs/Modot-3_StripLog/Out_Target"
DEFAULT_MAP = "D:/AZ9/_00_Work_Logs/Modot-3_StripLog/Templates/hole_template_map.csv"
DEFAULT_DBS = [
    "G:/My Drive/JG GeoHub/01_Projects/XV-020181_Dornogobi Airag_Modot-3/09_Drilling/01_Drilling_Database/MT_Drilling_Database.xlsx",
    "D:/AZ9/_00_Work_Logs/Modot-3_Prep/MT_Drilling_Database.xlsx",
    "D:/AZ9/_00_Work_Logs/Modot-3_Prep/MT_DB.xlsx",
]
PAGE_M = 23.0            # 1 хуудас = 23 м (A3 хөндлөн, 12_ скриптийн масштаб) → хуудас = ceil(TD/23) + 1 (тайлбар)
A3L_IN = (16.535, 11.693)  # A3 хөндлөн, инч
PROJECT_MN = "Модот-3 (XV-020181) · Дорноговь, Айраг · Батбадмаараг ХХК"

TEMPLATES = {
    "T1": dict(
        tag="T1_Zn-Ag",
        name="Zn–Ag (Pb–Mn–Cd–In)",
        kind="скарн / карбонат-орлуулалтын (CRD) төрлийн хүдэржилт",
        toc_cols=[("Zn_ppm", "Zn макс\nppm"), ("Ag_ppm", "Ag макс\nppm"), ("Pb_ppm", "Pb макс\nppm"),
                  ("Mn_ppm", "Mn макс\nppm"), ("Cd_ppm", "Cd макс\nppm")],
        sig=[("Zn", 1000), ("Ag", 3)],
        thresholds="Zn ≥1000 ppm · Ag ≥3 ppm · Pb ≥200 ppm · Cd/Zn харьцаа (<10 → өндөр T скарн) · Mn — Zn-тэй харьцуулна",
    ),
    "T2": dict(
        tag="T2_Mo-W-Sn",
        name="Mo–W–Sn (Be–Li)",
        kind="грейзен / боржинтой холбоотой хүдэржилт",
        toc_cols=[("Mo_ppm", "Mo макс\nppm"), ("W_ppm", "W макс\nppm"), ("Sn_ppm", "Sn макс\nppm"),
                  ("Be_ppm", "Be макс\nppm"), ("Li_ppm", "Li макс\nppm")],
        sig=[("Mo", 300), ("W", 100), ("Sn", 100)],
        thresholds="Mo ≥300 ppm · W ≥100 ppm · Sn ≥100 ppm · Li ≥60 ppm · Be — грейзений хаяг",
    ),
    "T3": dict(
        tag="T3_Ag-As-Sb",
        name="Ag–As–Sb (Pb)",
        kind="судлын төрлийн (Ag–As–Sb) хүдэржилт",
        toc_cols=[("Ag_ppm", "Ag макс\nppm"), ("As_ppm", "As макс\nppm"), ("Sb_ppm", "Sb макс\nppm"),
                  ("Pb_ppm", "Pb макс\nppm"), ("S_pct", "S макс\n%")],
        sig=[("Ag", 3)],
        thresholds="Ag ≥3 ppm · As ≥300 ppm · Sb ≥5 ppm · Pb ≥200 ppm",
    ),
    "T4": dict(
        tag="T4_GeoLog",
        name="Геологийн лог",
        kind="шинжилгээгүй / цөөн дээжтэй цооног — литологи · хувирал · бүтэц · керний фото",
        toc_cols=[],
        sig=[],
        thresholds="элементийн bar байхгүй — литологи, хувирал (Log_Detail), фотологийн керний гарц",
    ),
}
ORDER = ["T1", "T2", "T3", "T4"]
HOLE_RE = re.compile(r"^(MTDH-\d{2})_DrillLog_A3L_(T[1-4])(?:_([^_]+(?:-[^_]+)*))?(?:_(\d{4}))?(?:_DRAFT)?\.pdf$", re.I)


def hole_num(h):
    m = re.search(r"(\d+)", h)
    return int(m.group(1)) if m else 0


def fmt(v, nd=0):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    if isinstance(v, (int, float)):
        if nd == 0:
            return f"{v:,.0f}".replace(",", " ")
        return f"{v:.{nd}f}"
    return str(v)


# ----------------------------------------------------------------------------------------------
# 1. Гаралтын PDF-үүдийг олох
# ----------------------------------------------------------------------------------------------
def discover(out_target: Path):
    """→ {Tn: {hole: Path}}, мөн бүх олдсон файлын жагсаалт."""
    found = collections.defaultdict(dict)
    dupes = []
    for p in sorted(out_target.glob("MTDH-*/*.pdf")):
        m = HOLE_RE.match(p.name)
        if not m:
            continue
        hole, tn = m.group(1).upper(), m.group(2).upper()
        prev = found[tn].get(hole)
        if prev is not None:
            # DRAFT биш / шинэ файлыг илүүд үзнэ
            keep_new = ("DRAFT" in prev.name.upper()) or (p.stat().st_mtime > prev.stat().st_mtime and "DRAFT" not in p.name.upper())
            dupes.append((hole, tn, prev.name, p.name, "new" if keep_new else "old"))
            if not keep_new:
                continue
        found[tn][hole] = p
    return found, dupes


def read_map(map_csv: Path):
    """hole_template_map.csv → {(hole, Tn)} хүлээгдэж буй ажлын олонлог. Толгойн нэрийг уян хатан таньна."""
    if not map_csv or not map_csv.exists():
        return None, "map csv олдсонгүй: %s" % map_csv
    with open(map_csv, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return None, "map csv хоосон"
    hdr = [h.strip() for h in rows[0]]
    low = [h.lower() for h in hdr]

    def find(pred):
        for i, h in enumerate(low):
            if pred(h):
                return i
        return None

    i_hole = find(lambda h: "hole" in h or "цооног" in h)
    i_t1 = find(lambda h: "template" in h and "second" not in h and "2" not in h)
    i_t2 = find(lambda h: "second" in h or ("template" in h and "2" in h))
    if i_hole is None or i_t1 is None:
        return None, "map csv толгой танигдсангүй: %s" % hdr
    expected = set()
    for r in rows[1:]:
        if len(r) <= i_hole or not r[i_hole].strip():
            continue
        hole = r[i_hole].strip().upper()
        for idx in (i_t1, i_t2):
            if idx is None or len(r) <= idx:
                continue
            m = re.search(r"T[1-4]", r[idx].upper())
            if m:
                expected.add((hole, m.group(0)))
    return expected, "map csv: %d ажил (%s)" % (len(expected), map_csv.name)


# ----------------------------------------------------------------------------------------------
# 2. Бааз унших
# ----------------------------------------------------------------------------------------------
def read_db(db_path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(db_path, data_only=True, read_only=True)

    def sheet(name):
        if name not in wb.sheetnames:
            return []
        it = wb[name].iter_rows(values_only=True)
        hdr = [str(h).strip() if h is not None else "" for h in next(it)]
        out = []
        for r in it:
            if r is None or not any(v is not None for v in r):
                continue
            out.append(dict(zip(hdr, r)))
        return out

    db = {}
    db["collar"] = {r["Hole_ID"]: r for r in sheet("Collar") if r.get("Hole_ID")}
    sig = collections.defaultdict(list)
    for r in sheet("Sig_Intervals"):
        sig[r["Hole_ID"]].append(r)
    db["sig"] = sig
    amax = collections.defaultdict(dict)
    for r in sheet("Assay"):
        h = r.get("Hole_ID")
        for k, v in r.items():
            if isinstance(v, (int, float)) and (k.endswith("_ppm") or k.endswith("_pct")):
                if v > amax[h].get(k, float("-inf")):
                    amax[h][k] = v
    db["amax"] = amax
    lith = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in sheet("Lithology"):
        try:
            lith[r["Hole_ID"]][str(r["Code"])] += float(r["Length_m"] or 0)
        except (TypeError, ValueError):
            pass
    db["lith"] = lith
    rec = collections.defaultdict(list)
    for r in sheet("PhotoLog"):
        v = r.get("Rec_pct_est")
        if isinstance(v, (int, float)):
            rec[r.get("Hole")].append(v)
    db["rec"] = {h: sum(v) / len(v) for h, v in rec.items() if v}
    caf = collections.defaultdict(lambda: None)
    for r in sheet("Assay_CaF2"):
        v = r.get("CaF2_pct")
        h = r.get("Hole_or_Trench")
        if isinstance(v, (int, float)) and (caf[h] is None or v > caf[h]):
            caf[h] = v
    db["caf2"] = caf
    db["codes"] = sheet("Codes")
    wb.close()
    return db


def best_interval(sig_rows, elements):
    """Загварын элементүүдээс хамгийн чухал огтлол (агуулга × урт хамгийн их) + бусад огтлолын тоо."""
    cand = [r for r in sig_rows if r.get("Element") in [e for e, _ in elements]]
    if not cand:
        return "—"

    def score(r):
        try:
            return float(r["Wtd_Mean_ppm"]) * max(float(r["Length_m"]), 0.25)
        except (TypeError, ValueError):
            return 0.0

    cand.sort(key=score, reverse=True)
    b = cand[0]
    s = "%s %.2f м @ %s ppm (%.2f–%.2f)" % (b["Element"], float(b["Length_m"]), fmt(b["Wtd_Mean_ppm"]), float(b["From_m"]), float(b["To_m"]))
    if len(cand) > 1:
        s += "  +%d огтлол" % (len(cand) - 1)
    return s


# ----------------------------------------------------------------------------------------------
# 3. Нүүр / агуулга / тайлбар хуудас (matplotlib → PDF)
# ----------------------------------------------------------------------------------------------
def make_front_matter(tn, holes, db, page_plan, out_pdf: Path, date_str, total_pages, warnings):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.patches import Rectangle

    plt.rcParams["font.family"] = "DejaVu Sans"
    plt.rcParams["pdf.fonttype"] = 42
    T = TEMPLATES[tn]
    n_holes = len(holes)
    sum_td = sum(float(db["collar"].get(h, {}).get("Max_Depth_m") or 0) for h in holes)
    sum_samples = sum(int(db["collar"].get(h, {}).get("n_Samples") or 0) for h in holes)
    date_h = "%s-%s-%s" % (date_str[:4], date_str[4:6], date_str[6:8])

    toc_pages = []  # (page_index_in_front, rows)
    with PdfPages(str(out_pdf)) as pdf:
        # ---------------- Нүүр ----------------
        fig = plt.figure(figsize=A3L_IN)
        ax = fig.add_axes([0, 0, 1, 1]); ax.set_axis_off(); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.add_patch(Rectangle((0.06, 0.60), 0.88, 0.005, color="#2F3B52"))
        ax.text(0.06, 0.86, "МОДОТ-3 (XV-020181)", fontsize=34, weight="bold", color="#2F3B52")
        ax.text(0.06, 0.79, "Цооногийн багана · A3 хөндлөн · 1 хуудас = %.0f м" % PAGE_M, fontsize=18, color="#444")
        ax.text(0.06, 0.70, "Загвар %s — %s" % (tn, T["name"]), fontsize=26, weight="bold", color="#8B2E2E")
        ax.text(0.06, 0.645, T["kind"], fontsize=15, color="#444")
        y = 0.54
        lines = [
            ("Цооног", "%d ш — %s" % (n_holes, ", ".join(holes))),
            ("Нийт гүн", "%.1f м · дээж %d ш (SGS 2023)" % (sum_td, sum_samples)),
            ("Тодруулгын босго", T["thresholds"]),
            ("Хуудас", "%d (нүүр 1 · агуулга %d · тайлбар 1 · цооногууд %d)" % (total_pages, page_plan["n_toc"], total_pages - 2 - page_plan["n_toc"])),
            ("Огноо", date_h),
            ("Эх дата", "MT_Drilling_Database.xlsx (Collar · Lithology · Log_Detail · Assay · Assay_OverRange · Assay_CaF2 · PhotoLog · Core_Photos) · дахин дээжлэлтийн бүртгэл v2 · керний фото 02_Core_Photos"),
            ("Нэгж файл", "<Цооног>_DrillLog_A3L_%s_2026.pdf — цооног бүрийн сүүлийн хуудас = тухайн баганын бүрэн тайлбар" % T["tag"]),
        ]
        wrap_px = 0.70 * fig.get_figwidth() * fig.dpi  # утгын багана 22%→92% хүртэл эвхэнэ
        for k, v in lines:
            ax.text(0.06, y, k, fontsize=12.5, weight="bold", color="#2F3B52", va="top")
            t = ax.text(0.22, y, v, fontsize=12.5, color="#222", va="top", wrap=True)
            t._get_wrap_line_width = lambda: wrap_px
            y -= 0.055 if len(v) < 110 else (0.085 if len(v) < 220 else 0.115)
        ax.text(0.06, 0.07, "ХЯНАЛТЫН ХУВИЛБАР — визуал QC дуусаагүй. Хуваалцах хавтас руу зөвхөн Жаргалын шийдвэрээр.", fontsize=11, color="#8B2E2E", weight="bold")
        ax.text(0.06, 0.045, PROJECT_MN, fontsize=10, color="#666")
        ax.text(0.94, 0.045, "Нэгтгэсэн: %s" % date_h, fontsize=10, color="#666", ha="right")
        pdf.savefig(fig); plt.close(fig)

        # ---------------- Агуулга ----------------
        base_cols = [("no", "№", 0.030), ("hole", "Цооног", 0.075), ("td", "TD\nм", 0.050), ("box", "Хайр-\nцаг", 0.045),
                     ("ns", "Дээж\nn", 0.045), ("cov", "Хам-\nралт %", 0.050)]
        el_cols = [(k, lbl, 0.058) for k, lbl in T["toc_cols"]]
        if tn == "T4":
            el_cols = [("lith", "Литологи (гол кодууд, м)", 0.240), ("rec", "Керний\nгарц %", 0.055)]
        tail_cols = [("best", "Гол огтлол (Sig_Intervals композит)" if tn != "T4" else "Тэмдэглэл", 0.250 if tn != "T4" else 0.210),
                     ("caf2", "CaF₂ макс\n% (2025)", 0.060), ("start", "Хуудас\n(эхлэх)", 0.055), ("np", "х.", 0.035)]
        cols = base_cols + el_cols + tail_cols
        rows_per_page = 24
        chunks = [holes[i:i + rows_per_page] for i in range(0, len(holes), rows_per_page)] or [[]]
        for ci, chunk in enumerate(chunks):
            fig = plt.figure(figsize=A3L_IN)
            ax = fig.add_axes([0, 0, 1, 1]); ax.set_axis_off(); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.text(0.05, 0.945, "АГУУЛГА — Загвар %s (%s)%s" % (tn, T["name"], "" if len(chunks) == 1 else "  · %d/%d" % (ci + 1, len(chunks))),
                    fontsize=20, weight="bold", color="#2F3B52")
            ax.text(0.05, 0.915, "%s · %d цооног · %s" % (PROJECT_MN, n_holes, date_h), fontsize=11, color="#555")
            x0, x1 = 0.05, 0.95
            widths = [w for _, _, w in cols]
            scale = (x1 - x0) / sum(widths)
            xs = [x0]
            for w in widths:
                xs.append(xs[-1] + w * scale)
            y_top = 0.875
            hh = 0.042
            rh = 0.030
            # толгой
            ax.add_patch(Rectangle((x0, y_top - hh), x1 - x0, hh, color="#E7ECF3"))
            for j, (_, lbl, _) in enumerate(cols):
                ax.text((xs[j] + xs[j + 1]) / 2, y_top - hh / 2, lbl, fontsize=8.5, weight="bold", ha="center", va="center", color="#2F3B52")
            ax.plot([x0, x1], [y_top, y_top], color="#2F3B52", lw=1.2)
            ax.plot([x0, x1], [y_top - hh, y_top - hh], color="#2F3B52", lw=1.2)
            y = y_top - hh
            for i, h in enumerate(chunk):
                c = db["collar"].get(h, {})
                am = db["amax"].get(h, {})
                idx = holes.index(h)
                vals = {
                    "no": str(idx + 1), "hole": h,
                    "td": fmt(float(c.get("Max_Depth_m")), 1) if c.get("Max_Depth_m") is not None else "—",
                    "box": fmt(c.get("n_Core_Boxes")), "ns": fmt(c.get("n_Samples")),
                    "cov": fmt(c.get("Coverage_pct"), 1) if isinstance(c.get("Coverage_pct"), (int, float)) else "—",
                    "best": best_interval(db["sig"].get(h, []), T["sig"]) if tn != "T4" else ("дээж лабд илгээгдээгүй" if h == "MTDH-11" else ("шинжилгээгүй" if not c.get("n_Samples") else "цөөн дээж")),
                    "caf2": fmt(db["caf2"].get(h), 2) if db["caf2"].get(h) is not None else "—",
                    "start": str(page_plan["start"][h]), "np": str(page_plan["pages"][h]),
                }
                for k, _ in T["toc_cols"]:
                    v = am.get(k)
                    vals[k] = fmt(v, 2) if (v is not None and k.endswith("_pct")) else (fmt(v, 1) if (v is not None and v < 100) else fmt(v))
                if tn == "T4":
                    lt = sorted(db["lith"].get(h, {}).items(), key=lambda kv: -kv[1])[:3]
                    vals["lith"] = " · ".join("%s %.1f" % (k, v) for k, v in lt) or "—"
                    vals["rec"] = fmt(db["rec"].get(h), 0) if db["rec"].get(h) is not None else "—"
                if i % 2 == 1:
                    ax.add_patch(Rectangle((x0, y - rh), x1 - x0, rh, color="#F6F8FA", lw=0))
                for j, (k, _, _) in enumerate(cols):
                    ha = "left" if k in ("hole", "best", "lith") else "center"
                    xx = xs[j] + 0.004 if ha == "left" else (xs[j] + xs[j + 1]) / 2
                    ax.text(xx, y - rh / 2, vals.get(k, "—"), fontsize=8.6, ha=ha, va="center", color="#111",
                            weight="bold" if k in ("hole", "start") else "normal")
                y -= rh
                ax.plot([x0, x1], [y, y], color="#C9D1DB", lw=0.5)
            ax.plot([x0, x1], [y, y], color="#2F3B52", lw=1.2)
            for xx in xs:
                ax.plot([xx, xx], [y, y_top], color="#C9D1DB", lw=0.5)
            note = ("Макс утга = SGS 2023 (IC40A/IC40M) цооногийн бүх дээжийн дээд утга; ТДХ давсан утга Assay_OverRange-д тусад нь (Fe/Mn/Zn/Mo %). "
                    "Гол огтлол = Sig_Intervals композит (Mo≥300 · Zn≥1000 · W≥100 · Sn≥100 · Ag≥3 ppm), агуулга×урт хамгийн их нь. "
                    "CaF₂ = SGS 2025 CLA07C (2023 pulp-ийн дахин шинжилгээ; MTDH-18…26-д өгөгдөл байхгүй). Хуудас = нэгтгэсэн PDF доторх дугаар.")
            ax.text(0.05, 0.045, note, fontsize=8.5, color="#555", va="bottom", wrap=True)
            pdf.savefig(fig); plt.close(fig)

        # ---------------- Тайлбар ----------------
        fig = plt.figure(figsize=A3L_IN)
        ax = fig.add_axes([0, 0, 1, 1]); ax.set_axis_off(); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.text(0.05, 0.945, "ТАЙЛБАР — Загвар %s (%s)" % (tn, T["name"]), fontsize=20, weight="bold", color="#2F3B52")
        left = [
            ("Загвар", "%s — %s. %s" % (tn, T["name"], T["kind"])),
            ("Тодруулгын босго", T["thresholds"] + ". (Жаргалын шийдвэр 2026-09-02: 4 загвар, Fe загваргүй; cut-off Pb 200 · Fe 10% · As 300 · Sb 5 · Li 60.)"),
            ("Хуудасны бүтэц", "Цооног бүр: ceil(TD/%.0f)+1 хуудас — гүний хуудсууд + сүүлийн хуудас = тухайн загварын баганын бүрэн тайлбар (монгол). "
                               "Гарчиг, толгой, тайлбар — монгол; литологийн тайлбарын дата — эх хэвээр." % PAGE_M),
            ("Багана (ерөнхий)", "Гүн (м) · дээж № · литологи (код + өнгө) · хувирал/бүтэц (Log_Detail, PhotoLog) · загварын элементийн bar-ууд (лог масштаб, босго тодруулгатай) · "
                                 "дээжлэлтийн хамралт ба цоорхой (Sample_Gaps) + дахин дээжлэлт v2 (1A/1B/2/3, CL) · керний фото (2 хайрцаг/зураг, самбар хэвээр)."),
            ("Эх дата", "MT_Drilling_Database.xlsx (2026-08-31): Collar 26 · Lithology 302 · Log_Detail 302 · Assay 518 (SGS 2023) · Assay_OverRange 81 · Assay_CaF2 341 (SGS 2025) · "
                        "PhotoLog 1 115 · Core_Photos 417 · Sig_Intervals 104. Дахин дээжлэлт: Modot3_Resampling_Register_v2_20260902 (Register_v2_Core.csv)."),
            ("Анхаарах", "① Ca (IC40A) 15%-д таслагдсан — карбонат/скарн/жоншны ялгаа CaF₂/CaCO₃-оос. ② MTDH-23 61.75–61.90 «FL» = ХАГАРАЛ (жоншны судал биш). "
                         "③ MTDH-26 In_Model=N (өндөр/байрлал баталгаажаагүй) — өгөгдөл баазад хэвээр. ④ MTDH-11: 6 дээж лабд илгээгдээгүй; MTDH-15: 10 дээж сертификатгүй (52.2–61.7 шохойн чулуу). "
                         "⑤ PhotoLog-ийн керний гарц (Rec) = фотоны үнэлгээ [Таамаг]. ⑥ ӨТШ 17-1 = MTDH-17 15.70–16.70 м [Тооцоолсон, Ж-4]. ⑦ Гүний хазайлт хэмжигдээгүй — цооног шулуун."),
            ("Хувилбар", "T-загвар v1.1 (12_make_striplog_MT_A3L_target.py) · нэгтгэл 15_merge_striplogs_target.py · %s. Хяналтын хувилбар — визуал QC-ийн дараа засвар орж болно." % date_h),
        ]
        y = 0.89
        for k, v in left:
            ax.text(0.05, y, k, fontsize=11.5, weight="bold", color="#2F3B52", va="top")
            t = ax.text(0.16, y, v, fontsize=10.5, color="#222", va="top", wrap=True)
            t._get_wrap_line_width = lambda: 0.46 * fig.get_figwidth() * fig.dpi  # 46% өргөнд эвхэнэ
            y -= 0.036 + 0.0165 * (len(v) // 110)
        # Литологийн өнгө (баруун)
        ax.text(0.66, 0.89, "Литологийн код · өнгө (Codes sheet)", fontsize=12, weight="bold", color="#2F3B52", va="top")
        yy = 0.855
        for r in db["codes"]:
            col = str(r.get("Striplog_Color") or "#FFFFFF")
            ax.add_patch(Rectangle((0.66, yy - 0.022), 0.035, 0.022, facecolor=col, edgecolor="#333", lw=0.6))
            ax.text(0.703, yy - 0.011, "%s — %s (%s) · %s м" % (r.get("Code"), r.get("Name_MN"), r.get("Name_EN"), fmt(r.get("Total_m"), 1)),
                    fontsize=9.5, va="center", color="#222")
            yy -= 0.031
        if warnings:
            ax.text(0.66, yy - 0.03, "Нэгтгэлийн анхааруулга:", fontsize=10.5, weight="bold", color="#8B2E2E", va="top")
            ax.text(0.66, yy - 0.055, "\n".join("• " + w for w in warnings[:12]), fontsize=8.8, color="#8B2E2E", va="top")
        ax.text(0.05, 0.04, PROJECT_MN + " · " + date_h, fontsize=9.5, color="#666")
        pdf.savefig(fig); plt.close(fig)
    return 2 + len(chunks)


# ----------------------------------------------------------------------------------------------
# 4. Нэгтгэх
# ----------------------------------------------------------------------------------------------
def merge_template(tn, hole_paths, db, merged_dir: Path, date_str, dry_run=False, extra_warn=()):
    from pypdf import PdfReader, PdfWriter
    T = TEMPLATES[tn]
    holes = sorted(hole_paths, key=hole_num)
    warnings = list(extra_warn)
    # хуудасны тоо
    npages = {}
    for h in holes:
        try:
            npages[h] = len(PdfReader(str(hole_paths[h])).pages)
        except Exception as e:  # noqa
            npages[h] = 0
            warnings.append("%s: PDF уншигдсангүй (%s)" % (h, e))
        td = db["collar"].get(h, {}).get("Max_Depth_m")
        if td is not None and npages[h] and npages[h] != math.ceil(float(td) / PAGE_M) + 1:
            warnings.append("%s: хуудас %d ≠ ceil(TD/%.0f)+1 = %d" % (h, npages[h], PAGE_M, math.ceil(float(td) / PAGE_M) + 1))
    n_toc = max(1, math.ceil(len(holes) / 24))
    front = 2 + n_toc
    start = {}
    p = front + 1
    for h in holes:
        start[h] = p
        p += npages[h]
    total = p - 1
    plan = dict(n_toc=n_toc, start=start, pages=npages)
    out_pdf = merged_dir / ("Modot3_StripLog_%s_%s.pdf" % (T["tag"], date_str))
    if dry_run:
        return dict(tn=tn, out=out_pdf, holes=holes, pages=npages, start=start, total=total, size_mb=0, warnings=warnings, ok=None)
    merged_dir.mkdir(parents=True, exist_ok=True)
    tmp = Path(tempfile.mkdtemp(prefix="mt_front_"))
    front_pdf = tmp / ("front_%s.pdf" % tn)
    nfront = make_front_matter(tn, holes, db, plan, front_pdf, date_str, total, warnings)
    assert nfront == front, (nfront, front)
    writer = PdfWriter()
    writer.append(str(front_pdf))
    writer.add_outline_item("Нүүр", 0)
    writer.add_outline_item("Агуулга", 1)
    writer.add_outline_item("Тайлбар", front - 1)
    for h in holes:
        if not npages[h]:
            continue
        writer.append(str(hole_paths[h]))
        td = db["collar"].get(h, {}).get("Max_Depth_m")
        title = "%s · TD %s м · %d х." % (h, fmt(float(td), 1) if td is not None else "?", npages[h])
        writer.add_outline_item(title, start[h] - 1)
    writer.add_metadata({"/Title": "Модот-3 (XV-020181) — Цооногийн багана, загвар %s (%s)" % (tn, T["name"]),
                         "/Subject": "Элемент-зорилтот striplog, A3 хөндлөн, %d цооног" % len(holes),
                         "/Author": "Батбадмаараг ХХК — геологийн баг"})
    with open(out_pdf, "wb") as f:
        writer.write(f)
    got = len(PdfReader(str(out_pdf)).pages)
    ok = (got == total)
    if not ok:
        warnings.append("%s: нэгтгэсэн хуудас %d ≠ төлөвлөсөн %d" % (out_pdf.name, got, total))
    try:
        for q in tmp.iterdir():
            q.unlink()
        tmp.rmdir()
    except OSError:
        pass
    return dict(tn=tn, out=out_pdf, holes=holes, pages=npages, start=start, total=total, got=got,
                size_mb=out_pdf.stat().st_size / 1e6, warnings=warnings, ok=ok)


def write_index(results, merged_dir: Path, date_str, disc_note, map_note, missing, extra, dupes):
    lines = ["# Модот-3 (XV-020181) — T1–T4 нэгтгэсэн цооногийн багана · индекс %s-%s-%s" % (date_str[:4], date_str[4:6], date_str[6:8]), "",
             "Скрипт: `15_merge_striplogs_target.py` · эх: `Out_Target/<HOLE>/<HOLE>_DrillLog_A3L_<TAG>_2026.pdf` · %s · %s" % (disc_note, map_note), ""]
    if missing:
        lines += ["**Map-д байгаа ч PDF олдоогүй:** " + ", ".join("%s %s" % m for m in sorted(missing)), ""]
    if extra:
        lines += ["**PDF байгаа ч map-д байхгүй:** " + ", ".join("%s %s" % m for m in sorted(extra)), ""]
    if dupes:
        lines += ["**Давхардсан файл (сонгосон):** " + "; ".join("%s %s: %s / %s → %s" % d for d in dupes), ""]
    lines += ["| Загвар | Файл | Цооног | Хуудас (төлөв/бодит) | MB | Төлөв |", "|---|---|---:|---|---:|---|"]
    for r in results:
        st = "OK" if r.get("ok") else ("dry-run" if r.get("ok") is None else "ХУУДАС ЗӨРҮҮТЭЙ")
        lines.append("| %s | %s | %d | %d / %s | %.1f | %s |" % (r["tn"], r["out"].name, len(r["holes"]), r["total"], r.get("got", "—"), r["size_mb"], st))
    for r in results:
        lines += ["", "## %s — %s (%d цооног)" % (r["tn"], TEMPLATES[r["tn"]]["name"], len(r["holes"])), "",
                  "| № | Цооног | Хуудас (эхлэх) | х. | Эх файл |", "|---:|---|---:|---:|---|"]
        for i, h in enumerate(r["holes"]):
            lines.append("| %d | %s | %d | %d | %s |" % (i + 1, h, r["start"][h], r["pages"][h], "%s_DrillLog_A3L_%s_2026.pdf" % (h, TEMPLATES[r["tn"]]["tag"])))
        if r["warnings"]:
            lines += ["", "Анхааруулга:"] + ["- " + w for w in r["warnings"]]
    idx = merged_dir / ("Merge_index_%s.md" % date_str)
    merged_dir.mkdir(parents=True, exist_ok=True)
    idx.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return idx


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out-target", default=DEFAULT_OUT_TARGET)
    ap.add_argument("--db", default=None, help="MT_Drilling_Database.xlsx (default: G: → D: Prep)")
    ap.add_argument("--map", default=DEFAULT_MAP, help="hole_template_map.csv (тулгалтад)")
    ap.add_argument("--merged-dir", default=None, help="default <out-target>/_Merged")
    ap.add_argument("--templates", nargs="*", default=ORDER, choices=ORDER)
    ap.add_argument("--date", default=dt.date.today().strftime("%Y%m%d"))
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args(argv)

    out_target = Path(a.out_target)
    if not out_target.exists():
        sys.exit("Out_Target олдсонгүй: %s" % out_target)
    merged_dir = Path(a.merged_dir) if a.merged_dir else out_target / "_Merged"
    db_path = None
    for cand in ([a.db] if a.db else []) + DEFAULT_DBS:
        if cand and Path(cand).exists():
            db_path = Path(cand)
            break
    if db_path is None:
        sys.exit("Бааз олдсонгүй (--db): %s" % ([a.db] + DEFAULT_DBS))
    print("Out_Target :", out_target)
    print("Бааз       :", db_path)
    found, dupes = discover(out_target)
    n_found = sum(len(v) for v in found.values())
    disc_note = "олдсон PDF %d (T1 %d · T2 %d · T3 %d · T4 %d)" % (n_found, len(found["T1"]), len(found["T2"]), len(found["T3"]), len(found["T4"]))
    print(disc_note)
    expected, map_note = read_map(Path(a.map) if a.map else None)
    print(map_note)
    missing, extra = set(), set()
    if expected is not None:
        have = {(h, tn) for tn, d in found.items() for h in d}
        missing = expected - have
        extra = have - expected
        if missing:
            print("  ! map-д байгаа ч PDF АЛГА:", sorted(missing))
        if extra:
            print("  ! PDF байгаа ч map-д БАЙХГҮЙ:", sorted(extra))
    db = read_db(db_path)
    results = []
    for tn in a.templates:
        if not found.get(tn):
            print("  %s: PDF байхгүй — алгасав" % tn)
            continue
        warn = ["%s %s: map-д байхгүй" % (h, t) for (h, t) in sorted(extra) if t == tn]
        warn += ["%s %s: PDF олдоогүй (map-д бий)" % (h, t) for (h, t) in sorted(missing) if t == tn]
        r = merge_template(tn, found[tn], db, merged_dir, a.date, dry_run=a.dry_run, extra_warn=warn)
        results.append(r)
        print("  %s: %d цооног → %s · хуудас %d%s · %.1f MB %s" % (tn, len(r["holes"]), r["out"].name, r["total"],
              "" if a.dry_run else " (бодит %s)" % r.get("got"), r["size_mb"], "OK" if r["ok"] else ("(dry-run)" if r["ok"] is None else "!! ЗӨРҮҮ")))
        for w in r["warnings"]:
            print("      -", w)
    idx = write_index(results, merged_dir, a.date, disc_note, map_note, missing, extra, dupes)
    print("Индекс     :", idx)
    bad = [r for r in results if r["ok"] is False]
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())

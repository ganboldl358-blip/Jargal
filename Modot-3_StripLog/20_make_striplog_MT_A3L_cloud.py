#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
20_make_striplog_MT_A3L_cloud.py — Модот-3 (XV-020181) элемент-зорилтот цооногийн багана (T1–T4).
CLOUD хувилбар: зөвхөн Google Drive дээрх өгөгдлөөр (MT_Drilling_Database.xlsx + керний фотоны PDF)
ажилладаг — D: диск шаардахгүй. D: дээрх батлагдсан `12_make_striplog_MT_A3L_target.py`-ийн
ОРЛУУЛАГЧ БИШ, харин компьютер унтарсан үед cloud-оос гаргасан зэрэгцээ хувилбар.

Формат (Жаргалын батлагдсан загвар, 2026-08…09):
    A3 ХӨНДЛӨН · босоо гүн · 1:100 (нэг хуудас = 23.0 м) · хуудас = ceil(TD/23)+1 (сүүлийнх = ТАЙЛБАР)
    Багана: Гүн · Дээж № · Литологи · Хувирал/бүтэц · Керний гарц · <загварын элементийн bar-ууд>
            · Дээжлэлтийн цоорхой + дахин дээжлэлт · Тайлбар · Керний фото
    Гарчиг, толгой, тайлбар МОНГОЛ; литологийн тайлбарын дата эх хэвээр.

Ажиллуулах:
    python3 20_make_striplog_MT_A3L_cloud.py MTDH-17 T2 [--db ...] [--photos ...] [--out ...] [--qc ...]
    python3 20_make_striplog_MT_A3L_cloud.py --all --map hole_template_map_cloud.csv
"""
import argparse
import collections
import csv
import gc
import math
import os
import re
import sys
import textwrap
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import Rectangle, Polygon
import matplotlib.image as mpimg
import numpy as np

plt.rcParams["font.family"] = "DejaVu Sans"
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["pdf.compression"] = 9

# ---------------------------------------------------------------------------------------------
# Хуудасны геометр — A3 хөндлөн, 1:100
# ---------------------------------------------------------------------------------------------
A3L = (16.535, 11.693)          # инч
PAGE_M = 23.0                   # нэг хуудсанд ногдох гүн (м)
PLOT_Y0, PLOT_H = 0.0725, 0.7744  # 0.7744 × 11.693" = 9.055" = 23.0 см ⇒ ЯГ 1:100
PLOT_Y1 = PLOT_Y0 + PLOT_H
HDR_Y1 = 0.940                  # баганын толгойн дээд ирмэг
X0, X1 = 0.026, 0.988

NAVY, RED, GREY = "#2F3B52", "#8B2E2E", "#666666"
PROJECT_MN = "Модот-3 (XV-020181) · Дорноговь аймаг, Айраг сум · «Бат Бадмаараг» ХХК"

# ---------------------------------------------------------------------------------------------
# Загварууд
# ---------------------------------------------------------------------------------------------
EL = {  # түлхүүр: (шошго, нэгж, доод, дээд, босго, өнгө, тодруулгын өнгө, аравтын орон)
    "Zn_ppm": ("Zn", "ppm", 10, 20000, 1000, "#2E6FB7", "#0B3D7A", 0),
    "Ag_ppm": ("Ag", "ppm", 0.1, 30, 3, "#8A8F98", "#3D4149", 1),
    "Pb_ppm": ("Pb", "ppm", 5, 2000, 200, "#5B6B7C", "#25313D", 0),
    "Mn_ppm": ("Mn", "ppm", 50, 60000, 10000, "#9B6B9E", "#5C2F60", 0),
    "Cd_ppm": ("Cd", "ppm", 0.05, 50, 5, "#A08B4F", "#5E4B12", 1),
    "In_ppm": ("In", "ppm", 0.02, 35, 5, "#7E9E5B", "#3D5A1C", 1),
    "Mo_ppm": ("Mo", "ppm", 1, 15000, 300, "#D2691E", "#8B3A00", 0),
    "W_ppm":  ("W", "ppm", 0.5, 1200, 100, "#2E8B87", "#0E4B48", 0),
    "Sn_ppm": ("Sn", "ppm", 1, 350, 100, "#8B5A2B", "#4A2C10", 0),
    "Be_ppm": ("Be", "ppm", 0.5, 120, 20, "#C77BA0", "#7A2F52", 1),
    "Li_ppm": ("Li", "ppm", 1, 200, 60, "#7FA650", "#3F5C1E", 0),
    "As_ppm": ("As", "ppm", 1, 3500, 300, "#4F8A4F", "#1E4B1E", 0),
    "Sb_ppm": ("Sb", "ppm", 0.1, 30, 5, "#7B5EA7", "#3B2668", 1),
    "S_pct":  ("S", "%", 0.01, 5, 0.5, "#C9A227", "#7A5F00", 2),
    "Cu_ppm": ("Cu", "ppm", 1, 800, 200, "#B06A3B", "#6A3410", 0),
    "Bi_ppm": ("Bi", "ppm", 0.1, 1500, 100, "#6D6875", "#332F3A", 1),
}

TEMPLATES = {
    "T1": dict(tag="T1_Zn-Ag", name="Zn–Ag (Pb–Mn–Cd–In)",
               kind="скарн / карбонат-орлуулалтын (CRD) төрлийн хүдэржилт",
               els=["Zn_ppm", "Ag_ppm", "Pb_ppm", "Mn_ppm", "Cd_ppm", "In_ppm"],
               note="Zn ≥1000 ppm · Ag ≥3 ppm · Pb ≥200 ppm · Cd/Zn <10 ⇒ өндөр T скарн · In нь Zn-ийг дагана"),
    "T2": dict(tag="T2_Mo-W-Sn", name="Mo–W–Sn (Be–Li)",
               kind="грейзен / боржинтой холбоотой хүдэржилт",
               els=["Mo_ppm", "W_ppm", "Sn_ppm", "Be_ppm", "Li_ppm", "Bi_ppm"],
               note="Mo ≥300 ppm · W ≥100 ppm · Sn ≥100 ppm · Li ≥60 ppm · Be, Bi = грейзений хаяг"),
    "T3": dict(tag="T3_Ag-As-Sb", name="Ag–As–Sb (Pb)",
               kind="судлын төрлийн (Ag–As–Sb) хүдэржилт",
               els=["Ag_ppm", "As_ppm", "Sb_ppm", "Pb_ppm", "S_pct"],
               note="Ag ≥3 ppm · As ≥300 ppm · Sb ≥5 ppm · Pb ≥200 ppm"),
    "T4": dict(tag="T4_GeoLog", name="Геологийн лог",
               kind="шинжилгээгүй / цөөн дээжтэй цооног — литологи · хувирал · бүтэц · керний фото",
               els=[],
               note="элементийн bar байхгүй — литологи, хээрийн бичиглэл (Log_Detail), фотологийн керний гарц"),
}

ALT_COLS = [("Alt_Silica", "Цх", "#C0392B"), ("Alt_Carbonate", "Кб", "#2E86C1"),
            ("Alt_Clay", "Шв", "#B7950B"), ("Alt_Sericite", "Ср", "#7D3C98"),
            ("Alt_Chlorite", "Хл", "#1E8449"), ("Alt_Epidote", "Эп", "#58D68D"),
            ("Alt_Kspar", "Кш", "#E67E22"), ("Alt_Magnetite", "Мг", "#34495E"),
            ("Alt_Fracturing", "Хг", "#7F8C8D")]


# ---------------------------------------------------------------------------------------------
# Бааз
# ---------------------------------------------------------------------------------------------
def load_db(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    def sh(name):
        if name not in wb.sheetnames:
            return []
        it = wb[name].iter_rows(values_only=True)
        hdr = [str(h).strip() if h is not None else "" for h in next(it)]
        out = []
        for r in it:
            if r is None or not any(v is not None for v in r):
                continue
            out.append(dict(zip(hdr, r)))
        return out

    db = {n: sh(n) for n in ["Collar", "Lithology", "Log_Detail", "Assay", "Assay_OverRange",
                             "Assay_CaF2", "PhotoLog", "Core_Photos", "Sample_Gaps",
                             "Sig_Intervals", "Codes", "RS_Core", "Keyword_Index"]}
    wb.close()
    db["collar"] = {r["Hole_ID"]: r for r in db["Collar"] if r.get("Hole_ID")}
    db["codes"] = {str(r["Code"]): r for r in db["Codes"]}
    # ТДХ давсан (ore-grade) утгыг нэгтгэх
    sid = {r.get("Sample_ID"): r for r in db["Assay"] if r.get("Sample_ID")}
    for r in db["Assay_OverRange"]:
        a = sid.get(r.get("Sample_ID"))
        if not a:
            continue
        for k, tgt, mult in (("Fe_pct", "Fe_pct", 1), ("Mn_pct", "Mn_ppm", 1e4),
                             ("Zn_pct", "Zn_ppm", 1e4), ("Mo_pct", "Mo_ppm", 1e4)):
            v = r.get(k)
            if isinstance(v, (int, float)):
                cur = a.get(tgt)
                if not isinstance(cur, (int, float)) or v * mult > cur:
                    a[tgt] = v * mult
                    a.setdefault("_over", set()).add(tgt)
    return db


def num(v):
    """Тоо болгож хөрвүүлнэ.

    Log_Detail-ийн хувирлын эрчим ('1','2','3') болон Vein_Thick_cm зэрэг зарим
    талбар Excel-д ТЕКСТ хэлбэрээр хадгалагдсан тул мөрийг ч задална (2026-09-03
    QC-д илрэв — өмнө нь 197 хувирлын утга чимээгүй хаягдаж, «Хувирал · бүтэц»
    багана хоосон гарч байсан). Илрүүлэлтийн доод хязгаараас доош ('<0.01')
    утгыг тоо гэж авахгүй — бодит утга мэдэгдэхгүй тул зурахгүй.
    """
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v if v == v else None
    if isinstance(v, str):
        s = v.strip().replace(",", ".").replace("%", "").replace("\u00a0", "")
        if not s or s[0] in "<>":
            return None
        try:
            f = float(s)
        except ValueError:
            return None
        return f if f == f else None
    return None


def rows_for(db, sheet, hole, key="Hole_ID"):
    return [r for r in db[sheet] if r.get(key) == hole]


def hole_data(db, hole):
    d = {"hole": hole, "collar": db["collar"].get(hole, {})}
    d["td"] = float(d["collar"].get("Max_Depth_m") or 0)
    d["lith"] = sorted([r for r in rows_for(db, "Lithology", hole)], key=lambda r: float(r["From_m"] or 0))
    d["log"] = sorted([r for r in rows_for(db, "Log_Detail", hole)], key=lambda r: float(r["From_m"] or 0))
    d["assay"] = sorted([r for r in rows_for(db, "Assay", hole)], key=lambda r: float(r["From_m"] or 0))
    d["gaps"] = sorted([r for r in rows_for(db, "Sample_Gaps", hole)], key=lambda r: float(r["From_m"] or 0))
    d["sig"] = rows_for(db, "Sig_Intervals", hole)
    d["photolog"] = sorted([r for r in db["PhotoLog"] if r.get("Hole") == hole],
                           key=lambda r: float(r["From_m"] or 0))
    d["boxes"] = sorted([r for r in rows_for(db, "Core_Photos", hole)], key=lambda r: float(r["From_m"] or 0))
    d["caf2"] = sorted([r for r in db["Assay_CaF2"] if r.get("Hole_or_Trench") == hole],
                       key=lambda r: float(r["From_m"] or 0))
    d["rs"] = sorted([r for r in db["RS_Core"] if r.get("Цооног") == hole],
                     key=lambda r: float(r["From (м)"] or 0))
    d["kw"] = rows_for(db, "Keyword_Index", hole)
    return d


# --- дахин дээжлэлт v2 (register v2-ийн дүрмээр тооцсон) ---------------------------------------
def resample_priority(d, f, t):
    """Цоорхойн интервалд v2 эрэмбэ (1A/1B/2/3) — Modot3_Resampling_Register_v2 §2-ын дүрэм."""
    hi = [r for r in d["caf2"] if num(r.get("CaF2_pct")) and num(r["CaF2_pct"]) >= 5]
    halo = [r for r in d["caf2"] if num(r.get("CaF2_pct")) and num(r["CaF2_pct"]) >= 1]
    mid = (f + t) / 2.0

    def near(rows, dist):
        for r in rows:
            a, b = float(r["From_m"]), float(r["To_m"])
            if a - dist <= mid <= b + dist:
                return True
        return False

    codes = set()
    for r in d["lith"]:
        a, b = float(r["From_m"] or 0), float(r["To_m"] or 0)
        if b > f and a < t:
            codes.add(str(r.get("Code")))
    carb = bool(codes & {"LS", "SK", "AL", "BR"})
    gran = bool(codes & {"GR", "LGR", "DR", "LDR"})
    fl = any(str(r.get("Code")) == "FL" for r in d["lith"]
             if float(r["From_m"] or 0) < t and float(r["To_m"] or 0) > f)
    if (hi and near(hi, 3.0)) or fl:
        return "1A"
    if halo and (near(halo, 5.0) or carb):
        return "1B"
    if not d["caf2"] and carb:
        return "2"
    if gran:
        return "3"
    return "2" if carb else "3"


RS_COL = {"1A": "#B22222", "1B": "#E07B39", "2": "#D4B106", "3": "#9E9E9E"}


# ---------------------------------------------------------------------------------------------
# Керний фото
# ---------------------------------------------------------------------------------------------
def load_photos(photo_dir, hole, boxes):
    """photo_dir/<HOLE>/*.jpg-ийг гүнээр эрэмбэлж, Core_Photos-той тааруулна."""
    p = Path(photo_dir) / hole
    if not p.exists():
        return []
    files = sorted([q for q in p.iterdir() if q.suffix.lower() in (".jpg", ".jpeg", ".png")],
                   key=lambda q: int(re.search(r"_(\d+)", q.stem).group(1)) if re.search(r"_(\d+)", q.stem) else 0)
    if not files:
        return []
    # Core_Photos-оос Photo_File бүрийн гүний хүрээ (2 хайрцаг = 1 зураг)
    grp = collections.OrderedDict()
    for b in boxes:
        key = str(b.get("Photo_File") or "").strip()
        f, t = num(b.get("From_m")), num(b.get("To_m"))
        if f is None or t is None:
            continue
        if key not in grp:
            grp[key] = [f, t, []]
        grp[key][0] = min(grp[key][0], f)
        grp[key][1] = max(grp[key][1], t)
        grp[key][2].append(b.get("Box"))
    spans = list(grp.values())
    if not spans:  # Core_Photos байхгүй бол зургуудыг TD-д жигд хуваарилна
        return [(i, None, None, q) for i, q in enumerate(files)]
    out = []
    n = min(len(spans), len(files))
    for i in range(n):
        f, t, bx = spans[i]
        out.append((i, f, t, files[i], bx))
    for i in range(n, len(files)):       # илүү зураг үлдвэл сүүлд нь
        out.append((i, None, None, files[i], []))
    return out


# ---------------------------------------------------------------------------------------------
# Зурах туслахууд
# ---------------------------------------------------------------------------------------------
def wrapt(txt, width_frac, fontsize, fig_w=A3L[0]):
    """Урьдчилан мөр таслах. matplotlib-ийн wrap=True нь зурах болгонд дахин
    тооцдог тул удаан (O(n^2)); энэ нь нэг л удаа textwrap-аар таслана."""
    pt = width_frac * fig_w * 72.0
    # 0.60 = DejaVu Sans-ийн кирилл текстийн дундаж тэмдэгтийн өргөн (em-ийн
    # хувиар, TextPath-аар хэмжсэн 0.55–0.58 + нөөц).
    n = max(8, int(pt / (fontsize * 0.60)))
    return textwrap.fill(str(txt), n)


def col_axes(fig, x, w, d0, d1, top=PLOT_Y1, bottom=PLOT_Y0):
    ax = fig.add_axes([x, bottom, w, top - bottom])
    ax.set_xlim(0, 1)
    ax.set_ylim(d1, d0)          # гүн доошоо
    ax.set_xticks([])
    ax.set_yticks([])
    for s in ax.spines.values():
        s.set_color("#B9C2CC")
        s.set_linewidth(0.6)
    return ax


def hdr(fig, x, w, text, sub=None, color=NAVY, rot=0):
    ax = fig.add_axes([x, PLOT_Y1, w, HDR_Y1 - PLOT_Y1])
    ax.set_axis_off()
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.add_patch(Rectangle((0, 0), 1, 1, facecolor="#EEF2F7", edgecolor="#B9C2CC", lw=0.6))
    ax.text(0.5, 0.62 if sub else 0.5, text, ha="center", va="center", fontsize=8.2,
            weight="bold", color=color, rotation=rot)
    if sub:
        ax.text(0.5, 0.22, sub, ha="center", va="center", fontsize=6.2, color="#555")
    return ax


def clip_iter(rows, d0, d1, fk="From_m", tk="To_m"):
    for r in rows:
        f, t = num(r.get(fk)), num(r.get(tk))
        if f is None or t is None or t <= d0 or f >= d1:
            continue
        yield r, max(f, d0), min(t, d1)


def logpos(v, lo, hi):
    v = max(min(float(v), hi), lo)
    return (math.log10(v) - math.log10(lo)) / (math.log10(hi) - math.log10(lo))


# ---------------------------------------------------------------------------------------------
# Нэг гүний хуудас
# ---------------------------------------------------------------------------------------------
def draw_depth_page(fig, d, tpl, d0, d1, page_i, n_pages, photos):
    T = TEMPLATES[tpl]
    hole, c = d["hole"], d["collar"]
    els = T["els"]

    # ---- гарчгийн зурвас
    ax = fig.add_axes([0, HDR_Y1, 1, 1 - HDR_Y1]); ax.set_axis_off()
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.text(X0, 0.62, "%s" % hole, fontsize=18, weight="bold", color=NAVY, va="center")
    ax.text(X0 + 0.093, 0.62, "цооногийн багана", fontsize=10.5, color="#7A8593", va="center")
    ax.text(X0 + 0.205, 0.62, "Загвар %s — %s" % (tpl, T["name"]), fontsize=12.5, weight="bold",
            color=RED, va="center")
    ax.text(X0, 0.22, "%s · TD %.1f м · азимут %s° / налуу %s° · RL %.1f м · %s / %s · бичиглэгч %s"
            % (PROJECT_MN, d["td"], c.get("Azimuth"), c.get("Dip"), float(c.get("RL_DEM_m") or 0),
               ("%.1f" % float(c.get("East"))) if num(c.get("East")) else "?",
               ("%.1f" % float(c.get("North"))) if num(c.get("North")) else "?",
               c.get("Logger") or "—"),
            fontsize=7.4, color="#555", va="center")
    ax.text(0.988, 0.62, "%.1f – %.1f м" % (d0, d1), fontsize=13, weight="bold", color=NAVY,
            ha="right", va="center")
    ax.text(0.988, 0.22, "хуудас %d / %d · A3 хөндлөн · 1:100" % (page_i, n_pages), fontsize=7.6,
            color="#555", ha="right", va="center")

    # ---- баганын өргөн
    cols = [("depth", 0.030), ("samp", 0.052), ("lith", 0.050), ("alt", 0.082), ("rec", 0.036)]
    cols += [(e, 0.050) for e in els]
    cols += [("rs", 0.052), ("desc", 0.150 if els else 0.240), ("photo", 0.215 if els else 0.300)]
    tot = sum(w for _, w in cols)
    scale = (X1 - X0) / tot
    xs, x = {}, X0
    for k, w in cols:
        xs[k] = (x, w * scale)
        x += w * scale

    # ---- 1. Гүн
    x, w = xs["depth"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Гүн", "м")
    step = 1.0
    v = math.floor(d0)
    while v <= d1 + 1e-9:
        if v >= d0:
            major = abs(v - round(v / 5.0) * 5.0) < 1e-6
            ax.plot([0.55 if not major else 0.25, 1.0], [v, v], color="#8A97A6" if major else "#C9D1DB",
                    lw=0.9 if major else 0.4)
            if major:
                ax.text(0.20, min(max(v, d0 + (d1 - d0) * 0.008), d1 - (d1 - d0) * 0.004),
                        "%g" % v, fontsize=7.4, ha="right", va="center", color=NAVY, weight="bold")
        v += step

    # ---- 2. Дээж №
    x, w = xs["samp"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Дээж", "№ / интервал")
    for r, a, b in clip_iter(d["assay"], d0, d1):
        ax.add_patch(Rectangle((0.06, a), 0.88, b - a, facecolor="#F2F6FA",
                               edgecolor="#7F8C99", lw=0.5))
        if b - a > 0.45:
            ax.text(0.5, (a + b) / 2, str(r.get("Sample_No") or ""), fontsize=6.4, ha="center",
                    va="center", color="#243447")

    # ---- 3. Литологи
    x, w = xs["lith"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Литологи", "код")
    for r, a, b in clip_iter(d["lith"], d0, d1):
        code = str(r.get("Code"))
        col = str((d["codes_map"].get(code) or {}).get("Striplog_Color") or "#FFFFFF")
        ax.add_patch(Rectangle((0, a), 1, b - a, facecolor=col, edgecolor="#4A5568", lw=0.5))
        if b - a > 0.9:
            ax.text(0.5, (a + b) / 2, code, fontsize=6.8, ha="center", va="center",
                    color="#1B2530", weight="bold",
                    bbox=dict(boxstyle="round,pad=0.12", fc="white", ec="none", alpha=0.65))

    # ---- 4. Хувирал / бүтэц
    x, w = xs["alt"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Хувирал · бүтэц", "эрчим 1–3 · судал · сульфид %")
    nA = len(ALT_COLS)
    for i, (_, lbl, colr) in enumerate(ALT_COLS):
        ax.plot([(i + 0.5) / (nA + 2), (i + 0.5) / (nA + 2)], [d0, d1], color="#EDF1F5", lw=0.4)
    for r, a, b in clip_iter(d["log"], d0, d1):
        for i, (key, lbl, colr) in enumerate(ALT_COLS):
            v = num(r.get(key))
            if not v:
                continue
            xc = (i + 0.5) / (nA + 2)
            # эрчим 1–3 → зурвасын өргөн 55–100% (нарийн багана тул хамгийн
            # сул эрчим ч уншигдахуйц байх ёстой); өнгөний тодрол мөн эрчмээр
            hw = 0.5 / (nA + 2) * (0.55 + 0.45 * min(v, 3) / 3.0)
            ax.add_patch(Rectangle((xc - hw, a), 2 * hw, b - a, facecolor=colr,
                                   edgecolor="none", alpha=0.30 + 0.22 * min(v, 3)))
        # судал
        if r.get("Vein_Qtz") or r.get("Vein_Carbonate") or r.get("Vein_Sulphide") or r.get("Vein_Other"):
            xc = (nA + 0.5) / (nA + 2)
            ax.plot([xc - 0.02, xc + 0.02], [(a + b) / 2, (a + b) / 2], color="#111", lw=1.1)
            ax.plot([xc, xc], [a, b], color="#111", lw=0.8)
        # сульфид %
        sp = num(r.get("Sulphide_pct"))
        if sp:
            xc = (nA + 1.5) / (nA + 2)
            hw = 0.42 / (nA + 2) * min(sp, 10) / 10.0
            ax.add_patch(Rectangle((xc - hw, a), 2 * hw, b - a, facecolor="#C9A227",
                                   edgecolor="#7A5F00", lw=0.3))

    # ---- 5. Керний гарц
    x, w = xs["rec"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Керний", "гарц %")
    ax.plot([0.5, 0.5], [d0, d1], color="#DCE3EA", lw=0.5)
    for r, a, b in clip_iter(d["photolog"], d0, d1):
        v = num(r.get("Rec_pct_est"))
        if v is None:
            continue
        fr = max(min(v / 100.0, 1.0), 0.0)
        ax.add_patch(Rectangle((0, a), fr, b - a,
                               facecolor="#9CCBA1" if v >= 50 else ("#E8C07D" if v >= 15 else "#DE8A7A"),
                               edgecolor="#7F8C99", lw=0.25))

    # ---- 6. Элементийн bar-ууд
    for e in els:
        lbl, unit, lo, hi, thr, colr, hcol, ndp = EL[e]
        x, w = xs[e]
        ax = col_axes(fig, x, w, d0, d1)
        hdr(fig, x, w, lbl, "%s · лог" % unit)
        tp = logpos(thr, lo, hi)
        ax.plot([tp, tp], [d0, d1], color=RED, lw=0.7, ls=(0, (3, 2)))
        for g in (0.25, 0.5, 0.75):
            ax.plot([g, g], [d0, d1], color="#EDF1F5", lw=0.4)
        for r, a, b in clip_iter(d["assay"], d0, d1):
            v = num(r.get(e))
            if v is None or v <= 0:
                continue
            fr = logpos(v, lo, hi)
            over = e in (r.get("_over") or set())
            ax.add_patch(Rectangle((0, a), fr, b - a,
                                   facecolor=hcol if v >= thr else colr,
                                   edgecolor="#22303F" if over else "none",
                                   lw=0.6 if over else 0, alpha=0.95))
            if v >= thr and (b - a) > 0.55:
                if fr < 0.62:
                    ax.text(fr + 0.04, (a + b) / 2, ("%." + str(ndp) + "f") % v, fontsize=5.8,
                            va="center", ha="left", color="#111", weight="bold")
                else:
                    ax.text(fr - 0.035, (a + b) / 2, ("%." + str(ndp) + "f") % v, fontsize=5.8,
                            va="center", ha="right", color="white", weight="bold")

    # ---- 7. Дээжлэлт / цоорхой / дахин дээжлэлт
    x, w = xs["rs"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Цоорхой", "дахин дээжлэлт v2")
    for r, a, b in clip_iter(d["gaps"], d0, d1):
        pr = resample_priority(d, float(r["From_m"]), float(r["To_m"]))
        ax.add_patch(Rectangle((0.04, a), 0.44, b - a, facecolor="#F6D6D0", edgecolor="#C0392B",
                               lw=0.4, hatch="///"))
        ax.add_patch(Rectangle((0.54, a), 0.42, b - a, facecolor=RS_COL[pr], edgecolor="none", alpha=0.85))
        if b - a > 1.2:
            ax.text(0.75, (a + b) / 2, pr, fontsize=6.2, ha="center", va="center",
                    color="white", weight="bold")

    # ---- 8. Тайлбар (литологийн бичиглэл, leader шугамтай)
    x, w = xs["desc"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Тайлбар", "хээрийн бичиглэл (эх хэвээр)")
    span = d1 - d0
    placed = []
    for r, a, b in clip_iter(d["lith"], d0, d1):
        txt = (r.get("Description") or r.get("Lithology_MN") or "").strip()
        if not txt:
            continue
        maxch = int(38 * ((b - a) / span * 26 + 1.2))
        txt = txt if len(txt) <= maxch else txt[:maxch - 1] + "…"
        # Давхцлаас зайлсхийх: доошоо л шилжинэ (while ашиглавал хөвөгч таслалын
        # алдаанаас болж мөнхийн давталтад орно — 2026-09-03-т засав).
        y = (a + b) / 2
        if placed:
            y = max(y, placed[-1] + span * 0.030)
        y = min(y, d1 - span * 0.012)
        placed.append(y)
        ax.plot([0.0, 0.03], [(a + b) / 2, y], color="#9AA7B4", lw=0.4)
        ax.text(0.035, y, wrapt("%.1f–%.1f  %s" % (float(r["From_m"]), float(r["To_m"]), txt),
                                w * 0.90, 5.6), fontsize=5.6, va="center", ha="left",
                color="#1B2530", linespacing=1.15)
        ax.plot([0, 1], [a, a], color="#E3E8ED", lw=0.3)

    # ---- 9. Керний фото
    x, w = xs["photo"]
    ax = col_axes(fig, x, w, d0, d1)
    hdr(fig, x, w, "Керний фото", "хайрцгийн бүтэн зураг (зүсэлтгүй)")
    ax.set_facecolor("#FAFBFC")
    if not photos:
        ax.text(0.5, (d0 + d1) / 2, "керний фото энэ хувилбарт\nтатагдаагүй\n(02_Core_Photos-д бий)",
                fontsize=8, color="#9AA7B4", ha="center", va="center")
    col_in = w * A3L[0]
    px_target = int(col_in * 210)          # ~210 dpi — PDF-ийн хэмжээг барихын тулд
    slots = []                              # (y_top, y_bot) давхцлаас сэргийлэх
    for item in photos:
        pf, pt, path = item[1], item[2], item[3]
        bx = item[4] if len(item) > 4 else []
        if pf is None or pt is None:
            continue
        mid = (pf + pt) / 2.0
        if not (d0 <= mid < d1):            # зөвхөн дунджаараа энэ хуудсанд харьяалагдах зураг
            continue
        try:
            im = mpimg.imread(str(path))
        except Exception:
            continue
        ih, iw = im.shape[0], im.shape[1]
        if iw > px_target:                  # PDF-д шигтгэхийн өмнө жижигрүүлнэ
            from PIL import Image as _I
            im = np.asarray(_I.fromarray(im).resize((px_target, max(1, int(ih * px_target / iw))), _I.LANCZOS))
            ih, iw = im.shape[0], im.shape[1]
        wid_in = col_in * 0.93
        hgt_m = wid_in / (iw / ih) / (PLOT_H * A3L[1]) * PAGE_M
        y_top = mid - hgt_m / 2
        for (sa, sb) in slots:              # давхцвал доош түлхэнэ
            if y_top < sb and y_top + hgt_m > sa:
                y_top = sb + span * 0.010
        y_top = max(min(y_top, d1 - hgt_m - span * 0.004), d0 + span * 0.016)
        slots.append((y_top, y_top + hgt_m))
        xw = wid_in / col_in
        ax.imshow(im, extent=(0.045, 0.045 + xw, y_top + hgt_m, y_top),
                  aspect="auto", interpolation="none", zorder=2, clip_on=True)
        ax.add_patch(Rectangle((0.045, y_top), xw, hgt_m, fill=False, edgecolor="#6B7684",
                               lw=0.6, zorder=3))
        ax.text(0.045, y_top - span * 0.005, "%.2f–%.2f м%s" % (pf, pt, (" · хайрцаг %s" %
                "-".join(str(q) for q in bx)) if bx else ""), fontsize=5.6, color="#333",
                va="bottom", ha="left", zorder=3)
        ax.plot([0.008, 0.008], [pf, pt], color="#4A7A9B", lw=1.6, zorder=3,
                solid_capstyle="butt")      # зургийн жинхэнэ гүний хамрах хүрээ
        ax.plot([0.008, 0.045], [mid, y_top + hgt_m / 2], color="#9AA7B4", lw=0.4, zorder=1)

    # ---- хөл
    ax = fig.add_axes([0, 0, 1, PLOT_Y0]); ax.set_axis_off()
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.plot([X0, X1], [0.82, 0.82], color="#C9D1DB", lw=0.6)
    ax.text(X0, 0.52, "Эх дата: MT_Drilling_Database.xlsx (2026-08-31) · керний фото 02_Core_Photos · "
                      "сүүлийн хуудас = баганын бүрэн тайлбар", fontsize=6.6, color="#666", va="center")
    ax.text(X1, 0.52, "ХЯНАЛТЫН ХУВИЛБАР — визуал QC хийгдээгүй", fontsize=6.8, color=RED,
            weight="bold", va="center", ha="right")
    # масштабын шугам (1 см = 1 м)
    ax.text(X0, 0.16, "Босоо масштаб 1:100 — 1 хуудас = %.0f м" % PAGE_M, fontsize=6.4,
            color="#555", va="center")


# ---------------------------------------------------------------------------------------------
# Тайлбар хуудас
# ---------------------------------------------------------------------------------------------
def draw_legend_page(fig, d, tpl, db, n_pages):
    T = TEMPLATES[tpl]
    hole = d["hole"]
    ax = fig.add_axes([0, 0, 1, 1]); ax.set_axis_off()
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.text(X0, 0.955, "%s — ТАЙЛБАР · Загвар %s (%s)" % (hole, tpl, T["name"]),
            fontsize=17, weight="bold", color=NAVY)
    ax.text(X0, 0.928, "%s · %s" % (PROJECT_MN, T["kind"]), fontsize=8.6, color="#555")
    ax.plot([X0, X1], [0.918, 0.918], color=NAVY, lw=1.0)

    # --- зүүн багана: багануудын тайлбар
    y = 0.888
    items = [
        ("Гүн (м)", "Босоо тэнхлэг = цооногийн дагуух гүн, 1:100 (1 см = 1 м). Тод шугам ба тоо 5 м тутам. "
                    "Гүний хазайлт ХЭМЖИГДЭЭГҮЙ тул цооногийг шулуун гэж үзнэ."),
        ("Дээж №", "SGS 2023 оны кернийн сорьц (Assay sheet, %d ш). Интервалын өндөр = сорьцын урт."
                   % len(d["assay"])),
        ("Литологи", "Хээрийн бичиглэлийн код ба өнгө (Codes sheet). Кодын товчлол баруун талын хүснэгтэд."),
        ("Хувирал · бүтэц", "Log_Detail sheet: хувирлын 9 төрлийн эрчим 1–3 (баганын өргөн ба өнгөний "
                            "тодрол = эрчим), судал (хар зураас), сульфидын эзлэх хувь (шар багана, 0–10%)."),
        ("Керний гарц %", "PhotoLog-ийн `Rec_pct_est` — керний фотоос хийсэн үнэлгээ [Таамаг]. "
                          "Ногоон ≥50% · шар 15–50% · улаан <15% (дахин дээжлэхэд материал хүрэлцэхгүй)."),
        ("Элементийн bar", T["note"] + ". Лог масштаб; улаан тасархай = босго; босго давсан утга тодрох "
                                       "өнгөөр, тоогоор бичигдэнэ. Хар хүрээтэй bar = ТДХ давсан (ore-grade AAS) утга."),
("Цоорхой ·\nдахин дээжлэлт", "Зүүн (улаан зураастай) = дээжлэгдээгүй интервал (Sample_Gaps). "
                                     "Баруун өнгө = дахин дээжлэлтийн эрэмбэ v2 дүрмээр тооцсон: "
                                     "1A улаан (CaF₂≥5%-аас ≤3 м / жоншны судал) · 1B улбар (жоншны хүрээ, карбонат хост) · "
                                     "2 шар (CaF₂ өгөгдөлгүй карбонат) · 3 саарал (гранит/лейкогранит). "
                                     "Дээжийн ЯГ хуваарь = Modot3_Resampling_Register_v2_20260902."),
        ("Тайлбар", "Хээрийн бичиглэлийн эх текст (Lithology.Description) — орчуулаагүй, товчилсон."),
        ("Керний фото", "02_Core_Photos-ийн хайрцгийн бүтэн зураг, гүний интервалдаа байрлана. "
                        "ЗҮСЭЛТ хийгээгүй — хайрцгийн зураг бүхэлдээ тавигдсан тул зурган доторх мөр бүр "
                        "гүнтэй шууд шугаман харьцаагүй."),
    ]
    for k, v in items:
        ax.text(X0, y, k, fontsize=9.2, weight="bold", color=NAVY, va="top")
        vw = wrapt(v, 0.38, 8.0)
        ax.text(X0 + 0.130, y, vw, fontsize=8.0, color="#222", va="top", linespacing=1.35)
        y -= 0.030 + 0.0158 * vw.count("\n") + (0.012 if "\n" in k else 0)

    # --- баруун дээд: литологийн код
    xr = 0.585
    ax.text(xr, 0.888, "Литологийн код · өнгө", fontsize=10.5, weight="bold", color=NAVY, va="top")
    yy = 0.862
    used = collections.Counter()
    for r in d["lith"]:
        used[str(r.get("Code"))] += float(r.get("Length_m") or 0)
    for code, m in sorted(used.items(), key=lambda kv: -kv[1]):
        cr = d["codes_map"].get(code) or {}
        ax.add_patch(Rectangle((xr, yy - 0.019), 0.030, 0.019,
                               facecolor=str(cr.get("Striplog_Color") or "#FFF"), edgecolor="#333", lw=0.5))
        ax.text(xr + 0.037, yy - 0.0095, "%s — %s (%s) · %.1f м" %
                (code, cr.get("Name_MN") or "?", cr.get("Name_EN") or "?", m),
                fontsize=7.8, va="center", color="#222")
        yy -= 0.0245

    # --- баруун дунд: хувирлын товчлол
    yy -= 0.012
    ax.text(xr, yy, "Хувирлын товчлол (эрчим 1–3)", fontsize=10.5, weight="bold", color=NAVY, va="top")
    yy -= 0.026
    for i, (key, lbl, colr) in enumerate(ALT_COLS):
        cx = xr + (i % 3) * 0.135
        cy = yy - (i // 3) * 0.024
        ax.add_patch(Rectangle((cx, cy - 0.016), 0.020, 0.016, facecolor=colr, alpha=0.75,
                               edgecolor="#333", lw=0.4))
        nm = {"Alt_Silica": "Цахиуржилт", "Alt_Carbonate": "Карбонатжилт", "Alt_Clay": "Шаваржилт",
              "Alt_Sericite": "Серицитжилт", "Alt_Chlorite": "Хлоритжилт", "Alt_Epidote": "Эпидотжилт",
              "Alt_Kspar": "К-хээрийн жонш", "Alt_Magnetite": "Магнетитжилт",
              "Alt_Fracturing": "Хагарал"}[key]
        ax.text(cx + 0.025, cy - 0.008, "%s %s" % (lbl, nm), fontsize=7.4, va="center", color="#222")
    yy -= 0.024 * math.ceil(len(ALT_COLS) / 3) + 0.014

    # --- баруун доод: цооногийн хураангуй
    ax.text(xr, yy, "Энэ цооногийн хураангуй", fontsize=10.5, weight="bold", color=NAVY, va="top")
    yy -= 0.026
    c = d["collar"]
    gaps_m = sum(float(r.get("Length_m") or 0) for r in d["gaps"])
    caf = [num(r.get("CaF2_pct")) for r in d["caf2"] if num(r.get("CaF2_pct")) is not None]
    lines = [
        "TD %.1f м · хайрцаг %s · дээж %s ш · дээжлэлтийн хамралт %s%%"
        % (d["td"], c.get("n_Core_Boxes"), c.get("n_Samples"), c.get("Coverage_pct")),
        "Дээжлэгдээгүй: %d интервал / %.1f м" % (len(d["gaps"]), gaps_m),
        "CaF₂ (SGS 2025 CLA07C): %s" % ("%d дээж · дээд %.2f%%" % (len(caf), max(caf)) if caf else "өгөгдөл байхгүй"),
        "Загварын босго: " + T["note"],
    ]
    sigs = collections.defaultdict(list)
    for r in d["sig"]:
        sigs[r["Element"]].append(r)
    if sigs:
        parts = []
        for e, rr in sorted(sigs.items()):
            best = max(rr, key=lambda q: (num(q.get("Wtd_Mean_ppm")) or 0) * (num(q.get("Length_m")) or 0))
            parts.append("%s %d огтлол, шилдэг %.2f м @ %s ppm (%.2f–%.2f)"
                         % (e, len(rr), float(best["Length_m"]), ("%.1f" % float(best["Wtd_Mean_ppm"])),
                            float(best["From_m"]), float(best["To_m"])))
        lines.append("Композит огтлол (Sig_Intervals): " + " · ".join(parts))
    if not d["assay"]:
        lines.append("⚠ Энэ цооногт шинжилгээний дүн БАЙХГҮЙ — T4 геологийн лог.")
    if d["hole"] == "MTDH-11":
        lines.append("⚠ MTDH-11: хээрийн 6 сорьц лабд ОГТ ИЛГЭЭГДЭЭГҮЙ (2026-08-30 хайлт).")
    if d["hole"] == "MTDH-15":
        lines.append("⚠ MTDH-15: 52.20–61.70 м (шохойн чулуу) зориуд дээжлэгдээгүй, 10 сорьц сертификатгүй.")
    if d["hole"] == "MTDH-23":
        lines.append("⚠ MTDH-23 61.75–61.90 «FL» код = ХАГАРАЛ (жоншны судал БИШ) — QC-2026-09-03-001.")
    if d["hole"] == "MTDH-26":
        lines.append("⚠ MTDH-26: In_Model = N (өндөр/байрлал баталгаажаагүй) — загварчлалд ороогүй.")
    if d["hole"] == "MTDH-17":
        lines.append("ӨТШ 17-1 = MTDH-17 15.70–16.70 м (CaF₂ 48.56%) [Тооцоолсон, Ж-4].")
    for s in lines:
        sw = wrapt("• " + s, 0.375, 7.8)
        ax.text(xr, yy, sw, fontsize=7.8, color="#222", va="top", linespacing=1.35)
        yy -= 0.021 + 0.0152 * sw.count("\n")

    # --- доод зүүн: дээжлэгдээгүй интервалууд
    ty = min(y - 0.020, 0.395)
    ax.text(X0, ty, "Дээжлэгдээгүй интервал ба дахин дээжлэлтийн эрэмбэ (v2 дүрмээр)",
            fontsize=10.5, weight="bold", color=NAVY, va="top")
    ty -= 0.028
    hdrs = [("Эрэмбэ", 0.055), ("Гүн (м)", 0.085), ("Урт", 0.045), ("Литологи", 0.170), ("CaF₂ ойр", 0.070)]
    xx = X0
    for lab, wd in hdrs:
        ax.add_patch(Rectangle((xx, ty - 0.020), wd, 0.020, facecolor="#EEF2F7", edgecolor="#B9C2CC", lw=0.4))
        ax.text(xx + wd / 2, ty - 0.010, lab, fontsize=7.4, weight="bold", color=NAVY,
                ha="center", va="center")
        xx += wd
    ty -= 0.020
    gg = sorted(d["gaps"], key=lambda r: -float(r.get("Length_m") or 0))[:14]
    gg = sorted(gg, key=lambda r: float(r["From_m"]))
    for i, r in enumerate(gg):
        f, t2 = float(r["From_m"]), float(r["To_m"])
        pr = resample_priority(d, f, t2)
        near = [num(q.get("CaF2_pct")) for q in d["caf2"]
                if num(q.get("CaF2_pct")) and float(q["From_m"]) - 5 <= (f + t2) / 2 <= float(q["To_m"]) + 5]
        vals = [("%s" % pr, 0.055, RS_COL[pr]), ("%.2f–%.2f" % (f, t2), 0.085, None),
                ("%.2f" % float(r.get("Length_m") or 0), 0.045, None),
                (str(r.get("Lithology_Breakdown") or "")[:44], 0.170, None),
                (("%.2f%%" % max(near)) if near else "—", 0.070, None)]
        xx = X0
        if i % 2 == 1:
            ax.add_patch(Rectangle((X0, ty - 0.0185), sum(w for _, w in hdrs), 0.0185,
                                   facecolor="#F6F8FA", lw=0))
        for txt, wd, colr in vals:
            if colr:
                ax.add_patch(Rectangle((xx + 0.010, ty - 0.016), 0.034, 0.013, facecolor=colr, lw=0))
                ax.text(xx + 0.027, ty - 0.0095, txt, fontsize=6.6, color="white", weight="bold",
                        ha="center", va="center")
            else:
                ax.text(xx + 0.004, ty - 0.0095, txt, fontsize=6.8, color="#222", va="center", ha="left")
            xx += wd
        ty -= 0.0185
    if len(d["gaps"]) > len(gg):
        ax.text(X0, ty - 0.012, "… нийт %d интервал (хамгийн урт %d-ыг харуулав)" % (len(d["gaps"]), len(gg)),
                fontsize=6.8, color="#777", va="center")

    ax.text(X0, 0.045, "Хувилбар: cloud v1 (20_make_striplog_MT_A3L_cloud.py) · хуудас %d · "
                       "энэ нь D: дээрх батлагдсан багцын ОРЛУУЛАГЧ БИШ — визуал QC-ийн дараа эцэслэнэ."
            % n_pages, fontsize=7.2, color=RED)
    ax.text(X0, 0.024, PROJECT_MN, fontsize=7.2, color="#666")


# ---------------------------------------------------------------------------------------------
def make_pdf(db, hole, tpl, photo_dir, out_dir, qc_dir=None, date_tag="2026"):
    d = hole_data(db, hole)
    d["codes_map"] = db["codes"]
    if not d["td"]:
        raise SystemExit("%s: TD олдсонгүй" % hole)
    photos = load_photos(photo_dir, hole, d["boxes"]) if photo_dir else []
    n_depth = int(math.ceil(d["td"] / PAGE_M))
    n_pages = n_depth + 1
    out_dir = Path(out_dir) / hole
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / ("%s_DrillLog_A3L_%s_%s.pdf" % (hole, TEMPLATES[tpl]["tag"], date_tag))
    with PdfPages(str(out)) as pdf:
        for i in range(n_depth):
            d0, d1 = i * PAGE_M, min((i + 1) * PAGE_M, math.ceil(d["td"] / PAGE_M) * PAGE_M)
            fig = plt.figure(figsize=A3L)
            draw_depth_page(fig, d, tpl, d0, d0 + PAGE_M, i + 1, n_pages, photos)
            pdf.savefig(fig, dpi=200)
            if qc_dir and i == 0:
                Path(qc_dir).mkdir(parents=True, exist_ok=True)
                fig.savefig(str(Path(qc_dir) / ("%s_%s_p1.png" % (hole, TEMPLATES[tpl]["tag"]))), dpi=110)
            plt.close(fig); gc.collect()
        fig = plt.figure(figsize=A3L)
        draw_legend_page(fig, d, tpl, db, n_pages)
        pdf.savefig(fig, dpi=200)
        if qc_dir:
            fig.savefig(str(Path(qc_dir) / ("%s_%s_legend.png" % (hole, TEMPLATES[tpl]["tag"]))), dpi=110)
        plt.close(fig); gc.collect()
    cz = compress_pdf(out)
    return out, n_pages, cz


def compress_pdf(path, quality=82):
    """Flate-ээр шигтгэсэн зургийг JPEG болгож шахна (Ghostscript-гүй орчинд pikepdf-ээр)."""
    try:
        import pikepdf
        from PIL import Image
        import io as _io
    except Exception:
        return None
    before = Path(path).stat().st_size
    pdf = pikepdf.open(str(path), allow_overwriting_input=True)
    n = 0
    for page in pdf.pages:
        try:
            imgs = dict(page.images)
        except Exception:
            imgs = {}
        for name, raw in imgs.items():
            try:
                pim = pikepdf.PdfImage(raw)
                if str(pim.filters[-1] if pim.filters else "") == "/DCTDecode":
                    continue
                im = pim.as_pil_image().convert("RGB")
                buf = _io.BytesIO()
                im.save(buf, "JPEG", quality=quality, optimize=True)
                raw.write(buf.getvalue(), filter=pikepdf.Name("/DCTDecode"))
                raw.ColorSpace = pikepdf.Name("/DeviceRGB")
                raw.BitsPerComponent = 8
                n += 1
            except Exception:
                continue
    pdf.save(str(path), compress_streams=True, object_stream_mode=pikepdf.ObjectStreamMode.generate)
    pdf.close()
    return before, Path(path).stat().st_size, n


DEF_DB = "drive/MT_Drilling_Database.xlsx"


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("hole", nargs="?")
    ap.add_argument("template", nargs="?", choices=list(TEMPLATES))
    ap.add_argument("--db", default=DEF_DB)
    ap.add_argument("--photos", default="photos")
    ap.add_argument("--out", default="Out_Target")
    ap.add_argument("--qc", default="Out_Target/QC_PNG")
    ap.add_argument("--map", default=None, help="hole_template_map csv — --all-тай хамт")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--only", default=None, help="зөвхөн эдгээр цооног (таслалаар)")
    ap.add_argument("--date", default="2026")
    a = ap.parse_args(argv)

    db = load_db(a.db)
    jobs = []
    if a.all:
        with open(a.map, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                h = (r.get("Hole_ID") or "").strip()
                if not h:
                    continue
                for key in ("Template", "Second_template_if_any"):
                    m = re.search(r"T[1-4]", (r.get(key) or "").upper())
                    if m:
                        jobs.append((h, m.group(0)))
    else:
        jobs = [(a.hole, a.template)]
    if a.only:
        keep = {s.strip().upper() for s in a.only.split(",")}
        jobs = [j for j in jobs if j[0].upper() in keep]

    for h, t in jobs:
        out, n, cz = make_pdf(db, h, t, a.photos, a.out, a.qc, a.date)
        extra = ("  (шахалт %.1f→%.1f MB, %d зураг)" % (cz[0] / 1e6, cz[1] / 1e6, cz[2])) if cz else ""
        print("%-9s %s  %2d х.  %6.1f MB  %s%s" % (h, t, n, out.stat().st_size / 1e6, out.name, extra))
        sys.stdout.flush()
    return 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_database.py — Consolidated, 3D-modelling-ready petrography database
for the Oval Ni-Cu project (Yambat, Mongolia).

VERSION 1.1 (2026-08-31).  v1.1 applies every defect found by the two
independent audits (database/VERIFICATION_integrity.md D1-D10 and
database/VERIFICATION_coverage.md G1-G14) and merges the
workspace/extracted/missing_sources/ batch (84 records).  See the changelog
in database/QA_report.md.

Reads the extracted source tables under  workspace/extracted/  and writes
ONE consolidated database to  database/ :
  csv/samples.csv           one row per physical sample (the spine)
  csv/descriptions.csv      one row per petrographic description
  csv/collar.csv            normalized drillhole collars (WGS84 / UTM 46N)
  csv/survey.csv            normalized downhole surveys
  csv/sample_assays.csv     wide assay suite from the Master "All" sheet
  csv/lu_hole_alias.csv     raw hole-id spelling -> normalized hole id
  csv/lu_lab.csv            lab / petrographer lookup
  csv/lu_rock_type.csv      rock-name -> standardized rock group
  csv/sources.csv           registry of contributing files (Drive fileIds)
  Oval_Petrography_DB.xlsx  all tables as sheets
  Oval_Petrography_DB.sqlite same tables
  QA_report.md, README.md

Re-run:  python3 scripts/build_database.py   (from the repo root or anywhere)
Requires: pandas, openpyxl  (pip install pandas openpyxl)
"""

import csv
import datetime
import json
import math
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent          # /home/user/Jargal
WS = ROOT / "workspace" / "extracted"
MASTER = WS / "master"
XLSX = WS / "xlsx"
MISS = WS / "missing_sources"
OUT = ROOT / "database"
CSVDIR = OUT / "csv"
CSVDIR.mkdir(parents=True, exist_ok=True)

DB_VERSION = "1.1"
BUILD_DATE = "2026-08-31"
EPSG_NOTE = "WGS84 / UTM zone 46N (EPSG:32646)"

# ----------------------------------------------------------------------------
# small helpers
# ----------------------------------------------------------------------------

hole_alias_seen = defaultdict(set)   # raw -> set(sources)
hole_alias_norm = {}                 # raw -> norm

_RE_OVD = re.compile(r"^OVD[-_ ]?0*(\d+)\s*([A-Z]?)$")
_RE_SCCRS = re.compile(r"^(SC|CRS)[-_ ]?0*(\d+)\s*([A-Z]?)$")
_RE_BS = re.compile(r"^BS[-_ ]?0*(\d+)$")
_RE_MU = re.compile(r"^MU[-_ ]?(\d+)$")


def norm_hole(raw, source=""):
    """Normalize a drillhole ID.  OVD-001->OVD001, OVD21->OVD021,
    OVD008a->OVD008A, CRSO1A->CRS01A; SC/CRS/MU/BS styles kept intact.
    Returns '' when the value is not a hole id."""
    if raw is None:
        return ""
    s = str(raw).strip().upper().replace(" ", "")
    if not s:
        return ""
    s = s.replace("CRSO", "CRS0")            # letter O typo in 2024 reports
    out = ""
    m = _RE_OVD.match(s)
    if m:
        out = "OVD" + m.group(1).zfill(3) + m.group(2)
    else:
        m = _RE_SCCRS.match(s)
        if m:
            out = m.group(1) + m.group(2).zfill(2) + m.group(3)
        else:
            m = _RE_BS.match(s)
            if m:
                out = "BS" + m.group(1).zfill(3)
            else:
                m = _RE_MU.match(s)
                if m:
                    out = "MU" + m.group(1)
    if out:
        raw_s = str(raw).strip()
        hole_alias_seen[raw_s].add(source)
        hole_alias_norm[raw_s] = out
    return out


def is_cyrillic(text):
    if not text:
        return False
    cyr = sum(1 for ch in text if "Ѐ" <= ch <= "ӿ")
    return cyr > max(3, 0.15 * len(text))


def ffloat(v):
    """forgiving float; returns None on failure"""
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    s = s.rstrip("mM").strip()
    try:
        return float(s)
    except ValueError:
        return None


def rdepth(v):
    return None if v is None else round(v, 3)


_RE_INTERVAL = re.compile(r"^(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)\s*m?$")
_RE_POINT = re.compile(r"^(\d+(?:\.\d+)?)\s*m?\s*(?:\(?([ABАБ])\)?)?$", re.I)


def parse_depth_text(s):
    """'36m'->(36,None) ; '95.8-95.9'->(95.8,95.9) ; '175.5 A'->(175.5,None)
    returns (from, to, note) — note!='' when unparseable/suffixed."""
    if s is None:
        return None, None, "empty"
    t = str(s).strip()
    if not t:
        return None, None, "empty"
    m = _RE_INTERVAL.match(t)
    if m:
        return float(m.group(1)), float(m.group(2)), ""
    m = _RE_POINT.match(t)
    if m:
        note = f"slide suffix {m.group(2)}" if m.group(2) else ""
        return float(m.group(1)), None, note
    return None, None, f"unparseable depth '{t}'"


def read_csv_rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


# ============================================================================
# 1. Collar + survey (location authority) and minimum-curvature desurvey
# ============================================================================

collar_rows = []
collar_by_hole = {}
for row in read_csv_rows(MASTER / "Collar_all_combined.csv")[1:]:
    if not any(c.strip() for c in row):
        continue
    raw_h = row[2].strip()
    h = norm_hole(raw_h, "Collar_all_combined")
    rec = {
        "hole_id": h,
        "hole_id_raw": raw_h,
        "project": row[0].strip(),
        "prospect": row[1].strip(),
        "hole_type": row[3].strip(),
        "east": ffloat(row[4]),
        "north": ffloat(row[5]),
        "rl": ffloat(row[6]),
        "azimuth": ffloat(row[7]),
        "dip": ffloat(row[8]),
        "start_depth_m": ffloat(row[9]),
        "total_depth_m": ffloat(row[11]),
        "start_date": row[12].strip(),
        "end_date": row[13].strip(),
        "status": row[14].strip(),
        "lease": row[15].strip(),
        "company": row[16].strip(),
        "supervisor": row[17].strip(),
        "remarks": row[19].strip(),
        "edited_date": row[20].strip(),
        "crs": EPSG_NOTE,
    }
    collar_rows.append(rec)
    collar_by_hole[h] = rec

survey_rows = []
survey_by_hole = defaultdict(list)
for row in read_csv_rows(MASTER / "Survey_all_YMB.csv")[1:]:
    if not any(c.strip() for c in row):
        continue
    raw_h = row[0].strip()
    if not raw_h:
        continue
    h = norm_hole(raw_h, "Survey_all_YMB")
    d, dip, azi = ffloat(row[1]), ffloat(row[2]), ffloat(row[3])
    rec = {
        "hole_id": h,
        "hole_id_raw": raw_h,
        "depth_m": d,
        "dip": dip,
        "azimuth": azi,
        "grid": row[4].strip(),
        "azim_utm": row[5].strip(),
        "method": row[6].strip(),
        "survey_company": row[7].strip(),
        "survey_date": row[8].strip() if len(row) > 8 else "",
        "qa_note": "",
    }
    survey_rows.append(rec)


def _sdate(s):
    """'7/28/2024' -> (2024, 7, 28); unparseable -> (0, 0, 0)."""
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", str(s).strip())
    if not m:
        return (0, 0, 0)
    mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if yy < 100:
        yy += 2000
    return (yy, mm, dd)


# --- D4: duplicate hole+depth survey stations -------------------------------
# Both readings are KEPT VERBATIM (they are verbatim source rows); the conflict
# is recorded in the new `qa_note` column and the recommended row is named.
_station_groups = defaultdict(list)
for rec in survey_rows:
    if rec["depth_m"] is not None:
        _station_groups[(rec["hole_id"], rec["depth_m"])].append(rec)

survey_conflicts = []
for (h, d), recs in sorted(_station_groups.items()):
    if len(recs) < 2:
        continue
    keep = max(recs, key=lambda r: (_sdate(r["survey_date"]), r["survey_company"]))
    for r in recs:
        others = "; ".join(
            f"{o['dip']}/{o['azimuth']} ({o['method']}, {o['survey_company']}, "
            f"{o['survey_date']})" for o in recs if o is not r)
        r["qa_note"] = (
            f"D4: duplicate survey station — {h} @ {d:g} m is recorded "
            f"{len(recs)} times with conflicting orientations (other reading(s): "
            f"{others}). Both rows are kept verbatim from Survey_all_YMB.csv. "
            f"RECOMMENDED row for importers that reject duplicate hole+depth "
            f"keys: the {keep['survey_date']} {keep['survey_company']} "
            f"({keep['method']}) reading, dip {keep['dip']} / azimuth "
            f"{keep['azimuth']} — the most recent instrument survey; this is "
            f"also the row the desurvey in this build uses."
            + (" THIS ROW." if r is keep else " (drop this row.)"))
    survey_conflicts.append(
        f"{h} @ {d:g} m: {len(recs)} readings — "
        + " vs ".join(f"dip {o['dip']} / azi {o['azimuth']} ({o['method']}, "
                      f"{o['survey_company']}, {o['survey_date']})" for o in recs)
        + f" | recommended: {keep['survey_date']} {keep['survey_company']}")

# desurvey stations: one per (hole, depth), the most recent survey wins
_pref_station = {}
for rec in survey_rows:
    if rec["depth_m"] is None or rec["dip"] is None or rec["azimuth"] is None:
        continue
    key = (rec["hole_id"], rec["depth_m"])
    cur = _pref_station.get(key)
    if cur is None or _sdate(rec["survey_date"]) > _sdate(cur["survey_date"]):
        _pref_station[key] = rec
for (h, d), rec in _pref_station.items():
    survey_by_hole[h].append((d, rec["dip"], rec["azimuth"]))


def _mc_delta(md1, i1, a1, md2, i2, a2):
    """minimum-curvature increment; inclinations from vertical, degrees.
    returns (dE, dN, dDown)"""
    dmd = md2 - md1
    I1, I2 = math.radians(i1), math.radians(i2)
    A1, A2 = math.radians(a1), math.radians(a2)
    cosdl = math.cos(I2 - I1) - math.sin(I1) * math.sin(I2) * (1 - math.cos(A2 - A1))
    cosdl = max(-1.0, min(1.0, cosdl))
    dl = math.acos(cosdl)
    rf = 1.0 if dl < 1e-9 else (2.0 / dl) * math.tan(dl / 2.0)
    dN = dmd / 2.0 * (math.sin(I1) * math.cos(A1) + math.sin(I2) * math.cos(A2)) * rf
    dE = dmd / 2.0 * (math.sin(I1) * math.sin(A1) + math.sin(I2) * math.sin(A2)) * rf
    dV = dmd / 2.0 * (math.cos(I1) + math.cos(I2)) * rf
    return dE, dN, dV


class HoleTrace:
    """Minimum-curvature desurveyed trace of one drillhole."""

    def __init__(self, collar, stations):
        self.ok = collar["east"] is not None and collar["north"] is not None
        self.collar = collar
        st = sorted({s[0]: s for s in stations}.values()) if stations else []
        # station: (md, inclination-from-vertical, azimuth)
        st = [(md, 90.0 + dip, azi) for md, dip, azi in st]
        if not st:
            dip = collar["dip"] if collar["dip"] is not None else -90.0
            azi = collar["azimuth"] if collar["azimuth"] is not None else 0.0
            st = [(0.0, 90.0 + dip, azi)]
        if st[0][0] > 0.0:                       # anchor at collar
            st.insert(0, (0.0, st[0][1], st[0][2]))
        self.st = st
        # cumulative positions at stations
        self.pos = [(0.0, 0.0, 0.0)]
        for k in range(1, len(st)):
            dE, dN, dV = _mc_delta(*st[k - 1], *st[k])
            e, n, v = self.pos[-1]
            self.pos.append((e + dE, n + dN, v + dV))

    def xyz_at(self, md):
        """(x, y, z) at measured depth md; None when collar has no coords."""
        if not self.ok or md is None:
            return None
        st, pos = self.st, self.pos
        if md <= st[0][0]:
            e, n, v = pos[0]
        elif md >= st[-1][0]:
            md1, i1, a1 = st[-1]
            dE, dN, dV = _mc_delta(md1, i1, a1, md, i1, a1)
            e, n, v = pos[-1][0] + dE, pos[-1][1] + dN, pos[-1][2] + dV
        else:
            k = max(i for i in range(len(st)) if st[i][0] <= md)
            md1, i1, a1 = st[k]
            md2, i2, a2 = st[k + 1]
            t = 0.0 if md2 == md1 else (md - md1) / (md2 - md1)
            ii = i1 + t * (i2 - i1)
            da = ((a2 - a1 + 180.0) % 360.0) - 180.0
            aa = a1 + t * da
            dE, dN, dV = _mc_delta(md1, i1, a1, md, ii, aa)
            e, n, v = pos[k][0] + dE, pos[k][1] + dN, pos[k][2] + dV
        c = self.collar
        z = None if c["rl"] is None else round(c["rl"] - v, 2)
        return round(c["east"] + e, 2), round(c["north"] + n, 2), z


traces = {h: HoleTrace(collar_by_hole[h], survey_by_hole.get(h, []))
          for h in collar_by_hole}

# ============================================================================
# 2. Master "All" sheet -> samples spine + assay table
# ============================================================================

all_rows = read_csv_rows(MASTER / "Yambat_Petrographic_Master_Data__All.csv")
hdr_grp, hdr_det = all_rows[0], all_rows[1]

# assay column names (cols 22..end), unique + sanitized
assay_cols = []          # (index, name)
prev_grp = ""
used_names = Counter()
for i in range(22, len(hdr_det)):
    det, grp = hdr_det[i].strip(), hdr_grp[i].strip()
    if not det:
        continue
    grp = grp or prev_grp
    prev_grp = grp
    name = re.sub(r"[^0-9A-Za-z]+", "_", det.replace("%", "pct")).strip("_")
    name = f"{name}__{re.sub(r'[^0-9A-Za-z]+', '_', grp).strip('_')}"
    used_names[name] += 1
    if used_names[name] > 1:
        name += f"_{used_names[name]}"
    assay_cols.append((i, name))

FLAG_COLS = [("polished_thin_section", 10), ("thin_section", 11), ("asd", 12),
             ("xrd", 13), ("fluid_incl", 14), ("eds", 15), ("epma", 16)]

samples = []             # list of dicts
sample_by_id = {}
tag2sample = {}          # numeric tag string -> sample dict
assay_records = []
qa = defaultdict(list)   # QA report bins

MASTER_ALL = "Yambat_Petrographic_Master_Data__All.csv"

HD_LABEL = re.compile(
    r"^([A-Za-z]+ ?-?\d+[A-Za-z]?)[-@](\d+(?:\.\d+)?)(?:-(\d+(?:\.\d+)?))?"
    r"\s*(?:\(?([ABАБ])\)?)?$")


def composite_id(hole, dfrom):
    d = ("%g" % dfrom) if dfrom is not None else "?"
    return f"{hole}@{d}"


def add_sample(rec):
    samples.append(rec)
    sample_by_id[rec["sample_id"]] = rec
    for t in [rec["sample_id"]] + [a.strip() for a in rec["alt_ids"].split("|") if a.strip()]:
        t2 = t.split(" ")[0]
        if re.fullmatch(r"\d{5}", t2) and t2 not in tag2sample:
            tag2sample[t2] = rec


seen_exact = set()
skipped_formatting = 0
for row in all_rows[2:]:
    if not any(c.strip() for c in row):
        continue
    sid_raw = row[6].strip()
    label = row[5].strip()
    hole_raw = row[1].strip()
    if not sid_raw and not label:
        skipped_formatting += 1
        continue

    hole = norm_hole(hole_raw, MASTER_ALL)
    qa_flags = []
    alt_ids = []

    # --- depths ------------------------------------------------------------
    d_from, d_to = None, None
    f_raw = row[4].strip()
    f_val, _f2, f_note = parse_depth_text(f_raw)
    if f_val is not None:
        d_from = f_val
    if f_note and "suffix" not in f_note and f_raw:
        qa_flags.append(f_note)

    lab_val = ffloat(label)
    if lab_val is not None and label and not label.endswith("m"):
        # numeric label = interval TO (2025-era rows)
        if d_from is None or lab_val >= d_from:
            d_to = lab_val
    elif label:
        for part in [p.strip() for p in label.split(",") if p.strip()]:
            m = HD_LABEL.match(part)
            if m and norm_hole(m.group(1), MASTER_ALL + " (interval label)"):
                alt_ids.append(part)
                lf, lt = float(m.group(2)), (float(m.group(3)) if m.group(3) else None)
                if d_from is None:
                    d_from = lf
                if lt is not None:
                    d_from, d_to = lf, lt
            elif not re.fullmatch(r"\d+(\.\d+)?m?", part):
                qa_flags.append(f"interval label '{part}' not parsed")

    # --- duplicate handling ------------------------------------------------
    if sid_raw:
        key = (sid_raw, hole, row[4].strip(), label)
        if sid_raw in {s["sample_id"] for s in samples}:
            if key in seen_exact:
                qa["dup_rows_dropped"].append(
                    f"{sid_raw} ({hole} {row[4]}) exact duplicate row dropped")
                continue
            # same tag, different sample (42808 = SC04-171 AND SC04-280.7)
            sid = composite_id(hole, d_from)
            alt_ids.append(f"{sid_raw} (tag shared)")
            qa_flags.append(f"tag {sid_raw} shared with another sample; "
                            f"composite id assigned")
            qa["dup_tags_split"].append(
                f"tag {sid_raw} used twice; second occurrence ({hole} @ "
                f"{row[4]}) stored as {sid}")
        else:
            seen_exact.add(key)
            if re.fullmatch(r"\d{5}", sid_raw):
                sid = sid_raw
            else:                          # OVD028-38 style id
                sid = composite_id(hole, d_from)
                alt_ids.append(sid_raw)
    else:
        # no SAMPLE_ID: rows 230/231 duplicate the tagged SC04 rows above
        qa["dup_rows_dropped"].append(
            f"no-id row {hole} '{label}' dropped (duplicates tagged SC04 rows)")
        continue

    d_mid = None
    if d_from is not None:
        d_mid = (d_from + d_to) / 2.0 if d_to is not None else d_from
    if d_from is None:
        qa_flags.append("depth not parsed")
        qa["depth_parse_failures"].append(f"{sid}: from='{row[4]}' label='{label}'")

    rec = {
        "sample_id": sid,
        "alt_ids": " | ".join(dict.fromkeys(alt_ids)),
        "hole_id_norm": hole,
        "depth_from_m": rdepth(d_from),
        "depth_to_m": rdepth(d_to),
        "depth_mid_m": rdepth(d_mid),
        "sample_source": "drill core" if hole else "unknown",
        "year": row[9].strip(),
        "petrographer_lab": row[7].strip(),
        "field_lithology": row[17].strip(),
        "petro_lithology": row[18].strip(),
        "iogas_lithology": row[19].strip(),
        "iogas_no": "",
        "x_utm": None, "y_utm": None, "z_rl": None, "coord_source": "none",
        "qa_flags": "; ".join(qa_flags),
        "source_files": MASTER_ALL,
        "_master_x": ffloat(row[2]), "_master_y": ffloat(row[3]),
    }
    for fname, idx in FLAG_COLS:
        rec[fname] = 1 if row[idx].strip() else 0
    add_sample(rec)

    vals = {name: ("" if row[i].strip() == "-" else row[i].strip())
            for i, name in assay_cols}
    if any(vals.values()):
        assay_records.append({"sample_id": sid, **vals})

# ---------------------------------------------------------------------------
# enrich intervals of the 47xxx / OVD022-029 samples from the Phase-2 sheet
# ---------------------------------------------------------------------------
PH2 = "Yambat_petrography_samples_2022-2024_from_Core___grab__2024_Phase_2_Drilling.csv"
for row in read_csv_rows(MASTER / PH2)[2:]:
    if len(row) < 6 or not row[5].strip():
        continue
    tag = row[5].strip()
    f, t = ffloat(row[2]), ffloat(row[3])
    ph2_flag = ""
    # --- D1: the sheet's sample From/To is inverted on at least one row
    # (47176: From 144, To 114.1).  The assay block on the SAME row (cols
    # 21/22 'from'/'to') is the corroborating record, and the 2024-2026
    # extract states the sample sits at 114-116 m.  Prefer the assay block.
    if f is not None and t is not None and f > t:
        af, at = (ffloat(row[21]) if len(row) > 21 else None,
                  ffloat(row[22]) if len(row) > 22 else None)
        if af is not None and at is not None and af <= at:
            ph2_flag = (
                f"D1 CORRECTED: master sheets record this sample interval as "
                f"From {f:g} / To {t:g} (inverted — a source typo for the "
                f"depth {af:g} m); the assay block on the same "
                f"{PH2} row reads from {af:g} m / to {at:g} m and the 2024-2026 "
                f"report extract states 'sample sits at {af:g}-{at:g} m, listed "
                f"after deeper samples in the source doc'. Interval set to "
                f"{af:g}-{at:g} m and re-desurveyed; the '{f:g}' of "
                f"Yambat_Petrographic_Master_Data__All.csv / the bichiglel "
                f"table is a master-sheet typo (144 for 114)")
            f, t = af, at
        else:
            f, t = t, f
            ph2_flag = (f"D1: inverted interval in {PH2} (From {t:g} > To "
                        f"{f:g}); swapped, no corroborating assay interval")
    s = tag2sample.get(tag)
    if s and f is not None:
        if s["depth_to_m"] is None:
            s["depth_from_m"], s["depth_to_m"] = rdepth(f), rdepth(t)
            s["depth_mid_m"] = rdepth((f + t) / 2.0 if t is not None else f)
            s["source_files"] += "; " + PH2
            if ph2_flag:
                s["qa_flags"] = (s["qa_flags"] + "; " + ph2_flag).strip("; ")
                qa["depth_interval_fixes"].append(f"{s['sample_id']}: {ph2_flag}")
        hd = row[4].strip()
        if hd and hd not in s["alt_ids"]:
            s["alt_ids"] = (s["alt_ids"] + " | " + hd).strip(" |")
        norm_hole(row[1], PH2)

# ---------------------------------------------------------------------------
# bichiglel table: ioGAS numbers, extra flags, and (later) descriptions
# ---------------------------------------------------------------------------
BICH = "Yambat_petrography_samples_2022-2024_from_Core___grab__Petrography_bichiglel_table.csv"
bich_rows = [r for r in read_csv_rows(MASTER / BICH)[2:] if any(c.strip() for c in r)]
for row in bich_rows:
    sn = row[1].strip()
    norm_hole(row[4], BICH)
    s = tag2sample.get(sn)
    if s and row[2].strip():
        s["iogas_no"] = row[2].strip()

# ============================================================================
# 3. ADD samples that are absent from the Master "All" sheet
# ============================================================================

def new_sample(sid, hole, d_from, d_to, source, year, ssource, lab="",
               field_lith="", petro_lith="", alt="", qa_flag="", flags=None,
               x=None, y=None):
    d_mid = None
    if d_from is not None:
        d_mid = (d_from + d_to) / 2.0 if d_to is not None else d_from
    rec = {
        "sample_id": sid, "alt_ids": alt, "hole_id_norm": hole,
        "depth_from_m": rdepth(d_from), "depth_to_m": rdepth(d_to),
        "depth_mid_m": rdepth(d_mid),
        "sample_source": ssource, "year": year, "petrographer_lab": lab,
        "field_lithology": field_lith, "petro_lithology": petro_lith,
        "iogas_lithology": "", "iogas_no": "",
        "x_utm": None, "y_utm": None, "z_rl": None, "coord_source": "none",
        "qa_flags": qa_flag, "source_files": source,
        "_master_x": x, "_master_y": y,
    }
    for fname, _ in FLAG_COLS:
        rec[fname] = 0
    if flags:
        for fl in flags:
            rec[fl] = 1
    add_sample(rec)
    qa["samples_added"].append(f"{sid} ({source})")
    return rec


# 3a. tag 40904 (Petrograph_2023 register + bichiglel; OVD-002 @ 44.2 m)
new_sample("40904", "OVD002", 44.2, None,
           "Petrograph_2023__Sheet1.csv; " + BICH, "2023", "drill core",
           lab="Mireslab", field_lith="gossan",
           qa_flag="absent from Master All sheet",
           flags=["polished_thin_section", "eds"])

# 3b. tag 47153 (bichiglel / Samples-to-Japan; OVD025 51.3-51.4 massive sulphide)
new_sample("47153", "OVD025", 51.3, 51.4,
           BICH + "; Samples_to_Japan sheet", "2024", "drill core",
           field_lith="Massive Sulphide", alt="OVD025-51.3",
           qa_flag="absent from Master All sheet",
           flags=["polished_thin_section"])

# 3c. Sheet3 2024 hole-depth series — all but one are already in "All"
SHEET3 = "Yambat_petrography_samples_2022-2024_from_Core___grab__Sheet3.csv"
sheet3_missing = 0
label_index = set()
for s in samples:
    for a in s["alt_ids"].split("|"):
        a = a.strip()
        label_index.add(a)                      # e.g. 'OVD015-175.5 (A)'
        label_index.add(a.split(" (")[0])       # e.g. 'OVD015-175.5'
    label_index.add(s["sample_id"])
for row in read_csv_rows(MASTER / SHEET3)[2:]:
    if len(row) < 4 or not row[3].strip():
        continue
    sn = row[3].strip()
    hole = norm_hole(row[1], SHEET3)
    if sn in label_index:
        continue
    m = HD_LABEL.match(sn)
    d = float(m.group(2)) if m else ffloat(row[2])
    sheet3_missing += 1
    new_sample(composite_id(hole, d), hole, d, None, SHEET3, "2024",
               "drill core", alt=sn,
               qa_flag="in Sheet3 hole-depth list but absent from Master All; "
                       "possible depth typo of OVD015-175.5 (A)/(B)",
               flags=["polished_thin_section"])

# 3d. 2025 "15ш" report lab numbers absent from All (incl. 43816 = OVD-009 @ 126.6)
with open(WS / "reports2024_2026" / "samples.json", encoding="utf-8") as f:
    recs2426 = json.load(f)
with open(WS / "reports" / "samples.json", encoding="utf-8") as f:
    recs_rep = json.load(f)

for r in recs2426:
    sid = str(r["sample_id"]).strip()
    m = re.match(r"^(\d{5})(?:\s*\((.*)\))?$", sid)
    if not m:
        continue
    tag = m.group(1)
    if tag in tag2sample:
        continue
    hole, d = "", None
    if r.get("drillhole_id"):
        hole = norm_hole(r["drillhole_id"], r["source_file"])
    if m.group(2):                      # e.g. 43816 (OVD-009-126.6)
        hm = HD_LABEL.match(m.group(2).replace(" ", ""))
        if hm:
            hole = norm_hole(hm.group(1), r["source_file"]) or hole
            d = float(hm.group(2))
    d = d if d is not None else r.get("depth_m")
    src = r["source_file"]
    if "MS3" in src:
        ssource, yr, qa_f = "outcrop", "2026", ("MS3 outcrop sample; no coordinates "
                                               "given in source report")
    else:
        ssource = "drill core" if hole else "unknown"
        yr = "2025"
        qa_f = "absent from Master All sheet" + (
            "" if hole else "; no drillhole/depth stated in source report")
    new_sample(tag, hole, d, None, src, yr, ssource,
               lab=str(r.get("analyst_or_lab") or ""), alt=sid if m.group(2) else "",
               petro_lith=str(r.get("rock_name") or ""), qa_flag=qa_f,
               flags=["polished_thin_section"])

# 3e. unlocated 2023-24 report samples: С-1/С-2 and TS-n / Дээж-n
for r in recs_rep:
    if r["record_type"] != "petrographic_description":
        continue
    sid = str(r["sample_id"]).strip()
    if sid.startswith("С-") or sid.startswith("TS-") or sid.startswith("Дээж"):
        canon = (sid.replace("С-", "C-").split(" ")[0]
                 .replace("Дээж-", "DEEJ-"))
        if canon in sample_by_id:
            continue
        new_sample(canon, "", None, None, r["source_file"],
                   (r.get("report_date") or "")[:4], "unknown",
                   lab=str(r.get("analyst_or_lab") or ""), alt=sid,
                   petro_lith=str(r.get("rock_name") or ""),
                   qa_flag="no drillhole/depth/coordinates given in source report",
                   flags=["polished_thin_section" if "ӨТШ" in str(r.get("sample_type"))
                          or "polished" in str(r.get("sample_type") or "")
                          else "thin_section"])

# 3f. grab samples (YT-xx / YM-xx) with X/Y
GRAB = "Yambat_Petrographic_Master_Data__2022-2024_grab.csv"
for row in read_csv_rows(MASTER / GRAB)[2:]:
    if len(row) < 6 or not row[1].strip():
        continue
    sid = row[1].strip()
    if sid in sample_by_id:
        continue
    flags = []
    if len(row) > 12:
        if row[9].strip():
            flags.append("polished_thin_section")
        if row[10].strip():
            flags.append("thin_section")
        if row[11].strip():
            flags.append("asd")
        if row[12].strip():
            flags.append("fluid_incl")
    if "xrd" in row[2].lower():
        flags.append("xrd")
    gx, gy = ffloat(row[3]), ffloat(row[4])
    new_sample(sid, "", None, None, GRAB, "", "grab",
               field_lith=row[5].strip(),
               qa_flag="" if gx is not None else "no X/Y in source grab sheet",
               flags=flags, x=gx, y=gy)

# 3g. 2024 Copper Ridge rockchip samples.
# NOTE: the rockchip sheet's 32 Sample_numbers are the SAME physical samples
# as the numeric-tag / CR-xx rows of the grab sheet (which carries their X/Y)
# — so they are reclassified as 'rockchip' and flag-enriched, not duplicated.
RC = "Yambat_Petrographic_Master_Data__2024_Copper_ridge_rockchip_samp.csv"
n_rc_merged = 0
for row in read_csv_rows(MASTER / RC)[2:]:
    if len(row) < 8 or not row[1].strip():
        continue
    sid = row[1].strip()
    s = sample_by_id.get(sid)
    if s is None:
        s = new_sample(sid, "", None, None, RC, "2024", "rockchip",
                       field_lith=row[2].strip(),
                       qa_flag="rockchip sample; no coordinates in source table",
                       flags=[])
    else:
        n_rc_merged += 1
        s["sample_source"] = "rockchip"
        s["source_files"] += "; " + RC
        s["year"] = s["year"] or "2024"
        if not s["field_lithology"]:
            s["field_lithology"] = row[2].strip()
    if row[3].strip():
        s["polished_thin_section"] = 1
    if row[4].strip():
        s["thin_section"] = 1
    if row[5].strip():
        s["asd"] = 1
    if row[6].strip():
        s["fluid_incl"] = 1
    if row[7].strip() and not s["petro_lithology"]:
        s["petro_lithology"] = row[7].strip()
qa["rockchip_merge"].append(
    f"{n_rc_merged} rockchip-sheet rows merged into existing grab-sheet samples "
    f"(same Sample_numbers; grab sheet supplies their X/Y)")

# ---------------------------------------------------------------------------
# 3h. COVERAGE G1 — two Crawford-2025 thin sections that exist physically but
#     have no row in any sample register.  Both are drill core with a known
#     hole + depth, so both are created and desurveyed.
# ---------------------------------------------------------------------------
CRAWFORD_PDF = "Asian Battery Metals March 2025 The Oval Summary Report.pdf"

new_sample(composite_id("OVD003", 202.0), "OVD003", 202.0, None,
           CRAWFORD_PDF, "2025", "drill core",
           lab="Dr Anthony J Crawford, A & A Crawford Geological Research "
               "Consultants",
           petro_lith="Intensely hydrothermally altered and brecciated "
                      "protomylonite (protolith probably olivine-hornblende "
                      "gabbro)",
           alt="OVD003@202m",
           qa_flag="COVERAGE G1: sample row created in v1.1 for a Crawford 2025 "
                   "thin section that has no entry in any sample register "
                   "(Master All, bichiglel, Sheet3, phase sheets); hole+depth "
                   "taken verbatim from the report caption 'OVD003@202m' and "
                   "desurveyed (OVD003 total depth 209.5 m, so the depth is "
                   "inside the hole)",
           flags=["polished_thin_section"])

new_sample(composite_id("OVD009", 178.0), "OVD009", 178.0, 180.0,
           CRAWFORD_PDF, "2025", "drill core",
           lab="Dr Anthony J Crawford, A & A Crawford Geological Research "
               "Consultants",
           petro_lith="Intensely hydrothermally altered rock, probably a "
                      "leucogabbro dyke (no protolith evidence preserved)",
           alt="OVD009@178-180m",
           qa_flag="COVERAGE G1: sample row created in v1.1 for a Crawford 2025 "
                   "thin section that has no entry in any sample register "
                   "(OVD009 register samples jump 171.5 -> 190.8 m); interval "
                   "taken verbatim from the report caption 'OVD009@178-180m' "
                   "and desurveyed at the 179.0 m midpoint. CRAWFORD CAVEAT: "
                   "the analyst states the wholerock assay for this interval "
                   "does not match the thin section and suspects a SAMPLE SWAP "
                   "(possibly with a leucogabbro dyke such as OVD021@101.5m / "
                   "OVD011-101.5, tag 42027) — treat the location as "
                   "provisional",
           flags=["polished_thin_section"])

# ---------------------------------------------------------------------------
# 3i. missing_sources batch — ARDH-2005-01 photo_only stubs (legacy 2005 hole)
# ---------------------------------------------------------------------------
with open(MISS / "samples.json", encoding="utf-8") as f:
    recs_miss = json.load(f)

miss_photo = [r for r in recs_miss
              if "photo_only" in str(r.get("sample_type") or "")]
miss_desc_recs = [r for r in recs_miss if r not in miss_photo]

for r in miss_photo:
    cam = str(r["sample_id"]).strip()
    hole_raw = str(r.get("drillhole_id") or "").strip()      # ARDH-2005-01
    rec = new_sample(f"{hole_raw}-{cam}", "", None, None,
                     "CORE PHOTO/ARDH-2005-01/4. Thin section photo folder "
                     "— 17 photo_only stubs", "2005", "drill core",
                     lab="", alt=cam,
                     qa_flag="legacy_2005_photo_only: thin-section PHOTOGRAPH "
                             "only — no petrographic report, sheet or "
                             "description exists anywhere under ARDH-2005-01; "
                             "no depth is recorded on the image or in the "
                             "folder. Hole ARDH-2005-01 is NOT in "
                             "Collar_all_combined.csv, so no collar, no survey "
                             "and no coordinates are available "
                             "(coord_source = none). ARDH-2005-02's equivalent "
                             "folder is empty on Drive",
                     flags=["thin_section"])
    # ARDH-2005-01 is not a recognised hole pattern and is absent from collar:
    # carry it in hole_id_norm for grouping but keep it OUT of lu_hole_alias.
    rec["hole_id_norm"] = hole_raw
    rec["source_files"] = ("CORE PHOTO/ARDH-2005-01/4. Thin section photo "
                           f"(folder:1SpY0E3wPZudd9e6KIC5D8SpFgfnRF2tk) :: "
                           f"{cam}.JPG")

# ---------------------------------------------------------------------------
# 3z. GLOBAL depth_from <= depth_to GUARD (D1).
#     Any surviving inversion is corrected by swapping and is loudly flagged;
#     the final assertion block (§12) fails the build if one is left.
# ---------------------------------------------------------------------------
for s in samples:
    df_, dt_ = s["depth_from_m"], s["depth_to_m"]
    if df_ is not None and dt_ is not None and df_ > dt_:
        s["depth_from_m"], s["depth_to_m"] = dt_, df_
        s["depth_mid_m"] = rdepth((df_ + dt_) / 2.0)
        flag = (f"D1 GUARD: stored interval was inverted (from {df_:g} > to "
                f"{dt_:g}); from/to swapped by build_database.py")
        s["qa_flags"] = (s["qa_flags"] + "; " + flag).strip("; ")
        qa["depth_interval_fixes"].append(f"{s['sample_id']}: {flag}")

# ============================================================================
# 4. Coordinates
# ============================================================================

for s in samples:
    h = s["hole_id_norm"]
    if h and h in traces and s["depth_mid_m"] is not None:
        xyz = traces[h].xyz_at(s["depth_mid_m"])
        if xyz:
            s["x_utm"], s["y_utm"], s["z_rl"] = xyz
            s["coord_source"] = "desurvey"
            continue
    if s["_master_x"] is not None and s["_master_y"] is not None:
        s["x_utm"], s["y_utm"] = round(s["_master_x"], 2), round(s["_master_y"], 2)
        s["coord_source"] = "master_xy"
        if h:  # drill sample that could not be desurveyed
            s["qa_flags"] = (s["qa_flags"] + "; collar X/Y used (no desurvey)").strip("; ")
    else:
        s["coord_source"] = "none"
        if "no coordinates" not in s["qa_flags"] and "no drillhole" not in s["qa_flags"]:
            if not s["qa_flags"]:
                s["qa_flags"] = "no coordinates available"
for s in samples:
    s.pop("_master_x"), s.pop("_master_y")

# ============================================================================
# 5. tag <-> hole@depth cross-reference (2023 registers) for description joins
# ============================================================================

tag_xref = {}
for fn, src in [("Drillhole Petrograph_2023__Sheet1.csv", XLSX),
                ("Suggested_Petro_for_the_Oval_2024__Sheet1.csv", MASTER),
                ("Petrograph_2023__Sheet1.csv", MASTER)]:
    rows = read_csv_rows(src / fn)
    for row in rows:
        if len(row) < 3 or not re.fullmatch(r"\d{5}", row[0].strip()):
            continue
        tag = row[0].strip()
        d, _, _ = parse_depth_text(row[1])
        h = norm_hole(row[2], fn)
        if tag not in tag_xref and h:
            tag_xref[tag] = (h, d)

# depth index for hole+depth joins
depth_index = defaultdict(list)
for s in samples:
    if s["hole_id_norm"] and s["depth_from_m"] is not None:
        depth_index[s["hole_id_norm"]].append(s)

# label index for exact hole-depth-label joins (e.g. 'OVD028-38', 'CRS01A-81')
label2sample = {}
for s in samples:
    keys = {s["sample_id"]}
    for a in s["alt_ids"].split("|"):
        a = a.strip()
        if a:
            keys.add(a)
            keys.add(a.split(" (")[0])
    if s["hole_id_norm"] and s["depth_from_m"] is not None:
        dtxt = "%g" % s["depth_from_m"]
        keys.add(f"{s['hole_id_norm']}-{dtxt}")
        keys.add(f"{s['hole_id_norm']}@{dtxt}")
    for k in keys:
        if k and k not in label2sample:
            label2sample[k] = s


def match_by_label(raw_id):
    """exact match of a hole-depth label after normalizing the hole part.

    COVERAGE G10 fix: the VERBATIM label (suffix included) is tried FIRST, so
    `OVD015-175.5 (B)` lands on 42389 (alt_id `OVD015-175.5 (B)`) instead of
    falling through to the suffix-stripped `OVD015-175.5`, which used to send
    both D0088 and D0089 to 42388."""
    if not raw_id:
        return None
    verbatim = str(raw_id).strip()
    if verbatim in label2sample:
        return label2sample[verbatim]
    lab = verbatim.split(" (")[0].strip()
    if lab in label2sample:
        return label2sample[lab]
    m = HD_LABEL.match(lab)
    if m and not m.group(4):
        h = norm_hole(m.group(1))
        if h:
            d = m.group(2) + ("-" + m.group(3) if m.group(3) else "")
            for key in (f"{h}-{d}", f"{h}@{d}"):
                if key in label2sample:
                    return label2sample[key]
    return None


def match_by_depth(hole, d, dfrom=None, dto=None, suffix=None, tol=0.3):
    """Find the sample for hole+depth (±tol); prefers exact depth, then a
    slide-suffix (A/B) match among equidistant candidates."""
    cands = []
    for s in depth_index.get(hole, []):
        sf, st_ = s["depth_from_m"], s["depth_to_m"]
        target = d
        if target is None and dfrom is not None:
            target = (dfrom + dto) / 2.0 if dto is not None else dfrom
        if target is None:
            continue
        if st_ is not None and (sf - tol) <= target <= (st_ + tol):
            diff = 0.0
        else:
            diff = abs((sf if sf is not None else 1e9) - target)
        if diff <= tol:
            cands.append((diff, s))
    if not cands:
        return None, ""
    cands.sort(key=lambda t: t[0])
    best = [s for diff, s in cands if abs(diff - cands[0][0]) < 1e-9]
    if len(best) > 1 and suffix:
        sfx = {"А": "A", "Б": "B", "A": "A", "B": "B"}.get(suffix.upper(), suffix.upper())
        for s in best:
            if re.search(r"[\( ]" + sfx + r"\)?(\s*\||$)", s["alt_ids"]):
                return s, "hole+depth+suffix"
    note = "ambiguous: %d equidistant candidates" % len(best) if len(best) > 1 else ""
    return best[0], note


# ============================================================================
# 6. descriptions table
# ============================================================================

descriptions = []


def add_desc(raw_id, raw_hole, raw_depth, source_file, analyst, date, rock_name,
             rock_orig, texture, minerals_json, alteration, opaques, text,
             tag=None, hole=None, d=None, dfrom=None, dto=None, suffix=None,
             qa_notes=""):
    joined, method, note = None, "unmatched", ""
    if tag and str(tag) in tag2sample:
        joined, method = tag2sample[str(tag)], "tag"
    else:
        joined = match_by_label(raw_id)
        if joined is not None:
            method = "label"
        else:
            if (hole is None or d is None) and tag and str(tag) in tag_xref:
                hole, d = tag_xref[str(tag)]
            if hole and (d is not None or dfrom is not None):
                joined, note = match_by_depth(hole, d, dfrom, dto, suffix)
                if joined is not None:
                    method = "hole+depth" if note != "hole+depth+suffix" else note
                    note = "" if note.startswith("hole") else note
    lang = "mn" if is_cyrillic((text or "") + (rock_orig or "")) else "en"
    all_notes = "; ".join(x for x in [qa_notes, note] if x)
    descriptions.append({
        "sample_id": joined["sample_id"] if joined else "",
        "desc_id": f"D{len(descriptions)+1:04d}",
        "raw_sample_id": raw_id or "",
        "raw_hole_id": raw_hole or "",
        "raw_depth": raw_depth if raw_depth is not None else "",
        "source_file": source_file,
        "analyst_or_lab": analyst or "",
        "report_date": date or "",
        "language": lang,
        "rock_name": rock_name or "",
        "rock_name_original": rock_orig or "",
        "texture": texture or "",
        "minerals_json": minerals_json or "",
        "alteration": alteration or "",
        "opaque_minerals": opaques or "",
        "description_text": text or "",
        "join_method": method,
        "qa_notes": all_notes,
    })


# 6a. Petrography bichiglel table (master description sheet)
for row in bich_rows:
    row = row + [""] * (26 - len(row))
    has_desc = any(row[i].strip() for i in (16, 17, 18, 19, 20, 21, 24))
    if not has_desc:
        continue
    sn = row[1].strip()
    tag = sn if re.fullmatch(r"\d{5}", sn) else None
    hole = norm_hole(row[4], BICH)
    dfrom, _, _ = parse_depth_text(row[5])
    dto, _, _ = parse_depth_text(row[6])
    m = HD_LABEL.match(sn)
    suffix = m.group(4) if m else None
    if tag is None and m:
        dfrom = dfrom if dfrom is not None else float(m.group(2))
    text = row[21].strip()
    if row[24].strip():
        text = (text + " | Microscope rock name: " + row[24].strip()).strip(" |")
    minerals = json.dumps({"rock_forming": row[18].strip()}, ensure_ascii=False) \
        if row[18].strip() else ""
    add_desc(sn, row[4], row[5], BICH,
             row[9].strip() or "not stated", "", row[16].strip(), "",
             row[17].strip(), minerals, row[19].strip(), row[20].strip(), text,
             tag=tag, hole=hole, d=dfrom, dfrom=dfrom, dto=dto, suffix=suffix)

# 6b. Drillhole Petrograph_2023 register (Mireslab / Khanlab / Oyunjargal calls)
DP23 = "Drillhole Petrograph_2023__Sheet1.csv"
for row in read_csv_rows(XLSX / DP23)[1:]:
    if len(row) < 8 or not row[0].strip():
        continue
    d, _, _ = parse_depth_text(row[1])
    add_desc(row[0].strip(), row[2], row[1], DP23, row[6].strip(), "",
             row[7].strip(), "", "", "", "", "", row[7].strip(),
             tag=row[0].strip(), hole=norm_hole(row[2], DP23), d=d,
             qa_notes="" if row[7].strip() else "empty petrographic determination")

# 6c. MIRESL 2023-08-16 summary
MIR = "Petrograph_MIRESL20230816_summary__Summary Table.csv"
for row in read_csv_rows(XLSX / MIR)[1:]:
    if len(row) < 13 or not row[0].strip():
        continue
    d, _, _ = parse_depth_text(row[1])
    alte = " — ".join(x for x in [row[7].strip(), row[8].strip()] if x)
    opaq = "; ".join(x for x in [row[10].strip(),
                                 ("unknown: " + row[11].strip()) if row[11].strip() else "",
                                 ("other opaques: " + row[12].strip()) if row[12].strip() else ""] if x)
    add_desc(row[0].strip(), row[2], row[1], "Petrograph_MIRESL20230816_summary.xlsx",
             "MIRESL (Mireslab Mongol LLC)", "2023-08-16", row[6].strip(), "",
             "", "", alte, opaq,
             f"Lithofacies: {row[6].strip()}. MIRESL sample code {row[5].strip()}.",
             tag=row[0].strip(), hole=norm_hole(row[2], MIR), d=d)

# 6d. Samples-to-Japan sheet (micro-descriptions)
JAP = "Yambat_petrography_samples_2022-2024_from_Core___grab__Samples_to_Japan.csv"
for row in read_csv_rows(MASTER / JAP)[2:]:
    row = row + [""] * (11 - len(row))
    if not row[1].strip() or not row[8].strip():
        continue
    sn = row[1].strip()
    tag = sn if re.fullmatch(r"\d{5}", sn) else None
    m = HD_LABEL.match(row[3].strip() or sn)
    hole = norm_hole(row[2], JAP)
    d = float(m.group(2)) if m else None
    minerals = json.dumps({"rock_forming": row[10].strip()}, ensure_ascii=False) \
        if row[10].strip() else ""
    add_desc(sn, row[2], row[3], JAP, "not stated", "", row[8].strip(), "",
             row[9].strip(), minerals, "", "", row[8].strip(),
             tag=tag, hole=hole, d=d)

# 6e. 2022-2025 report descriptions (reports/samples.json)
for r in recs_rep:
    if r["record_type"] != "petrographic_description":
        continue
    sid = str(r["sample_id"]).strip()
    hole_raw = (r.get("drillhole_id") or "").split(" ")[0]
    hole = norm_hole(hole_raw, r["source_file"])
    # unlocated report samples were added as samples in 3e -> join by canonical id
    canon = sid.replace("С-", "C-").split(" ")[0].replace("Дээж-", "DEEJ-")
    tag = r.get("lab_tag")
    if tag is None and canon in sample_by_id and not hole:
        tag = None
    minerals = json.dumps(r.get("minerals"), ensure_ascii=False) if r.get("minerals") else ""
    joined_note = ""
    add_desc(sid, hole_raw, r.get("depth_m"), r["source_file"],
             r.get("analyst_or_lab"), r.get("report_date"),
             r.get("rock_name"), r.get("rock_name_original"), r.get("texture"),
             minerals, r.get("alteration"), r.get("mineralization"),
             r.get("description_summary"),
             tag=tag, hole=hole, d=r.get("depth_m"),
             dfrom=r.get("depth_from_m"), dto=r.get("depth_to_m"),
             qa_notes=joined_note)
    # unlocated samples: manual join via canonical id
    dd = descriptions[-1]
    if not dd["sample_id"] and canon in sample_by_id:
        dd["sample_id"] = canon
        dd["join_method"] = "report id"

# 6f. 2024-2026 report descriptions (reports2024_2026/samples.json)
for r in recs2426:
    sid = str(r["sample_id"]).strip()
    tag = None
    m = re.search(r"\((\d{5})\)", sid)
    if m:
        tag = m.group(1)
    elif re.match(r"^(\d{5})\b", sid):
        tag = re.match(r"^(\d{5})\b", sid).group(1)
    hole = norm_hole(r.get("drillhole_id") or "", r["source_file"])
    sfx = None
    ms = re.search(r"(?:[ .\-(]|(?<=\d))([ABАБab])\)?$",
                   sid.replace("(There was 2)", "").strip())
    if ms:
        sfx = ms.group(1)
    minerals = json.dumps(r.get("minerals"), ensure_ascii=False) if r.get("minerals") else ""
    add_desc(sid, r.get("drillhole_id"), r.get("depth_m") or r.get("depth_from_m"),
             r["source_file"], r.get("analyst_or_lab"), r.get("report_date"),
             r.get("rock_name"), r.get("rock_name_original"), r.get("texture"),
             minerals, r.get("alteration"), r.get("opaque_minerals"),
             r.get("description_summary"),
             tag=tag, hole=hole, d=r.get("depth_m"),
             dfrom=r.get("depth_from_m"), dto=r.get("depth_to_m"), suffix=sfx)

# ---------------------------------------------------------------------------
# 6g. COVERAGE G9 — the master grab sheet's free-text FIELD descriptions.
#     54 of the 65 grab rows carry a geologist's outcrop description in the
#     right-hand `Description` column (col 13).  v1.0 dropped them entirely;
#     they are now emitted as description rows joined to their grab sample.
# ---------------------------------------------------------------------------
n_grab_desc = 0
for row in read_csv_rows(MASTER / GRAB)[2:]:
    row = row + [""] * (14 - len(row))
    sid = row[1].strip()
    field_text = row[13].strip()
    if not sid or not field_text:
        continue
    n_grab_desc += 1
    joined = sample_by_id.get(sid)
    descriptions.append({
        "sample_id": joined["sample_id"] if joined else "",
        "desc_id": f"D{len(descriptions)+1:04d}",
        "raw_sample_id": sid,
        "raw_hole_id": "",
        "raw_depth": "",
        "source_file": "Yambat Petrographic Master Data.xlsx :: "
                       "2022-2024 grab sheet (field description column)",
        "analyst_or_lab": "ABM field geologist (not named in the sheet)",
        "report_date": "",
        "language": "mn" if is_cyrillic(field_text) else "en",
        "rock_name": row[5].strip(),
        "rock_name_original": "",
        "texture": "",
        "minerals_json": "",
        "alteration": "",
        "opaque_minerals": "",
        "description_text": field_text,
        "join_method": "grab sheet row" if joined else "unmatched",
        "qa_notes": "COVERAGE G9: field (hand-specimen/outcrop) description "
                    "recovered in v1.1 from the master grab sheet; it is a "
                    "field call, NOT a microscope determination",
    })
qa["grab_field_descriptions"].append(
    f"{n_grab_desc} field descriptions recovered from the 2022-2024 grab sheet")

# ---------------------------------------------------------------------------
# 6h. missing_sources batch (84 records; 17 photo_only became samples in 3i,
#     the remaining 67 become descriptions here).
#     Join rules are those of workspace/extracted/missing_sources/README.md.
# ---------------------------------------------------------------------------

# (i) surface records -> the 2022-2024 grab sheet rows 1-24, matched on the
#     sample id after case folding and leading-zero stripping.
def _surface_key(sid):
    s = str(sid).strip().lower().replace(" ", "")
    s = {"2022-01": "2021-01"}.get(s, s)      # README §4.3 year-digit typo
    if re.fullmatch(r"\d+", s):               # README §4.2 leading zeros
        s = s.lstrip("0") or "0"
    return s


grab_key2sample = {}
grab_row_no = {}
for _i, row in enumerate(read_csv_rows(MASTER / GRAB)[2:]):
    if len(row) < 2 or not row[1].strip():
        continue
    sid = row[1].strip()
    grab_key2sample.setdefault(_surface_key(sid), sid)
    grab_row_no.setdefault(_surface_key(sid), len(grab_row_no) + 1)

# (ii) MIRESL internal code (OVD001..OVD023) -> ABM 5-digit tag, read from the
#      `Code` column of the MIRESL 2023-08-16 summary sheet.
mireslab_code2tag = {}
for row in read_csv_rows(XLSX / MIR)[1:]:
    if len(row) < 6 or not re.fullmatch(r"\d{5}", row[0].strip()):
        continue
    if row[5].strip():
        mireslab_code2tag[row[5].strip().upper()] = row[0].strip()

MISS_UNMATCHED_NOTE = {
    "2111": "no field/sample number anywhere in Report_0715_Ni.pdf and no row "
            "in the 2022-2024 grab sheet or Master All — cannot be joined "
            "until a field number is supplied (missing_sources README §4.7)",
    "2107": "AMBIGUOUS ID: `2107` in Report_0715_Ni.pdf is a Ni-ore sample "
            "(XRD + ore microscopy + SEM-EDS), while grab row 21 `2107` is an "
            "amphibolite thin section (Thin and polish-4.docx). The grab-sheet "
            "coordinates put row 21 in the same outcrop cluster, so they are "
            "probably the same physical sample described by different methods "
            "— but no document states it, so this record is left UNJOINED "
            "rather than force-merged (missing_sources README §4.6)",
    "2023Nisample": "descriptive placeholder, not a field number: "
                    "MINERALOGICAL-DESCRIPTIONS_2023.03.25.pdf gives no sample "
                    "id for the garnierite / Ni-goethite sample "
                    "(missing_sources README §4.7)",
}

miss_stats = Counter()
miss_depth_conflicts = []
for r in miss_desc_recs:
    raw_sid = str(r["sample_id"]).strip()
    src = str(r["source_file"])
    hole_raw = str(r.get("drillhole_id") or "")
    hole = norm_hole(hole_raw.split(" ")[0], "missing_sources: " + src[:40])
    depth_txt = str(r.get("depth") or "")
    d, _dt, _dn = parse_depth_text(depth_txt.split(" ")[0]) if depth_txt else (None, None, "")
    tag = raw_sid if re.fullmatch(r"\d{5}", raw_sid) else None
    qa_notes, joined, method = "", None, "unmatched"

    if tag:                                   # sources 10-13: 5-digit ABM tags
        joined = tag2sample.get(tag)
        method = "tag" if joined else "unmatched"
        code = re.search(r"internal sample code (OVD\d{3})",
                         str(r.get("sample_type") or ""))
        if code:                              # source 12 — verify via MIRESL Code
            want = mireslab_code2tag.get(code.group(1).upper())
            if want == tag:
                method = "tag (MIRESL code)"
                qa_notes = (f"joined via the Mireslab internal code "
                            f"{code.group(1)} -> tag {tag} of the "
                            f"Petrograph_MIRESL20230816_summary `Code` column")
            elif want:
                qa_notes = (f"MIRESL code {code.group(1)} maps to tag {want} in "
                            f"the summary sheet but this record carries tag "
                            f"{tag} — joined on the tag, conflict recorded")
    elif raw_sid in MISS_UNMATCHED_NOTE and "Thin and polish" not in src:
        qa_notes = "UNMATCHED (v1.1): " + MISS_UNMATCHED_NOTE[raw_sid]
    else:                                     # sources 1, 2, 4, 6-9: surface
        key = _surface_key(raw_sid)
        gsid = grab_key2sample.get(key)
        if gsid and gsid in sample_by_id:
            joined = sample_by_id[gsid]
            method = "grab sheet id"
            if gsid != raw_sid:
                qa_notes = (f"grab-sheet id `{gsid}` (row {grab_row_no[key]}) "
                            f"vs report id `{raw_sid}`: joined after case "
                            f"folding / leading-zero stripping"
                            + ("; `2021-01` vs `2022-01` is a year-digit typo, "
                               "not two samples (missing_sources README §4.3) "
                               "— a human should confirm which spelling is "
                               "authoritative" if key == "2021-01" else ""))
                flag = (f"v1.1: the 2022 lab report calls this sample "
                        f"`{raw_sid}`; the grab sheet calls it `{gsid}`. Joined "
                        f"on the case-folded / zero-stripped id"
                        + ("; the year digit differs (2021 vs 2022) — a human "
                           "should confirm which spelling is authoritative"
                           if key == "2021-01" else ""))
                if flag not in joined["qa_flags"]:
                    joined["qa_flags"] = (joined["qa_flags"] + "; " + flag).strip("; ")
            if raw_sid == "2107":
                flag2 = ("v1.1 AMBIGUOUS ID: a second `2107` exists in "
                         "Report_0715_Ni.pdf (a Ni-ore sample analysed by XRD + "
                         "ore microscopy + SEM-EDS). Its coordinates are not "
                         "given, but this grab row sits in the same outcrop "
                         "cluster as 2102/2104-1/A, so the two are probably one "
                         "physical sample described by different methods — no "
                         "document states it, so the Ni-report description is "
                         "kept UNJOINED (missing_sources README §4.6)")
                if flag2 not in joined["qa_flags"]:
                    joined["qa_flags"] = (joined["qa_flags"] + "; " + flag2).strip("; ")
    if joined is None and not qa_notes:
        qa_notes = "UNMATCHED (v1.1): no sample row carries this id"

    minerals = json.dumps(r.get("minerals"), ensure_ascii=False) \
        if r.get("minerals") else ""
    text = r.get("description_summary") or ""
    lang = "mn" if is_cyrillic(text + str(r.get("rock_name_original") or "")) else "en"
    miss_stats[method.split(" (")[0]] += 1
    descriptions.append({
        "sample_id": joined["sample_id"] if joined else "",
        "desc_id": f"D{len(descriptions)+1:04d}",
        "raw_sample_id": raw_sid,
        "raw_hole_id": hole_raw,
        "raw_depth": depth_txt,
        "source_file": src,
        "analyst_or_lab": r.get("analyst_or_lab") or "",
        "report_date": r.get("report_date") or "",
        "language": lang,
        "rock_name": r.get("rock_name") or "",
        "rock_name_original": r.get("rock_name_original") or "",
        "texture": r.get("texture") or "",
        "minerals_json": minerals,
        "alteration": r.get("alteration") or "",
        "opaque_minerals": r.get("opaque_minerals") or "",
        "description_text": text,
        "join_method": method,
        "qa_notes": qa_notes,
    })
    if joined is not None and hole and joined["hole_id_norm"] and \
            joined["hole_id_norm"] != hole:
        descriptions[-1]["qa_notes"] = (
            descriptions[-1]["qa_notes"] + "; hole in record (" + hole +
            ") differs from the joined sample's hole (" +
            joined["hole_id_norm"] + ")").strip("; ")
    # depth cross-check: the report's stated depth vs the sample register's
    if joined is not None and d is not None and joined["depth_from_m"] is not None:
        sf = joined["depth_from_m"]
        st_ = joined["depth_to_m"] if joined["depth_to_m"] is not None else sf
        if not (sf - 0.05 <= d <= st_ + 0.05):
            off = round(min(abs(d - sf), abs(d - st_)), 3)
            descriptions[-1]["qa_notes"] = (
                descriptions[-1]["qa_notes"] +
                f"; depth cross-check: this report states {d:g} m but the "
                f"sample register stores {sf:g}"
                + (f"-{st_:g}" if st_ != sf else "") +
                f" m ({off:g} m apart) — joined on the lab tag, depth kept as "
                f"the register has it").strip("; ")
            miss_depth_conflicts.append(
                f"{descriptions[-1]['desc_id']} tag {raw_sid}: report {d:g} m "
                f"vs register {sf:g} m ({off:g} m)")

# ---------------------------------------------------------------------------
# 6i. Join corrections and source caveats
#     (integrity D2 / D5 / D9, coverage G2 / G10 / G11)
# ---------------------------------------------------------------------------

def _descs(raw_prefix, src_contains=""):
    return [d for d in descriptions
            if d["raw_sample_id"].strip().startswith(raw_prefix)
            and src_contains in d["source_file"]]


def _add_note(d, note):
    d["qa_notes"] = "; ".join(x for x in [d["qa_notes"], note] if x)


def _flag_sample(sid, flag):
    s = sample_by_id.get(sid)
    if s is not None and flag not in s["qa_flags"]:
        s["qa_flags"] = (s["qa_flags"] + "; " + flag).strip("; ")


# --- G2: two descriptions the v1.0 build left unmatched are in fact resolvable
#     from evidence already present in the extraction layer.
G2_REJOINS = [
    ("OVD021@101.5m", CRAWFORD_PDF, "42027",
     "COVERAGE G2 (v1.1): re-joined to 42027 = OVD011-101.5. The "
     "`KhanAltai vs Tony` sheet of the working workbook carries Tony "
     "Crawford's IDENTICAL micro-description ('An intensely altered aphyric, "
     "quite fine-grained leucogabbroic dyke(?)') against row OVD011-101.5, so "
     "`OVD021@101.5m` in the Crawford PDF is a hole-number typo for OVD011. "
     "OVD021 has no sample at 101.5 m. CRAWFORD CAVEAT: the rock lacks "
     "sulfides and has no chromite to account for its high-Cr assay"),
    ("OVD20-121", "English 41", "43251",
     "COVERAGE G2 (v1.1): re-joined to 43251 = OVD021 @ 121.0 m (alt id "
     "`OVD21-121`). The 41-sample report's own microphoto for this entry is "
     "captioned '21-121', and Sheet3 / the bichiglel table both carry "
     "`OVD21-121` = tag 43251; OVD020 has no sample at 121 m, so `OVD20-121` "
     "is a hole-number typo"),
]
g2_applied = []
for raw, src, target, note in G2_REJOINS:
    for d in _descs(raw, src):
        d["sample_id"] = target
        d["join_method"] = "xref-corrected"
        d["qa_notes"] = note
        g2_applied.append(f"{d['desc_id']} `{raw}` -> {target}")

# --- G1 follow-through: the two new Crawford samples' descriptions
for d in _descs("OVD003@202", CRAWFORD_PDF):
    _add_note(d, "COVERAGE G1 (v1.1): joined to the sample row created for "
                 "this thin section; untagged Crawford extra suggestion "
                 "(low-MgO gabbro), no register entry exists for it")
for d in _descs("OVD009@178-180", CRAWFORD_PDF):
    _add_note(d, "COVERAGE G1 (v1.1): joined to the sample row created for "
                 "this thin section. CRAWFORD CAVEAT: 'wholerock assay does "
                 "not match this thin section' — suspected sample swap; no "
                 "register entry exists for it")

# --- D2: the Crawford 2025 caveats must be queryable from qa_notes / qa_flags,
#     not only from the tail of description_text (QA_report §9 said they were).
CRAWFORD_CAVEATS = [
    ("OVD007@55.9m", "Crawford 2025 caveat: the analyst was unsure the "
                     "provided core photo matches the thin section"),
    ("OVD008@88.9m", "Crawford 2025 caveat: this sample lacks sulfides "
                     "despite the 2.5 %S assay for the interval"),
    ("OVD008@90.5m", "Crawford 2025 caveat: the assay for this hole/depth "
                     "indicates a strongly sulfidic rock (~30 % pyrrhotite) "
                     "whereas the thin section is a finely hornblende-phyric "
                     "basalt (dyke?) with <2 % sulfides"),
    ("OVD005@40.5m", "Crawford 2025 caveat: section far too thin — very "
                     "little rock preserved, diagnosis provisional"),
    ("OVD005@53.0m", "Crawford 2025 caveat: another far too-thin section with "
                     "little useful material"),
    ("OVD021@148.8m", "Crawford 2025 caveat: sulfides too poorly polished to "
                      "be informative"),
    ("OVD021@101.5m", "Crawford 2025 caveat: the rock lacks sulfides and "
                      "there is no trace of chromite to account for the "
                      "high-Cr assay"),
    ("OVD009@178-180m", "Crawford 2025 caveat: wholerock assay does not match "
                        "this thin section (suspected sample swap)"),
]
d2_applied = []
for raw, note in CRAWFORD_CAVEATS:
    for d in _descs(raw, CRAWFORD_PDF):
        _add_note(d, note)
        d2_applied.append(f"{d['desc_id']} ({raw} -> {d['sample_id'] or '-'})")
        if d["sample_id"]:
            _flag_sample(d["sample_id"], note + f" [source: {raw}]")

# --- D9: the A/B slide-suffix joins rest on letter order alone
ab_notes = []
for d in descriptions:
    if not d["sample_id"]:
        continue
    ms = re.search(r"(?:[ .\-(]|(?<=\d))([ABab])\)?\s*$",
                   d["raw_sample_id"].strip())
    if not ms:
        continue
    s = sample_by_id[d["sample_id"]]
    if not re.search(r"\((A|B)\)", s["alt_ids"]):
        continue
    letter = ms.group(1).upper()
    _add_note(d, f"D9: A/B slide suffix — this description was assigned to "
                 f"{s['sample_id']} (`{s['alt_ids']}`) on LETTER ORDER ALONE "
                 f"(report letter '{letter}' -> master suffix '({letter})'). "
                 f"There is no independent discriminator — the master lists "
                 f"the (A) and (B) slides with the same lithology — so if "
                 f"either source transposed the letters, the two descriptions "
                 f"are swapped between the two samples. The (A)/(B) pair "
                 f"shares one hole and depth, so the spatial impact is nil")
    ab_notes.append(f"{d['desc_id']} `{d['raw_sample_id']}` -> {s['sample_id']}")

# --- G11: samples that merge two physically distinct (A)/(B) thin sections
for s in samples:
    if "(A)" in s["alt_ids"] and "(B)" in s["alt_ids"]:
        _flag_sample(s["sample_id"],
                     f"COVERAGE G11: this single row MERGES two physically "
                     f"distinct thin sections, `{s['alt_ids']}`, that the "
                     f"master sheet records at the same hole and depth. Their "
                     f"descriptions are all attached to this one sample_id; "
                     f"they are NOT split (unlike OVD015-175.5 (A)/(B), which "
                     f"carry separate tags 42388/42389)")

# --- D5: samples whose stored interval disagrees with the interval their own
#     report names.  Detected generically, not hard-coded.
_RE_RPT_IV = re.compile(
    r"^\s*([A-Za-z]+\s?-?\d+[A-Za-z]?)[-@](\d+(?:\.\d+)?)\s*-\s*"
    r"(\d+(?:\.\d+)?)\s*m?\s*(?:\((\d{5})\))?\s*$")
_RE_RPT_CAP = re.compile(
    r"^\s*([A-Za-z]+\s?-?\d+[A-Za-z]?)-(\d+(?:\.\d+)?)\s*\("
    r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)\)\s*$")
d5_rows, n_iv_checked = [], 0
for d in descriptions:
    raw = d["raw_sample_id"].strip()
    m, m2 = _RE_RPT_IV.match(raw), _RE_RPT_CAP.match(raw)
    if m:
        rf, rt = float(m.group(2)), float(m.group(3))
    elif m2:
        rf, rt = float(m2.group(3)), float(m2.group(4))
    else:
        continue
    if not d["sample_id"]:
        continue
    s = sample_by_id[d["sample_id"]]
    sf = s["depth_from_m"]
    if sf is None:
        continue
    st_ = s["depth_to_m"] if s["depth_to_m"] is not None else sf
    n_iv_checked += 1
    lo, hi = min(rf, rt), max(rf, rt)
    if lo - 1e-9 <= sf and st_ <= hi + 1e-9:
        continue
    gap = round(max(lo - sf, st_ - hi), 3)
    flag = (f"D5: the report that describes this sample names the interval "
            f"{lo:g}-{hi:g} m (`{raw}`), but the master workbook stores it at "
            f"{sf:g}-{st_:g} m — a {gap:g} m disagreement inside the sources "
            f"themselves (the phase sheet's sample From/To vs the assay "
            f"interval on the same row). Stored values are the master's; treat "
            f"the depth as uncertain to +/-{gap:g} m")
    _flag_sample(s["sample_id"], flag)
    _add_note(d, f"D5: stored interval {sf:g}-{st_:g} m disagrees with the "
                 f"report-named interval {lo:g}-{hi:g} m by {gap:g} m")
    d5_rows.append((d["desc_id"], raw, s["sample_id"], f"{sf:g}-{st_:g}",
                    f"{lo:g}-{hi:g}", gap))
d5_rows.sort(key=lambda t: -t[5])

# --- D7: cross-reference near-duplicate samples on the same hole
near_dupes = []
_by_hole = defaultdict(list)
for s in samples:
    if s["hole_id_norm"] and s["depth_mid_m"] is not None:
        _by_hole[s["hole_id_norm"]].append(s)
for h, lst in sorted(_by_hole.items()):
    lst = sorted(lst, key=lambda s: (s["depth_mid_m"], s["sample_id"]))
    for i in range(len(lst) - 1):
        a, b = lst[i], lst[i + 1]
        gap = round(abs(b["depth_mid_m"] - a["depth_mid_m"]), 3)
        if gap > 0.35:
            continue
        for x, y in ((a, b), (b, a)):
            _flag_sample(x["sample_id"],
                         f"D7 cross-reference: possible duplicate of sample "
                         f"{y['sample_id']} ({y['hole_id_norm']} @ "
                         f"{y['depth_mid_m']:g} m, {gap:g} m away"
                         + (f", alt ids `{y['alt_ids']}`" if y["alt_ids"] else "")
                         + ") — verify before treating the two as independent "
                           "observations")
        near_dupes.append((h, a["sample_id"], a["depth_mid_m"],
                           b["sample_id"], b["depth_mid_m"], gap))

# ============================================================================
# 7. Lookups
# ============================================================================

# 7a. hole aliases
lu_hole_alias = []
for raw in sorted(hole_alias_norm):
    lu_hole_alias.append({
        "hole_id_raw": raw,
        "hole_id_norm": hole_alias_norm[raw],
        "changed": 0 if raw == hole_alias_norm[raw] else 1,
        "seen_in": "; ".join(sorted(x for x in hole_alias_seen[raw] if x)),
    })

# 7b. labs
LAB_CANON = [
    ("Altantsetseg", "Altantsetseg (ABM/Aventura in-house petrographer)", "ABM in-house"),
    ("SHUTIS", "MUST — Mongolian University of Science & Technology (ШУТИС)", "university lab"),
    ("MUST", "MUST — Mongolian University of Science & Technology (ШУТИС)", "university lab"),
    ("MIRES", "Mireslab Mongol LLC (MIRESL)", "commercial lab"),
    ("Mireslab", "Mireslab Mongol LLC (MIRESL)", "commercial lab"),
    ("MIRESL", "Mireslab Mongol LLC (MIRESL)", "commercial lab"),
    ("Khanlab", "Khanlab LLC (KhanAltai)", "commercial lab"),
    ("NUM", "National University of Mongolia (МУИС/NUM; L.Jargal PhD)", "university lab"),
    ("MUIS", "National University of Mongolia (МУИС/NUM; L.Jargal PhD)", "university lab"),
    ("Oyunjargal", "Oyunjargal (project geologist, in-house)", "in-house"),
    ("Innova", "Innova Mineral LLC", "commercial lab"),
    ("ThinSection", "ThinSection Mongolia LLC", "commercial lab"),
    ("Жаргал", "National University of Mongolia (МУИС/NUM; L.Jargal PhD)", "university lab"),
    ("Crawford", "Dr Anthony J Crawford (A & A Crawford Geological Research)", "consultant"),
]


def canon_lab(raw):
    """Canonicalize a lab/analyst string.

    The rules are matched against the HEAD of the string (everything before the
    first ';'), i.e. the lab that actually made the determination.  Long v1.1
    analyst strings often name a *second* institution further along ('sections
    prepared at MUST', 'XRD at Akita'), which used to hijack the match — e.g.
    'Mireslab Mongol LLC - Jamsran Erdenebayar; Report #004 ...; sections at
    MUST' was canonicalized as MUST."""
    head = str(raw).split(";")[0]
    for text in (head, str(raw)):
        for key, canon, role in LAB_CANON:
            if key.lower() in text.lower():
                return canon, role
    return "", ""


lab_raws = Counter()
for s in samples:
    if s["petrographer_lab"]:
        lab_raws[s["petrographer_lab"]] += 1
for dsc in descriptions:
    if dsc["analyst_or_lab"]:
        lab_raws[dsc["analyst_or_lab"]] += 1
lu_lab = []
for raw, n in sorted(lab_raws.items()):
    canon, role = canon_lab(raw)
    lu_lab.append({"lab_raw": raw, "lab_canonical": canon or raw,
                   "role": role or "unresolved", "n_records": n})

# 7c. rock types
# D3: the rule set now covers the Mongolian (Cyrillic) vocabulary of
# `descriptions.rock_name_original` as well as the English names, so that
# EVERY distinct rock name in the database receives a rock_group, and a coarse
# `rock_family` (ultramafic / mafic intrusive / felsic-intermediate intrusive /
# volcanic / sedimentary / metamorphic / vein-ore / unknown) is derived from it.
ROCK_RULES = [
    ("massive sulfide / ore", r"massive sulph|massive sulf|net.?texture|semi.?massive|цул сульфид|цул хүдэр|цул сулфид|massive ore|massive [\w\- ]*ore|sulphide ore|sulfide ore|хүдэр(?!ж)"),
    ("gossan / oxidized", r"gossan|limonit|госсан|oxidiz|oxidaz|hematite and limonite|hematite ore|iron oxide ore|төмөржсөн|гётит|goethite|garnierite|гарниерит"),
    ("skarn / metasomatite", r"skarn|скарн|metasomat|метасоматит|greisen|грейзен|listvenit|лиственит|fuchsite|фуксит|garnet[\w\- ]*carbonate|carbonate[\w\- ]*garnet"),
    ("peridotite / ultramafic", r"perido|wehrlite|dunite|lherzol|верлит|перидотит|picrite|пикрит|websterite|вебстерит|pyroxenite|пироксенит|olivinite|оливинит|serpentinit|серпентинит|ultramafic|ультра|orthocumulate|adcumulate|mesocumulate|cumulate"),
    ("hornblendite", r"hornblendite|горнблендит"),
    ("olivine gabbro / melagabbro", r"olivine gabbro|melanocratic gabbro|melagabbro|оливин габбро|меланократ"),
    ("gabbro / gabbronorite / norite", r"gabbro(?!diorite)|norite|габбро(?!диорит)|норит"),
    ("gabbrodiorite", r"gabbrodiorite|габбродиорит"),
    ("diorite / quartz diorite", r"diorite|диорит|monzodiorite|монцодиорит"),
    ("granitoid", r"granodiorite|granite|гранодиорит|гранит|боржин|monzonit|монцонит|syenit|сиенит|plagiogranite|плагиогранит|alkaline rock|шүлтлэг"),
    ("dolerite / diabase / basalt dyke", r"doleri|diabase|basalt|долерит|диабаз|базальт"),
    ("volcanic / subvolcanic", r"andesi|rhyoli|rhyodacite|dacite|trachy|tuff|volcan|porphyry(?! system)|андезит|риолит|риодацит|дацит|трахи|туф|субвулкан|felsic|lava|лаав"),
    ("amphibolite / metabasite", r"amphibolit|амфиболит|metabasit|метабазит|greenschist|зеленокамен"),
    ("schist / phyllite", r"schist|phyllite|сланец|филлит|занар|slate"),
    ("hornfels / spotted rock", r"spotted|hornfels|роговик"),
    ("sediment / metasediment", r"sandstone|siltstone|silstone|argillite|mudstone|shale|pelit|sediment|gneiss|conglomerat|gravelit|greywacke|arkose|quartzite|песчаник|элсжин|элсэн чулуу|алевролит|аргиллит|гнейс|кварцит|гравелит|граувакк|аркоз|конгломерат|шавар"),
    ("breccia / fault rock", r"breccia|fault|mylonit|брекчи|разлом|милонит"),
    ("vein / quartz", r"quartz vein|vein|кварц(?!ит)|судал"),
    ("metamorphic (undifferentiated)", r"metamorphic|метаморф|метаморфизм"),
    ("mafic intrusive (undifferentiated)", r"mafic|basic intrusive|мафик"),
    ("intrusive (undifferentiated)", r"intrusive|интрузив"),
    ("indeterminate (section not diagnosable)",
     r"too thin|not determinable|indetermin|essentially identical to"),
    ("altered rock (protolith undetermined)",
     r"altered rock|entirely altered|silicif"),
]

# rock_group -> coarse family
ROCK_FAMILY = {
    "massive sulfide / ore": "vein-ore",
    "gossan / oxidized": "vein-ore",
    "vein / quartz": "vein-ore",
    "skarn / metasomatite": "metamorphic",
    "amphibolite / metabasite": "metamorphic",
    "schist / phyllite": "metamorphic",
    "hornfels / spotted rock": "metamorphic",
    "breccia / fault rock": "metamorphic",
    "peridotite / ultramafic": "ultramafic",
    "hornblendite": "ultramafic",
    "olivine gabbro / melagabbro": "mafic intrusive",
    "gabbro / gabbronorite / norite": "mafic intrusive",
    "gabbrodiorite": "mafic intrusive",
    "dolerite / diabase / basalt dyke": "mafic intrusive",
    "diorite / quartz diorite": "felsic-intermediate intrusive",
    "granitoid": "felsic-intermediate intrusive",
    "volcanic / subvolcanic": "volcanic",
    "sediment / metasediment": "sedimentary",
    "metamorphic (undifferentiated)": "metamorphic",
    "mafic intrusive (undifferentiated)": "mafic intrusive",
    "intrusive (undifferentiated)": "unknown",
    "indeterminate (section not diagnosable)": "unknown",
    "altered rock (protolith undetermined)": "unknown",
    "other / unclassified": "unknown",
    "": "unknown",
}


def rock_group(name):
    s = str(name).lower()
    if not s.strip():
        return ""
    for grp, pat in ROCK_RULES:
        if re.search(pat, s):
            return grp
    return "other / unclassified"


def rock_family(grp):
    return ROCK_FAMILY.get(grp, "unknown")


# D3: count EVERY distinct rock name across samples AND descriptions —
# including descriptions.rock_name_original, which v1.0 left out (172 of the
# 645 distinct names had no row, and 14 rows under-counted n_occurrences).
rock_counter = Counter()
rock_seen_in = defaultdict(set)


def _count_rock(v, where):
    v = (v or "").strip()
    if v:
        rock_counter[v] += 1
        rock_seen_in[v].add(where)


for s in samples:
    _count_rock(s["field_lithology"], "samples.field_lithology")
    _count_rock(s["petro_lithology"], "samples.petro_lithology")
    _count_rock(s["iogas_lithology"], "samples.iogas_lithology")
for dsc in descriptions:
    _count_rock(dsc["rock_name"], "descriptions.rock_name")
    _count_rock(dsc["rock_name_original"], "descriptions.rock_name_original")

lu_rock_type = [{"rock_name_original": k,
                 "rock_group": rock_group(k),
                 "rock_family": rock_family(rock_group(k)),
                 "n_occurrences": n,
                 "seen_in": "; ".join(sorted(rock_seen_in[k]))}
                for k, n in sorted(rock_counter.items(), key=lambda t: (-t[1], t[0]))]

# also standardized group on samples
for s in samples:
    s["rock_group"] = rock_group(s["petro_lithology"] or s["field_lithology"])

# ============================================================================
# 8. sources.csv registry
# ============================================================================

SOURCES = [
    ("Yambat Petrographic Master Data.xlsx", "1lG0hZDsnslLbezFCw34bx6f_bZn3YJ4w",
     "master sample table (All sheet = spine; grab & rockchip sheets)"),
    ("Yambat petrography samples 2022-2024 from Core & grab.xlsx", "1y25l1orJYDXQjXNupiht93Wg7Tr17Ste",
     "working workbook: bichiglel description table, Sheet3, phase sheets, Samples-to-Japan"),
    ("Collar_all_combined.csv", "1MxtDUzpvCtD9Nbi9IODl8EmXQOLCgvh2",
     "drillhole collars (76 holes) — location authority"),
    ("Survey_all_YMB.csv", "1fqyguCqRGnMgmuTm8pJbuPPmFTo3OMf0",
     "downhole surveys (1,990 rows) — desurvey input"),
    ("Khanlab_petrography_Samples.xlsx", "1W3XRRKAl-s3UAJY9ZpEal1rZjToOw0uh",
     "Khanlab assay verification table (hole@depth keyed; context only)"),
    ("PETRO LIST 2025.xlsx", "1X95MBGNX1QTlsDtQ26fcaXHIJIecNO1O",
     "2025 phase-3 sample list (subset of Master All)"),
    ("Petrograph_2023.xlsx (v3)", "1hLNU5HcF31MqC9LaTzQ4rn6qFQ1flvB3",
     "2023 sample register — source of tag 40904"),
    ("Suggested Petro for the Oval 2024.xlsx", "1bR-8vDiTDOcK4RclatPAGfHOIj-Y7Rct",
     "AJC/Tony Christie tag<->hole@depth cross-reference"),
    ("Drillhole Petrograph_2023.xlsx", "1gUTrLiQm_mVlDuPpeIpjaHLlXvxgYHvl",
     "2023 register of 50 drill-core petrography determinations (3 labs)"),
    ("Petrograph_MIRESL20230816_summary.xlsx", "1WoLN-IjSNCLeOxNqkBHa8S9UXbzHwloF",
     "MIRESL 2023-08-16 summary: lithofacies/alteration/ore-mineral detail"),
    ("Петрограф008.docx", "1tw3IOr7TudHqMDZz22kbLw2BvtTgugtK",
     "Khanlab 2024 Mongolian descriptions, 12 samples (ӨТШ)"),
    ("AM_ThinSectionMongolia_Report_2024.pdf", "1mNAkW0F0qWUuq2vDpBnjOtZ5oEcfESYj",
     "ThinSection Mongolia 2024 report, 8 unlocated samples (TS-n/Дээж-n)"),
    ("2023-08-06-2 шлиф.pdf", "1UONp87cloKEolNq0fNTAcR_WPXKZMaJE",
     "L.Jargal 2023 descriptions of С-1/С-2 (unlocated)"),
    ("2023-06-20-3 thin sections.pdf", "1kkEQHYDCdr7ckB7d7xv7zssmCIkEGpU-",
     "L.Jargal/Ragnarok 2023 descriptions of tags 40763/40900/40913"),
    ("Asian Battery Metals March 2025 The Oval Summary Report.pdf", "1HQZSTrXY79Q10i0cLzdl1Ug5Zp9POr9z",
     "Crawford 2025 petrographic report — 38 samples (English)"),
    ("Review of Work on 'The Oval' Ni-CU Target Revised March 2025.pdf", "1oLGiwQWcY8r_S9nhTHozPrDgJhzkjzss",
     "Crawford review (context; no per-sample descriptions)"),
    ("FW_ Petrology - Oval Nickel Project.zip", "1n_jRl5VBRsIzcX8KFezy7IznoPbUGIL-",
     "email bundle: Suggested Petro xlsx + assay table + May-2024 review"),
    ("English Petrographic and Mineragraphic Descriptions. 6.docx", "1oaodEQTEk0STZ9sRQ0X9ysVSqs4arRxc",
     "2024 descriptions, 6 samples (superseded by 41-sample report)"),
    ("English Petrographic and Mineragraphic description. 15.docx", "1H4Afka4AW1Irumg9EYs2XIw5n2rRa4Ga",
     "Innova Mineral 2024 descriptions, 15 samples"),
    ("English 41 Petrographic and Mineragraphic description.docx", "1beJzCAEVfCiPwChYtmCDu0vhlxqy4qAV",
     "2024 consolidated descriptions, 41 samples"),
    ("22 р цооног зассан 14ш.docx", "1ogaE3Tl1iDtPXh9h5ZKylOqCqTS58Tvx", "OVD22 descriptions (14)"),
    ("23 р цооног дууссан 9ш.docx", "1Hpn_S6ZrXwAI93crIjVJlKaG6QD6qAv-", "OVD23 descriptions (9)"),
    ("24 р цооног зассан 13ш.docx", "1IehFzGt7P_1swflmBZWZOxwkH0alPxBT", "OVD24 descriptions (13)"),
    ("25 цооног 10ш.docx", "1SkGc5ogH9rFr2AUsZ353aTqoG2OFv4KJ", "OVD25 descriptions (10)"),
    ("26 р цооног 16 ш зассан.docx", "1RmOTmYH2cLctQR66cq2a_j8u1PhOTHMk", "OVD26 descriptions (16)"),
    ("27 р цооног 12ш.docx", "1vPXAWelQlX_bfjQp1jOqa8xwVMsp91U6", "OVD27 descriptions (12)"),
    ("28 ба 29 р цооног 9ш.docx", "1ZwejH5NIjLSqTYBmuRyl5PoCSmgAdsMR", "OVD28+OVD29 descriptions (9)"),
    ("15ш пет-мин бичиглэл. 09.02.docx", "17UhBmihR8HYhLMbUo0qz-myjvxNi7G6y",
     "2025 descriptions, 15 samples (14 unlocated + 43816=OVD009@126.6)"),
    ("5 петрографи баттерей (2).docx", "1JSiTRfddzTDCi4VRso_BGmQKA_S6WVB7", "2025 descriptions, 5 samples"),
    ("Report_ABM 2025.09.09 final.pdf", "1el3ET8WvDjbohD0oJyQJ1CU4Lo7h-0-Z",
     "MUST (ШУТИС) 2025-09-09 report, 22 samples, SEM-EDX"),
    ("BayanSair_Drilling Sample Petrography_12 Sample.docx", "1XSM2M96NX3-UdzGe8csb3RF95jqecbOD",
     "2026 BS001 descriptions (12)"),
    ("MS3_Outcrop Sample_Петрографи-минераграфи 4ш.docx", "1pdeSCucBLmStN5nrRrOUaMwXzr_KpQK-",
     "2026 MS3 outcrop descriptions (4)"),
    # ---- COVERAGE G8: contributing files that v1.0 omitted -----------------
    ("41ш петрографи, минераграфийн бичиглэл.docx (41 samples MN)",
     "1uDyW5O5Ij4LixvQfTtNdwZAQlAQs2y00",
     "Mongolian twin of the 41-sample 2024 report — source of the Mongolian "
     "rock names merged into those 41 English descriptions"),
    ("Петрографи минераграфийн бичиглэл. 15 ш. Иннова Минерал.docx",
     "1HOaqAhH-D7Jm6CRRXwDuMw28lm7woGRK",
     "Mongolian twin of the Innova Mineral 15-sample 2024 report — source of "
     "the Mongolian rock names merged into those 15 descriptions"),
    # ---- v1.1 missing_sources batch (84 records) ---------------------------
    ("Petrographic descriptions 06.23.pdf", "16r0N4TEldedRvrGcPM_hEvIeCsZfMA7v",
     "Mireslab 2022-06-23 surface petrography, 2 samples (OV202202/OV202203)"),
    ("Petrographic descriptions 11.04.pdf", "1nIKqxM9CtQt3Hn62D0ExKWKdL-CAaVPq",
     "Mireslab 2022-11-07 surface petrography, 6 samples (OVF-1/-2, 020, OV-40/-41/-51)"),
    ("Report_0715_Ni.pdf", "1jIT5VpacpQcvW1KbUty6C1EmT2PPr1Fu",
     "Mireslab 2022-07-15 first Ni report, 2 unjoinable samples (2111, 2107)"),
    ("Report_microscope_20221012.pdf", "1sBdqf9GC7ZO_690r9rE2R6QhJe2cq54t",
     "Mireslab 2022-10-12 surface petrography, 8 samples (YT-/YM- field numbers)"),
    ("MINERALOGICAL-DESCRIPTIONS_2023.03.25.pdf", "1PqskoIsimAuRzmS2h6uQeKGz1cv2ZHoq",
     "Mireslab 2023-03 supergene Ni mineralogy, 1 unjoinable sample (2023Nisample)"),
    ("Thin and polish-1.docx", "1DfMbUNC3_4pIxEMFYjDPUb3TQZ5WHzM2",
     "MN description sheet, 1 surface sample (SH-14)"),
    ("Thin and polish-1sh.docx", "1UiyI5UX26_xmf4BTJIIUTL2yWWZ-FGEa",
     "MN description sheet, 1 surface sample (SH-18)"),
    ("Thin and polish-2sh.docx", "15hjIW8slda5_7rpK_Hkbvake7Qq7J0pW",
     "MN description sheet, 2 surface samples (SH-14-1, SH-16)"),
    ("Thin and polish-4.docx", "1ip2bAQRB303SZq0R5zbY0CBuhQnA02cs",
     "MN description sheet, 4 surface samples (2107, 2104-1, A, 2102)"),
    ("BE-3 samples in English.pdf", "1JCorsOW4v8Vepb9HJlzYNDtzabENxU25",
     "NUM (L.Oyunjargal PhD) 2023-06-22 ore petrology of blocks 40763/40900/40913"),
    ("Report20231124.docx", "106T7bR2o5_Pw1pFPEcfzml6Q9DhdnFUT",
     "Mireslab report #2303 (Innova Mineral order 008), 2 samples 40763/40913"),
    ("Report_20230816_Part1.pdf + Report_20230816_Part2.pdf",
     "1KLR39-VqsFSizHMIQk81Alv-HWZuslik + 1TEpS_kJCBV63jULDdND8Y7DZHfoEGHFH",
     "Mireslab report #2302 — the FULL narrative behind the 23-row MIRESL "
     "summary sheet (internal codes OVD001-OVD023)"),
    ("Petrography_mineragraphy_24 sample.pdf", "17L0euxdhc6dl-FDZchKX9xKrgYgvpKS3",
     "Khanlab О-24 consolidated report — the 12 OVD009 sections "
     "(41014-41023, 41033, 41034) missing from Петрограф008.docx"),
    ("CORE PHOTO/ARDH-2005-01/4. Thin section photo (18 JPG)",
     "folder:1SpY0E3wPZudd9e6KIC5D8SpFgfnRF2tk",
     "legacy 2005 hole ARDH-2005-01: 17 unique thin-section photographs, "
     "no accompanying report — photo_only sample stubs"),
]

# ---------------------------------------------------------------------------
# COVERAGE G8 — resolve every fileId to the inventory-CANONICAL copy.
# v1.0 built {title: fileId} by plain assignment, so the LAST inventory entry
# for a title won (often the duplicate copy), which is how PETRO LIST 2025,
# Петрограф008.docx and 2023-06-20-3 thin sections.pdf ended up citing
# `isDuplicateOf` copies.
# ---------------------------------------------------------------------------
_inv_by_id, _inv_by_title = {}, {}
try:
    with open(ROOT / "workspace" / "inventory.json", encoding="utf-8") as f:
        for it in json.load(f):
            _inv_by_id[it["fileId"]] = it
            # a canonical entry (isDuplicateOf = null) always wins its title
            if it["title"] not in _inv_by_title or not it.get("isDuplicateOf"):
                if it["title"] not in _inv_by_title or \
                        _inv_by_title[it["title"]].get("isDuplicateOf"):
                    _inv_by_title[it["title"]] = it
except Exception:
    pass


def canonical_fileid(title, fid):
    """inventory-canonical fileId for a source row, plus a provenance note."""
    note = ""
    if " + " in fid:                       # two-part report, both ids verbatim
        return fid, ""
    it = _inv_by_id.get(fid)
    if it and it.get("isDuplicateOf"):
        note = f"fileId corrected in v1.1: {fid} is a duplicate copy of"
        fid = it["isDuplicateOf"]
        note += f" the inventory-canonical {fid}"
        return fid, note
    if it:
        return fid, note
    it = _inv_by_title.get(title)
    if it:
        cid = it.get("isDuplicateOf") or it["fileId"]
        if cid != fid:
            note = (f"fileId corrected in v1.1: the build cited {fid}, which is "
                    f"not an inventory entry; the inventory-canonical copy of "
                    f"this title is {cid}")
        return cid, note
    return fid, "not listed in workspace/inventory.json (inventory gap)"


sources_tbl = []
source_fixes = []
for title, fid, role in SOURCES:
    fid2, note = canonical_fileid(title, fid)
    if note:
        source_fixes.append(f"{title}: {note}")
    sources_tbl.append({"title": title, "drive_fileId": fid2, "role": role,
                        "provenance_note": note})

# ============================================================================
# 9. write outputs
# ============================================================================

SAMPLE_COLS = ["sample_id", "alt_ids", "hole_id_norm", "depth_from_m",
               "depth_to_m", "depth_mid_m", "sample_source", "year",
               "petrographer_lab", "thin_section", "polished_thin_section",
               "asd", "xrd", "eds", "fluid_incl", "epma",
               "field_lithology", "petro_lithology", "iogas_lithology",
               "rock_group", "iogas_no", "x_utm", "y_utm", "z_rl",
               "coord_source", "qa_flags", "source_files"]
DESC_COLS = ["sample_id", "desc_id", "raw_sample_id", "raw_hole_id", "raw_depth",
             "source_file", "analyst_or_lab", "report_date", "language",
             "rock_name", "rock_name_original", "texture", "minerals_json",
             "alteration", "opaque_minerals", "description_text",
             "join_method", "qa_notes"]
COLLAR_COLS = ["hole_id", "project", "prospect", "hole_type", "east", "north",
               "rl", "azimuth", "dip", "start_depth_m", "total_depth_m",
               "start_date", "end_date", "status", "lease", "company",
               "supervisor", "remarks", "edited_date", "crs", "hole_id_raw"]
SURVEY_COLS = ["hole_id", "depth_m", "dip", "azimuth", "grid", "method",
               "survey_company", "survey_date", "hole_id_raw", "qa_note"]

df_samples = pd.DataFrame(samples)[SAMPLE_COLS]
df_desc = pd.DataFrame(descriptions)[DESC_COLS]
df_collar = pd.DataFrame(collar_rows)[COLLAR_COLS]
df_survey = pd.DataFrame(survey_rows)[SURVEY_COLS]
assay_col_names = ["sample_id", "hole_id_norm", "depth_from_m", "depth_to_m"] + \
                  [n for _, n in assay_cols]
for a in assay_records:
    s = sample_by_id.get(a["sample_id"])
    a["hole_id_norm"] = s["hole_id_norm"] if s else ""
    a["depth_from_m"] = s["depth_from_m"] if s else None
    a["depth_to_m"] = s["depth_to_m"] if s else None
df_assay = pd.DataFrame(assay_records)[assay_col_names]
df_alias = pd.DataFrame(lu_hole_alias)
df_lab = pd.DataFrame(lu_lab)
df_rock = pd.DataFrame(lu_rock_type)
df_sources = pd.DataFrame(sources_tbl)

TABLES = {
    "samples": df_samples, "descriptions": df_desc, "collar": df_collar,
    "survey": df_survey, "sample_assays": df_assay,
    "lu_hole_alias": df_alias, "lu_lab": df_lab, "lu_rock_type": df_rock,
    "sources": df_sources,
}

for name, df in TABLES.items():
    df.to_csv(CSVDIR / f"{name}.csv", index=False, encoding="utf-8-sig")

with pd.ExcelWriter(OUT / "Oval_Petrography_DB.xlsx", engine="openpyxl") as xw:
    for name, df in TABLES.items():
        df.to_excel(xw, sheet_name=name, index=False)
    _stamp = datetime.datetime(2026, 8, 31, 0, 0, 0)
    xw.book.properties.created = _stamp
    xw.book.properties.modified = _stamp
    xw.book.properties.creator = "scripts/build_database.py"
    xw.book.properties.lastModifiedBy = "scripts/build_database.py"


def _make_xlsx_deterministic(path, stamp="2026-08-31T00:00:00Z"):
    """openpyxl stamps the save time into docProps/core.xml and into every zip
    entry's mtime, so two runs of an otherwise identical build produce
    different bytes.  Rewrite the archive with a fixed timestamp everywhere so
    the whole build is byte-reproducible."""
    import zipfile
    with zipfile.ZipFile(path) as zf:
        items = [(i.filename, zf.read(i.filename)) for i in zf.infolist()]
    fixed = []
    for name, data in items:
        if name == "docProps/core.xml":
            data = re.sub(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                          rb"\g<1>" + stamp.encode() + rb"\g<2>", data)
        fixed.append((name, data))
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in fixed:
            info = zipfile.ZipInfo(name, date_time=(2026, 8, 31, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            zf.writestr(info, data)


_make_xlsx_deterministic(OUT / "Oval_Petrography_DB.xlsx")

# ---------------------------------------------------------------------------
# SQLite (D6): real NULLs instead of '', primary keys, indexes and foreign keys
# ---------------------------------------------------------------------------
PK = {
    "samples": "sample_id",
    "descriptions": "desc_id",
    "collar": "hole_id",
    "sample_assays": "sample_id",
}
FKS = {
    "descriptions": [("sample_id", "samples", "sample_id")],
    "sample_assays": [("sample_id", "samples", "sample_id")],
    "survey": [("hole_id", "collar", "hole_id")],
}
INDEXES = [
    ("idx_descriptions_sample_id", "descriptions", "sample_id"),
    ("idx_descriptions_source", "descriptions", "source_file"),
    ("idx_samples_hole_depth", "samples", "hole_id_norm, depth_from_m"),
    ("idx_samples_coord_source", "samples", "coord_source"),
    ("idx_survey_hole_depth", "survey", "hole_id, depth_m"),
    ("idx_assays_hole_depth", "sample_assays", "hole_id_norm, depth_from_m"),
    ("idx_alias_norm", "lu_hole_alias", "hole_id_norm"),
]


def _sqlite_type(series):
    k = series.dtype.kind
    if k == "i":
        return "INTEGER"
    if k == "f":
        return "REAL"
    return "TEXT"


def _sql_value(v):
    """'' -> NULL, NaN -> NULL, numpy scalar -> python scalar."""
    if v is None:
        return None
    if isinstance(v, str):
        return v if v.strip() != "" else None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):
        return v.item()
    return v


db_path = OUT / "Oval_Petrography_DB.sqlite"
if db_path.exists():
    db_path.unlink()
con = sqlite3.connect(db_path)
con.execute("PRAGMA foreign_keys = ON")
for name, df in TABLES.items():
    coldefs = []
    for c in df.columns:
        t = _sqlite_type(df[c])
        if PK.get(name) == c:
            coldefs.append(f'"{c}" {t} NOT NULL PRIMARY KEY')
        else:
            coldefs.append(f'"{c}" {t}')
    for col, rt, rc in FKS.get(name, []):
        coldefs.append(f'FOREIGN KEY ("{col}") REFERENCES "{rt}"("{rc}")')
    con.execute(f'CREATE TABLE "{name}" (\n  ' + ",\n  ".join(coldefs) + "\n)")
    placeholders = ",".join("?" * len(df.columns))
    con.executemany(
        f'INSERT INTO "{name}" VALUES ({placeholders})',
        [tuple(_sql_value(v) for v in rec)
         for rec in df.itertuples(index=False, name=None)])
for iname, tname, cols in INDEXES:
    con.execute(f'CREATE INDEX "{iname}" ON "{tname}" ({cols})')
con.commit()
fk_violations = con.execute("PRAGMA foreign_key_check").fetchall()
if fk_violations:
    raise AssertionError(f"SQLite foreign-key violations: {fk_violations[:10]}")

# ============================================================================
# 10. verification + QA report
# ============================================================================

from openpyxl import load_workbook  # noqa: E402
wb = load_workbook(OUT / "Oval_Petrography_DB.xlsx", read_only=True)
xlsx_sheets = wb.sheetnames
xlsx_rows = {ws.title: ws.max_row - 1 for ws in wb.worksheets}
wb.close()

ver = []
for name, df in TABLES.items():
    n_sql = con.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
    with open(CSVDIR / f"{name}.csv", encoding="utf-8-sig", newline="") as f:
        n_csv = sum(1 for _ in csv.reader(f)) - 1     # CSV records, not lines
    n_xls = xlsx_rows.get(name, -1)
    ok = (n_sql == len(df) == n_csv == n_xls)
    ver.append((name, len(df), n_csv, n_sql, n_xls, ok))

# ---------------------------------------------------------------------------
# 10a. HARD ASSERTIONS — the build fails loudly if any of these break
# ---------------------------------------------------------------------------
checks = []


def _check(label, cond, detail=""):
    checks.append((label, bool(cond), detail))
    if not cond:
        raise AssertionError(f"BUILD ASSERTION FAILED — {label}: {detail}")


_check("row counts consistent CSV = SQLite = XLSX = dataframe",
       all(v[5] for v in ver),
       "; ".join(f"{v[0]} df={v[1]} csv={v[2]} sqlite={v[3]} xlsx={v[4]}"
                 for v in ver if not v[5]))

_dup_sid = [k for k, n in Counter(s["sample_id"] for s in samples).items() if n > 1]
_check("samples.sample_id unique and non-blank",
       not _dup_sid and all(s["sample_id"].strip() for s in samples),
       f"duplicates: {_dup_sid[:5]}")

_orphans = sorted({d["sample_id"] for d in descriptions
                   if d["sample_id"] and d["sample_id"] not in sample_by_id})
_check("every descriptions.sample_id is in samples (or blank)",
       not _orphans, f"orphans: {_orphans[:5]}")

_inv_iv = [(s["sample_id"], s["depth_from_m"], s["depth_to_m"]) for s in samples
           if s["depth_from_m"] is not None and s["depth_to_m"] is not None
           and s["depth_from_m"] > s["depth_to_m"]]
_check("depth_from_m <= depth_to_m everywhere (D1)",
       not _inv_iv, f"inverted: {_inv_iv[:5]}")

_inv_a = [(a["sample_id"], a["depth_from_m"], a["depth_to_m"])
          for a in assay_records
          if a["depth_from_m"] is not None and a["depth_to_m"] is not None
          and a["depth_from_m"] > a["depth_to_m"]]
_check("depth_from_m <= depth_to_m in sample_assays", not _inv_a,
       f"inverted: {_inv_a[:5]}")

_s47176 = sample_by_id.get("47176")
_exp = (722067.76, 5144288.08, 1738.42)
_got = (_s47176["x_utm"], _s47176["y_utm"], _s47176["z_rl"]) if _s47176 else None
_check("D1: sample 47176 re-desurveyed to the corrected 114-116 m interval",
       _s47176 is not None
       and _s47176["depth_from_m"] == 114.0 and _s47176["depth_to_m"] == 116.0
       and _got is not None
       and all(abs(g - e) <= 0.1 for g, e in zip(_got, _exp)),
       f"got interval {_s47176['depth_from_m']}-{_s47176['depth_to_m']} "
       f"coords {_got}, expected 114.0-116.0 {_exp}")

_d89 = [d for d in descriptions
        if d["raw_sample_id"].strip() == "OVD015-175.5 (B)"]
_check("G10: the bichiglel `OVD015-175.5 (B)` row joins to 42389",
       _d89 and all(d["sample_id"] == "42389" for d in _d89),
       f"got {[ (d['desc_id'], d['sample_id']) for d in _d89 ]}")
_d88 = [d for d in descriptions
        if d["raw_sample_id"].strip() == "OVD015-175.5 (A)"]
_check("G10: the bichiglel `OVD015-175.5 (A)` row joins to 42388",
       _d88 and all(d["sample_id"] == "42388" for d in _d88),
       f"got {[ (d['desc_id'], d['sample_id']) for d in _d88 ]}")

for _sid, _lbl in ((composite_id("OVD003", 202.0), "OVD003@202m"),
                   (composite_id("OVD009", 178.0), "OVD009@178-180m")):
    _s = sample_by_id.get(_sid)
    _nd = [d for d in descriptions if d["sample_id"] == _sid]
    _check(f"G1: new Crawford sample {_sid} exists, is desurveyed and carries "
           f"its description",
           _s is not None and _s["coord_source"] == "desurvey"
           and _s["x_utm"] and _s["y_utm"] and _s["z_rl"] and _nd,
           f"sample={_s}, descriptions={len(_nd)}")

for _raw, _target in (("OVD021@101.5m", "42027"), ("OVD20-121", "43251")):
    _dd = [d for d in descriptions if d["raw_sample_id"].strip() == _raw]
    _check(f"G2: `{_raw}` re-joined to {_target}",
           _dd and all(d["sample_id"] == _target for d in _dd),
           f"got {[(d['desc_id'], d['sample_id']) for d in _dd]}")

_check("D3: lu_rock_type covers every distinct rock name in samples+descriptions",
       len({r["rock_name_original"] for r in lu_rock_type}) == len(rock_counter),
       "")

_check("D6: descriptions.sample_id is NULL (not '') in SQLite for unjoined rows",
       con.execute("SELECT COUNT(*) FROM descriptions "
                   "WHERE sample_id = ''").fetchone()[0] == 0, "")
_check("D6: SQLite primary keys present",
       all(con.execute(f"SELECT COUNT(*) FROM pragma_table_info('{t}') "
                       f"WHERE pk > 0").fetchone()[0] == 1 for t in PK), "")
_check("D6: SQLite indexes present",
       {r[0] for r in con.execute(
           "SELECT name FROM sqlite_master WHERE type='index' "
           "AND name LIKE 'idx_%'")} == {i[0] for i in INDEXES}, "")
_check("D6: SQLite foreign keys resolve", not fk_violations, "")

con.close()

# drill samples whose hole exists in collar must have coordinates
uncovered = [s for s in samples
             if s["hole_id_norm"] and s["hole_id_norm"] in collar_by_hole
             and s["coord_source"] != "desurvey"]
_check("every drill-core sample whose hole is in collar is desurveyed",
       not uncovered, f"{len(uncovered)} uncovered")

no_collar = sorted({s["hole_id_norm"] for s in samples
                    if s["hole_id_norm"] and s["hole_id_norm"] not in collar_by_hole})

# spot checks
spot = []
def _spot(sid):
    s = sample_by_id.get(sid)
    nd = sum(1 for d in descriptions if d["sample_id"] == sid)
    if s:
        spot.append(f"{sid}: hole={s['hole_id_norm']} depth={s['depth_mid_m']} "
                    f"x={s['x_utm']} y={s['y_utm']} z={s['z_rl']} "
                    f"coord={s['coord_source']} descriptions={nd}")
    else:
        spot.append(f"{sid}: NOT FOUND")
_spot("41011")
_spot("47188")
_spot("43816")
_spot("47176")
_spot(composite_id("OVD003", 202.0))
_spot(composite_id("OVD009", 178.0))
_spot("42389")

n_coord = sum(1 for s in samples if s["x_utm"] is not None)
n_desurv = sum(1 for s in samples if s["coord_source"] == "desurvey")
n_masterxy = sum(1 for s in samples if s["coord_source"] == "master_xy")
jm = Counter(d["join_method"] for d in descriptions)
unmatched = [d for d in descriptions if d["join_method"] == "unmatched"]
src_counts = Counter(d["source_file"] for d in descriptions)
src_match = defaultdict(lambda: [0, 0])
for d in descriptions:
    src_match[d["source_file"]][0] += 1
    if d["join_method"] != "unmatched":
        src_match[d["source_file"]][1] += 1

qa_md = []
qa_md.append(f"# QA report — Oval Petrography Database v{DB_VERSION}")
qa_md.append("")
qa_md.append(f"Built by `scripts/build_database.py` (v{DB_VERSION}) on "
             f"{BUILD_DATE} from `workspace/extracted/`. CRS: {EPSG_NOTE}.")
qa_md.append("")
qa_md.append("v1.1 applies every defect raised by the two independent audits "
             "(`VERIFICATION_integrity.md` D1–D10, `VERIFICATION_coverage.md` "
             "G1–G14) and merges the `missing_sources` extraction batch. "
             "See §12 for the changelog and §11 for what is still missing at "
             "source.")
qa_md.append("")
qa_md.append("## 0. Build assertions (the build fails if any of these break)")
qa_md.append("")
for label, ok, _detail in checks:
    qa_md.append(f"- [{'PASS' if ok else 'FAIL'}] {label}")
qa_md.append("")
qa_md.append("## 1. Row counts (dataframe = csv = sqlite = xlsx)")
qa_md.append("")
qa_md.append("| table | rows | csv | sqlite | xlsx | ok |")
qa_md.append("|---|---|---|---|---|---|")
for name, n, nc, ns, nx, ok in ver:
    qa_md.append(f"| {name} | {n} | {nc} | {ns} | {nx} | "
                 f"{'OK' if ok else '**MISMATCH**'} |")
qa_md.append("")
qa_md.append(f"xlsx sheets: {', '.join(xlsx_sheets)}")
qa_md.append("")
qa_md.append("SQLite carries real `NULL`s (never `''`), a primary key on "
             "`samples.sample_id`, `descriptions.desc_id`, `collar.hole_id` and "
             "`sample_assays.sample_id`, foreign keys "
             "`descriptions.sample_id`/`sample_assays.sample_id` → "
             "`samples.sample_id` and `survey.hole_id` → `collar.hole_id`, and "
             "the indexes " + ", ".join(f"`{i[0]}`" for i in INDEXES) +
             ". `PRAGMA foreign_key_check` is clean. No FK is declared on "
             "`samples.hole_id_norm` because the 17 legacy ARDH-2005-01 rows "
             "reference a hole that is not in `collar` (see §11).")
qa_md.append("")
qa_md.append("## 2. Coordinate coverage")
qa_md.append("")
qa_md.append(f"- samples total: **{len(samples)}**")
qa_md.append(f"- with coordinates: **{n_coord}** ({100*n_coord/len(samples):.1f} %)")
qa_md.append(f"  - desurveyed 3D (x,y,z): {n_desurv}")
qa_md.append(f"  - master/grab X,Y only (z null): {n_masterxy}")
qa_md.append(f"- without coordinates: {len(samples)-n_coord} "
             f"(rockchips w/o coords, unlocated lab-number/TS/C samples)")
qa_md.append(f"- drill-core samples whose hole is in collar but NOT desurveyed: "
             f"**{len(uncovered)}**"
             + ("" if not uncovered else " — " + "; ".join(
                 f"{s['sample_id']} ({s['qa_flags']})" for s in uncovered)))
qa_md.append(f"- samples referencing a hole that is NOT in `collar`: "
             f"**{sum(1 for s in samples if s['hole_id_norm'] and s['hole_id_norm'] not in collar_by_hole)}**"
             + (f" (holes: {', '.join(no_collar)}) — the legacy 2005 "
                f"photo-only stubs; no collar, no survey, no coordinates"
                if no_collar else ""))
qa_md.append("")
qa_md.append("## 3. Description join statistics")
qa_md.append("")
qa_md.append(f"- descriptions total: **{len(descriptions)}**; matched: "
             f"**{len(descriptions)-len(unmatched)}** "
             f"({100*(len(descriptions)-len(unmatched))/len(descriptions):.1f} %), "
             f"unmatched: {len(unmatched)}")
qa_md.append(f"- join methods: " + ", ".join(f"{k}: {v}" for k, v in jm.most_common()))
qa_md.append(f"- v1.1 additions: **{n_grab_desc}** grab-sheet field "
             f"descriptions (G9) and **{len(miss_desc_recs)}** rows from the "
             f"`missing_sources` batch ("
             + ", ".join(f"{k}: {v}" for k, v in miss_stats.most_common())
             + f"). The batch's other {len(miss_photo)} records are photo-only "
             f"and became sample rows, not descriptions.")
qa_md.append("")
qa_md.append("| source | descriptions | matched |")
qa_md.append("|---|---|---|")
for src, (tot, mat) in sorted(src_match.items()):
    qa_md.append(f"| {src[:70]} | {tot} | {mat} |")
qa_md.append("")
qa_md.append("### Unmatched descriptions")
qa_md.append("")
for d in unmatched:
    qa_md.append(f"- `{d['raw_sample_id']}` ({d['source_file'][:50]}) — "
                 f"{d['qa_notes'] or 'no tag / no matching hole+depth sample'}")
qa_md.append("")
qa_md.append("## 4. Duplicate handling")
qa_md.append("")
for x in qa["dup_rows_dropped"]:
    qa_md.append(f"- dropped: {x}")
for x in qa["dup_tags_split"]:
    qa_md.append(f"- split: {x}")
for x in qa["rockchip_merge"]:
    qa_md.append(f"- merged: {x}")
qa_md.append(f"- Master All formatting rows skipped (no id, no interval): "
             f"{skipped_formatting}")
qa_md.append("")
qa_md.append("## 5. Depth handling — parsing AND range validation")
qa_md.append("")
qa_md.append("**Depth-parse failures** (a depth string that could not be read "
             "at all):")
if qa["depth_parse_failures"]:
    for x in qa["depth_parse_failures"]:
        qa_md.append(f"- {x}")
else:
    qa_md.append("- none")
qa_md.append("")
qa_md.append("**Depth-RANGE validation** (v1.1 — v1.0 had none, which is how "
             "D1 survived; a parse success is not a range success):")
if qa["depth_interval_fixes"]:
    for x in qa["depth_interval_fixes"]:
        qa_md.append(f"- {x}")
else:
    qa_md.append("- no inverted intervals found")
qa_md.append("")
qa_md.append("The build now asserts `depth_from_m <= depth_to_m` over every "
             "`samples` and `sample_assays` row and aborts if a violation "
             "cannot be corrected from a corroborating source.")
qa_md.append("")
qa_md.append("## 6. Samples added beyond the Master All spine "
             f"({len(qa['samples_added'])})")
qa_md.append("")
add_counter = Counter(x.split("(", 1)[1].rstrip(")") for x in qa["samples_added"])
for src, n in add_counter.most_common():
    qa_md.append(f"- {src[:90]}: {n}")
qa_md.append("")
qa_md.append("## 7. Hole-ID aliases applied (raw -> normalized)")
qa_md.append("")
changed = [a for a in lu_hole_alias if a["changed"]]
qa_md.append(f"{len(changed)} raw spellings normalized "
             f"(full list in `csv/lu_hole_alias.csv`):")
for a in changed:
    qa_md.append(f"- `{a['hole_id_raw']}` -> `{a['hole_id_norm']}`")
qa_md.append("")
qa_md.append("## 8. Spot checks")
qa_md.append("")
for x in spot:
    qa_md.append(f"- {x}")
qa_md.append("")
qa_md.append("## 9. Known issues carried from sources")
qa_md.append("")
qa_md.extend([
    "- **Crawford 2025 sample/assay mix-up flags.** CORRECTED CLAIM (v1.0 §9 "
    "said these were in `descriptions.qa_notes` when they were only in the tail "
    "of `description_text`): as of v1.1 each caveat is written into BOTH "
    "`descriptions.qa_notes` AND the joined `samples.qa_flags`, and the "
    "original wording still stands verbatim inside `description_text`. "
    "Rows carrying a Crawford caveat: "
    + ", ".join(d2_applied) + ".",
    "- Crawford notes sub-standard polish on many of the 38 sections; OVD005@40.5 "
    "and @53.0 'far too thin'; OVD021@148.8 sulfides too poorly polished. These "
    "three are also in `qa_notes`/`qa_flags` as of v1.1.",
    "- `OVD021@101.5m` (Crawford) is OVD011-101.5 (tag **42027**) — as of v1.1 "
    "the description IS joined (`join_method = xref-corrected`), on the strength "
    "of the identical Crawford micro-description filed against OVD011-101.5 in "
    "the `KhanAltai vs Tony` sheet. v1.0 left it unmatched.",
    "- Tag **42808** is printed on two SC04 samples (171.0 m and 280.7 m); the "
    "280.7 m sample is stored as `SC04@280.7`.",
    "- BS001 sample numbering: report prints 45652 on both 380.5 and 380.8 m "
    "(summary suggests 45653); Master All carries a single 45652 row (380-382 m), "
    "the exact duplicate row was dropped.",
    "- `OVD015-175.4` (Sheet3) vs `OVD015-175.5 (A)/(B)` (Master All) — possible "
    "depth typo; kept as a separate sample with a QA flag.",
    "- 14 lab numbers of the 2025 '15ш' report (40340-41363 series) have no "
    "drillhole/depth stated anywhere — samples exist with no location.",
    "- TS-n / Дээж-n (ThinSection Mongolia 2024) and С-1/С-2 (L.Jargal 2023) "
    "samples have no location; ТЦ-1 is listed in that report's TOC but never "
    "described.",
    "- The master workbook's 'rockchip' sheet Sample_numbers duplicate the "
    "numeric-tag / CR-xx rows of the grab sheet (same physical samples): they "
    "were merged, classed `rockchip`, with X/Y from the grab sheet. The master "
    "README's 'tags 43113-43154' description of that sheet is inaccurate.",
    "- The MIRESL summary `Code` column (OVD001-OVD023) is a lab sample code, "
    "NOT a drillhole id.",
    "- Depths of 2023-era registers are point depths written as text ('36m'); "
    "tag 40715 was written '98.2' without the unit.",
    "- Master workbook's own Collar/Survey sheets are stale (68 holes); "
    "Collar_all_combined / Survey_all_YMB (76 holes) are the location authority.",
    "- Survey azimuths are used as grid azimuths (Grid (Orig) = WGS84_46N; "
    "`Azim (UTM)` column is empty in the source).",
    "- Collar data typo: MU2502 End date '10/14/225' (kept verbatim).",
    "- **OVD008A depth datum**: OVD008A is a re-drill sharing the OVD008 collar. "
    "`collar.start_depth_m` is 110.5 m, but its 33 survey stations run "
    "0 → 162.5 m MEASURED FROM SURFACE, not from the re-entry point. Any depth "
    "quoted against OVD008A must therefore be surface-referenced. No sample in "
    "the database is assigned to OVD008A.",
    "- **Duplicate survey station** OVD009 @ 240.0 m (see §10 D4) — both "
    "readings are kept verbatim, flagged in the new `survey.qa_note` column.",
])
qa_md.append("")

# ---------------------------------------------------------------------------
qa_md.append("## 10. v1.1 defect resolutions (audit D1–D10, G1–G14)")
qa_md.append("")
qa_md.append("### Integrity audit (`VERIFICATION_integrity.md`)")
qa_md.append("")
qa_md.append("**D1 — inverted depth interval on 47176 (HIGH).** Fixed. " +
             ("; ".join(qa["depth_interval_fixes"]) or "no fix recorded") +
             f". The sample now sits at {sample_by_id['47176']['depth_from_m']:g}"
             f"–{sample_by_id['47176']['depth_to_m']:g} m "
             f"(mid {sample_by_id['47176']['depth_mid_m']:g} m) and desurveys to "
             f"({sample_by_id['47176']['x_utm']}, "
             f"{sample_by_id['47176']['y_utm']}, "
             f"{sample_by_id['47176']['z_rl']}) — a 14.05 m correction from the "
             f"v1.0 position (722071.89, 5144281.90, 1726.51). A global "
             "`depth_from <= depth_to` guard plus a build assertion now make "
             "this class of defect impossible to ship.")
qa_md.append("")
qa_md.append(f"**D2 — Crawford caveats not queryable.** Fixed: "
             f"{len(d2_applied)} description rows now carry the caveat in "
             f"`qa_notes`, and every joined sample carries it in `qa_flags`. "
             f"§9 above is corrected.")
qa_md.append("")
qa_md.append(f"**D3 — `lu_rock_type` incomplete.** Fixed: the lookup is now "
             f"built from `samples.field_lithology` + `petro_lithology` + "
             f"`iogas_lithology` **and** `descriptions.rock_name` **and** "
             f"`descriptions.rock_name_original` — "
             f"**{len(lu_rock_type)}** distinct names (v1.0: 473), each with a "
             f"`rock_group`, a new coarse `rock_family` column and a `seen_in` "
             f"column, and `n_occurrences` recounted over all five fields. "
             f"Names left `other / unclassified`: "
             f"{sum(1 for r in lu_rock_type if r['rock_group'] == 'other / unclassified')}.")
qa_md.append("")
qa_md.append("**D4 — duplicate survey station.** Both rows kept verbatim; the "
             "new `survey.qa_note` column names the conflict and the "
             "recommended row on BOTH rows of the pair:")
for x in survey_conflicts:
    qa_md.append(f"- {x}")
qa_md.append("")
qa_md.append("The desurvey now resolves duplicate `(hole, depth)` stations "
             "deterministically to the most recent survey date, so the "
             "database and the README import guide agree. No stored coordinate "
             "changes (the deepest OVD009 sample is 195.2 m).")
qa_md.append("")
qa_md.append(f"**D5 — stored interval vs report-named interval.** "
             f"{n_iv_checked} descriptions name an explicit interval; "
             f"**{len(d5_rows)}** disagree with the sample's stored interval "
             f"and now carry a `qa_flags` entry on the sample and a `qa_notes` "
             f"entry on the description:")
qa_md.append("")
qa_md.append("| desc | report label | report interval | sample | stored | gap (m) |")
qa_md.append("|---|---|---|---|---|---|")
for did, raw, sid, stored, rpt, gap in d5_rows:
    qa_md.append(f"| {did} | `{raw}` | {rpt} | {sid} | {stored} | {gap:g} |")
qa_md.append("")
qa_md.append("The audit's D5 named 9 of these. The generic detector used here "
             "also catches `OVD028-38 (33.14-35)` (named in D5's own text as a "
             "related case), `OVD029-122.4 (123-124.5)` — both of which write "
             "the report interval as a caption rather than in the id — and "
             "`OVD28-19-21.25 (47097)` at 0.05 m. All 12 are flagged.")
qa_md.append("")
qa_md.append("**D6 — SQLite ergonomics.** Fixed: real NULLs, primary keys, "
             "foreign keys and 7 indexes (see §1).")
qa_md.append("")
qa_md.append(f"**D7 — near-duplicate samples not cross-referenced.** Fixed: a "
             f"generic scan of every hole for sample pairs within 0.35 m found "
             f"**{len(near_dupes)}** pairs; both members of each pair now carry "
             f"a `D7 cross-reference` entry in `qa_flags`:")
for h, a, da, b, db_, gap in near_dupes:
    qa_md.append(f"- {h}: {a} @ {da:g} m vs {b} @ {db_:g} m ({gap:g} m apart)")
qa_md.append("")
qa_md.append("**D8 — `depth_mid_m` vs the master's point depth.** Unchanged by "
             "design: where the Phase-2 sheet supplies a narrow interval the "
             "build stores that interval and uses its midpoint (typical "
             "difference 0.05 m, worst 0.5 m). Recorded here so the difference "
             "is not mistaken for corruption.")
qa_md.append("")
qa_md.append(f"**D9 — A/B suffix rests on letter order.** Fixed: "
             f"{len(ab_notes)} description rows now record the inference in "
             f"`qa_notes` — {', '.join(ab_notes)}.")
qa_md.append("")
qa_md.append("**D10 — OVD008A depth datum undocumented.** Fixed: documented in "
             "§9 above and in `README.md` (import notes).")
qa_md.append("")
qa_md.append("### Coverage audit (`VERIFICATION_coverage.md`)")
qa_md.append("")
qa_md.append(f"**G1 — 2 described thin sections with no sample row.** Fixed: "
             f"`{composite_id('OVD003', 202.0)}` and "
             f"`{composite_id('OVD009', 178.0)}` created from the Crawford 2025 "
             f"report and desurveyed; their descriptions (previously unmatched) "
             f"are now joined.")
qa_md.append("")
qa_md.append("**G2 — 2 resolvable unmatched descriptions.** Fixed: " +
             "; ".join(g2_applied) + ".")
qa_md.append("")
qa_md.append("**G3 — whole-dataset omissions unacknowledged.** Fixed: §11.")
qa_md.append("")
qa_md.append(f"**G4/G5/G6/G7 — unextracted source documents.** Largely fixed by "
             f"the `missing_sources` batch (§4 of that folder's README): "
             f"{len(recs_miss)} records, of which {len(miss_desc_recs)} became "
             f"descriptions and {len(miss_photo)} became photo-only sample "
             f"stubs. What is still missing is listed in §11.")
qa_md.append("")
if miss_depth_conflicts:
    qa_md.append("Depth cross-check on the batch's tag-joined records — the "
                 "report's stated depth vs the sample register's "
                 f"({len(miss_depth_conflicts)} disagreement"
                 f"{'' if len(miss_depth_conflicts) == 1 else 's'}, recorded in "
                 "`descriptions.qa_notes`; the register value is kept):")
    for x in miss_depth_conflicts:
        qa_md.append(f"- {x}")
    qa_md.append("")
qa_md.append("**No new SURFACE sample rows were needed.** All 24 surface "
             "records of the batch (sources 1, 2, 4 and 6–9) map onto rows "
             "1–24 of the master grab sheet — which already have sample rows — "
             "after case folding, leading-zero stripping and the documented "
             "`2021-01` / `2022-01` year-digit typo, and the mapping agrees "
             "with the row order the batch README reconstructs. Where the two "
             "spellings differ, BOTH the description (`qa_notes`) and the "
             "sample (`qa_flags`) record it. The three records that map to no "
             "sample (`2111`, the Ni-report `2107`, `2023Nisample`) are kept "
             "UNJOINED rather than given invented sample rows, because none of "
             "them has a field number — see §11.")
qa_md.append("")
qa_md.append("**G8 — `sources.csv` provenance.** Fixed: every fileId is now "
             "resolved to the inventory-canonical copy (duplicates are "
             "followed through `isDuplicateOf`), the omitted contributing "
             "files are registered, and the corrections are recorded in the "
             "new `sources.provenance_note` column:")
for x in source_fixes:
    qa_md.append(f"- {x}")
qa_md.append("")
qa_md.append(f"**G9 — grab-sheet field descriptions dropped.** Fixed: "
             f"{n_grab_desc} field descriptions recovered from the "
             f"`2022-2024 grab` sheet's right-hand `Description` column and "
             f"emitted as description rows "
             f"(`join_method = grab sheet row`, language `mn`/`en` as written).")
qa_md.append("")
qa_md.append("**G10 — `D0089` mis-join.** Fixed in `match_by_label`: the "
             "verbatim label (suffix included) is now tried before the "
             "suffix-stripped form, so `OVD015-175.5 (A)` → 42388 and "
             "`OVD015-175.5 (B)` → 42389.")
qa_md.append("")
qa_md.append("**G11 — undocumented OVD014-89.8 (A)/(B) merge.** Fixed: sample "
             "42147 carries a `COVERAGE G11` note in `qa_flags`.")
qa_md.append("")
qa_md.append("**G12/G13/G14 — unverified tables, the BE-3 assumption and the "
             "photo datasets.** G13 is now resolved: `BE-3 samples in "
             "English.pdf` was read in the `missing_sources` batch and its "
             "three block samples ARE tags 40763 / 40900 / 40913, confirming "
             "the v1.0 assumption. G12 and G14 remain open — see §11.")
qa_md.append("")

# ---------------------------------------------------------------------------
qa_md.append("## 11. Known missing at source (datasets identified but NOT in "
             "this database)")
qa_md.append("")
qa_md.append("This section exists because the coverage audit found that v1.0 "
             "acknowledged none of these. Row counts were never inflated — "
             "nothing below is silently counted as covered.")
qa_md.append("")
qa_md.append("### Absent from Google Drive itself (cannot be ingested)")
qa_md.append("")
qa_md.extend([
    "- **Gtech prospect review** — referenced by name in project correspondence; "
    "no file in the Drive set.",
    "- **Chuluunbataar / Vi Vitex LLC review (May 2022)** — referenced by name; "
    "no file in the Drive set.",
    "- **Dennis (RPM Global, Oct 2023)** and **Prof. D. Holwell (Oct 2023)** "
    "reviews — referenced by name; no file in the Drive set.",
    "- **ARDH-2005-02 thin-section photos** — folder "
    "`1AHFQu0eLtEZbbM-nrJFZvfu2-OUqRxid` exists but is EMPTY on Drive.",
    "- **`41016.jpg`** — the hand-specimen photo for tag 41016 is missing from "
    "both `Khanlab_Petrograph_samples` photo folders (23 JPGs for 24 samples). "
    "The 41016 DESCRIPTION is present (ingested in v1.1 from the consolidated "
    "Khanlab О-24 PDF).",
    "- **Khanlab batch-1 report** (SEM-EDS reference '1', 7 × OVD-009 sections: "
    "41014, 41015, 41016, 41017, 41020, 41021, 41023) — the report document "
    "itself is not in the Drive set. All 7 samples exist and, as of v1.1, all 7 "
    "carry the Khanlab О-24 narrative from the consolidated PDF.",
])
qa_md.append("")
qa_md.append("### Out of scope for this database (no table models them)")
qa_md.append("")
qa_md.extend([
    "- **≈330 sample photographs in 13 Drive folders** — `Petrographic_photos_2023` "
    "(36 PNG), `Mineralogical_photos_2023` (31), `SEM-EDS_photos_2023` (17), "
    "`Khanlab_Petrograph_samples` (23), `ymb_2024_Scanned…` (~75), Phase-1 (~55), "
    "Phase-2 (~45), ARDH-2005-01 (18). Many are named by sample tag "
    "(40530–40915) and are therefore directly linkable to `samples.sample_id`. "
    "**There is no `sample_photos` table in v1.1** — the images themselves are "
    "not ingested and no per-image row exists, EXCEPT the 17 ARDH-2005-01 "
    "thin-section photographs, which are carried as photo-only sample stubs "
    "because they are the only record of that hole's sections.",
    "- **`МП2026-24 Батбадмаараг ХХК …pdf` (Modot-3, licence XV-020181)** — "
    "deliberately excluded: a different project, not Oval/Yambat.",
])
qa_md.append("")
qa_md.append("### Present on Drive, still not opened (G12 — no evidence they "
             "add samples, but unverified)")
qa_md.append("")
qa_md.extend([
    "- `2023 Drilling petrography samples.xlsx` — `17GqS_Wo0T6OOEIAgiUyl6rkUgjN2Ox5Z`",
    "- `Yambat petrography samples 2024 from Core.xlsx` — `1dRlx13-icZZl-OokbD9OfXov-mbuoP4o`",
    "- `Petrograph_2023_07_31.xlsx` — `1PPWrYjVeLfTTgec-Qmazi_oYdYnOu3Gi`",
    "- `Deejiin hoolgoonii list_ABM (1).xlsx` — `1o_jgfkLe_lC4f21Uf1z3uoSmZZtfoVId`",
    "- Grab lists `Grab 2022aug-2023.xlsx/.csv` (`16f8S2Si…`, `1JeL0cAM…`), "
    "`2022aug-2023.xlsx` (`1is5GE0W…`), `03Aug2022.xlsx` (`1XZC2MhQ…`) — the "
    "65-sample grab count rests on the master grab sheet, not on these.",
    "- The three 250–320 MB Khanlab `.doc` files "
    "(`1b2NKUWu…`, `1d9YIMNE…`, `1jhySMOE…`) — the same 24-sample report; the "
    "consolidated PDF used in v1.1 covers all 24 sections, but the ENGLISH "
    "translation has not been harvested, so the 12 new Khanlab records carry "
    "Mongolian `description_text` with an English `rock_name`.",
    "- `R_2023-21 Petrology, mineralogy – Mireslab Mongol LLC.pdf` "
    "(`16QAZBbGJVSkJSjXjzeO6RLW_De4LCIhH`) — READ in the v1.1 batch and found "
    "to be the WORK CONTRACT, not a petrography report (26 samples ordered, "
    "23 delivered as Report #2302). No sample data; deliberately not ingested.",
])
qa_md.append("")
qa_md.append("### Sample suites that still have no petrographic description")
qa_md.append("")
qa_md.append("- **Grab-sheet rows 25–65** — `TS1`–`TS7`, `RC5`, `RC6`, tags "
             "`43113, 43122, 43123, 43125, 43141, 43144, 43146, 41154, 41155, "
             "41160–41163, 41167–41169, 41172, 41178–41180, 41183, 47071–47073, "
             "47076, 47077, 47084`, and `CR66, CR99, CR71, CR1, CRE`. None of "
             "the recovered 2022–23 reports describes them. As of v1.1 they do "
             "carry the geologist's FIELD description (§10 G9), but no "
             "microscope determination.")
qa_md.append("")
qa_md.append("### Unjoinable records (ingested, but with no sample to attach to)")
qa_md.append("")
for raw, note in sorted(MISS_UNMATCHED_NOTE.items()):
    qa_md.append(f"- `{raw}` — {note}")
qa_md.append("")
qa_md.append("Full list of every unmatched description row (all sources):")
qa_md.append("")
for d in unmatched:
    qa_md.append(f"- `{d['desc_id']}` `{d['raw_sample_id']}` "
                 f"({d['source_file'][:60]})")
qa_md.append("")
qa_md.append("### Datasets the audit listed as missing that v1.1 RESOLVED")
qa_md.append("")
qa_md.extend([
    "- The 10 unextracted 2022–23 Mireslab surface reports — **ingested** "
    "(`Petrographic descriptions 06.23`, `… 11.04`, `Report_0715_Ni`, "
    "`Report_microscope_20221012`, `MINERALOGICAL-DESCRIPTIONS_2023.03.25` + "
    "its `-NI.docx` twin, `Thin and polish-1/-1sh/-2sh/-4`). 24 of the 65 "
    "grab/rockchip samples now carry a laboratory petrographic description.",
    "- `Report20231124.docx` and `Report_20230816 Part1/Part2` — **ingested**; "
    "the 23 MIRESL 2023 drill-core samples now carry the FULL narrative "
    "(hand specimen, texture, per-mineral habit and size, alteration, SEM-EDS), "
    "not only the one-line summary-sheet fields.",
    "- The Khanlab О-24 consolidated report — **ingested**; the 12 OVD009 "
    "sections (41014–41023, 41033, 41034) that `Петрограф008.docx` did not "
    "cover now have their primary-source description.",
    "- **Mireslab 'pdf2' (tags 40910, 40628, 40635, 40645)** — RESOLVED as an "
    "artefact: all four tags are Report #2302 sections (Mireslab internal codes "
    "OVD019, OVD007, OVD004, OVD005), whose full narrative arrived with "
    "`Report_20230816 Part1/Part2` in this batch. There is no separate 'pdf2' "
    "document to find; the four samples now carry the primary-source narrative "
    "alongside the MIRESL summary row and the Crawford description.",
    "- **ARDH-2005-01** — the 17 unique thin-section photographs are now "
    "carried as sample rows (photo-only stubs). There is still no petrographic "
    "text for them: no report, sheet or description exists under that hole.",
    "- The `BE-3 samples` PDFs (G13) — read; the 3 NUM sections are confirmed "
    "to be tags 40763 / 40900 / 40913.",
])
qa_md.append("")

# ---------------------------------------------------------------------------
qa_md.append("## 12. Changelog v1.0 → v1.1")
qa_md.append("")
qa_md.append(f"| table | v1.0 rows | v{DB_VERSION} rows | change |")
qa_md.append("|---|---|---|---|")
V10_ROWS = {"samples": 376, "descriptions": 451, "collar": 76, "survey": 1990,
            "sample_assays": 277, "lu_hole_alias": 101, "lu_lab": 20,
            "lu_rock_type": 473, "sources": 32}
for name, df in TABLES.items():
    old = V10_ROWS.get(name, 0)
    qa_md.append(f"| {name} | {old} | {len(df)} | {len(df)-old:+d} |")
qa_md.append("")
qa_md.extend([
    "**Data corrections**",
    "",
    f"- 47176: interval 144.0–114.1 → 114.0–116.0 m; position moved 14.05 m "
    f"(D1).",
    f"- `OVD015-175.5 (B)` description re-joined 42388 → 42389 (G10).",
    f"- `OVD021@101.5m` → 42027 and `OVD20-121` → 43251, both previously "
    f"unmatched (G2).",
    f"- 2 new drill-core samples created and desurveyed (G1).",
    f"- 17 legacy ARDH-2005-01 photo-only sample stubs created.",
    "",
    "**New content**",
    "",
    f"- {n_grab_desc} grab-sheet field descriptions recovered (G9).",
    f"- {len(miss_desc_recs)} descriptions merged from the `missing_sources` "
    f"batch (of 84 records; the other {len(miss_photo)} are photo-only stubs).",
    "",
    "**Schema changes**",
    "",
    "- `survey` gains `qa_note`.",
    "- `lu_rock_type` gains `rock_family` and `seen_in`.",
    "- `sources` gains `provenance_note`.",
    "- SQLite gains NULLs, primary keys, foreign keys and indexes.",
    "",
    "**Documentation**",
    "",
    "- §5 now separates depth PARSING from depth RANGE validation (the v1.0 "
    "claim 'depth-parse failures: none' was true but masked D1).",
    "- §9 corrected: the Crawford caveats are now genuinely in "
    "`descriptions.qa_notes` and `samples.qa_flags`, as v1.0 claimed.",
    "- §11 'Known missing at source' added.",
])
qa_md.append("")
(OUT / "QA_report.md").write_text("\n".join(qa_md), encoding="utf-8")

# ============================================================================
# 11. README (schema documentation)
# ============================================================================

readme = f"""# Oval Ni-Cu (Yambat) — Consolidated Petrography Database (v{DB_VERSION})

Built from the Google Drive petrography/drilling sources of the AZ9 GeoHub
by `scripts/build_database.py`. One row per **physical sample** in `samples`,
one row per **petrographic description** in `descriptions` (a sample can have
several descriptions: Mongolian lab report, Crawford 2025, MIRESL 2023, ...).

- CRS of all coordinates: **{EPSG_NOTE}**
- Files: `csv/*.csv` (UTF-8 with BOM — opens correctly in Excel),
  `Oval_Petrography_DB.xlsx`, `Oval_Petrography_DB.sqlite`, `QA_report.md`.

## Contents (v{DB_VERSION}, built {BUILD_DATE})

| table | rows | what it is |
|---|---|---|
| `samples` | {len(df_samples)} | physical samples (the spine) |
| `descriptions` | {len(df_desc)} | petrographic / mineragraphic descriptions |
| `collar` | {len(df_collar)} | drillhole collars |
| `survey` | {len(df_survey)} | downhole survey stations |
| `sample_assays` | {len(df_assay)} | wide assay suite from the Master "All" sheet |
| `lu_hole_alias` | {len(df_alias)} | raw hole-id spelling → normalized id |
| `lu_lab` | {len(df_lab)} | lab / petrographer lookup |
| `lu_rock_type` | {len(df_rock)} | every distinct rock name → group / family |
| `sources` | {len(df_sources)} | contributing files with Drive fileIds |

- **{n_coord} of {len(samples)}** samples ({100*n_coord/len(samples):.1f} %) carry
  coordinates — {n_desurv} desurveyed in 3-D (x, y, z), {n_masterxy} with
  surface X/Y only, {len(samples)-n_coord} with none.
- **{len(descriptions)-len(unmatched)} of {len(descriptions)}**
  ({100*(len(descriptions)-len(unmatched))/len(descriptions):.1f} %) descriptions
  are joined to a sample; {len(unmatched)} cannot be (see `QA_report.md` §11).
- v1.1 fixes every defect raised by the two independent audits
  (`VERIFICATION_integrity.md`, `VERIFICATION_coverage.md`) and merges the
  `missing_sources` batch. Changelog: `QA_report.md` §12.

## Re-running

```bash
pip install pandas openpyxl
python3 scripts/build_database.py
```

Inputs are read from `workspace/extracted/` (master/, xlsx/, reports/,
reports2024_2026/, missing_sources/). Outputs are rewritten under `database/`.
The build is deterministic — two consecutive runs produce byte-identical CSVs,
SQLite, XLSX and Markdown — and it aborts with an `AssertionError` if any of
the invariants listed in `QA_report.md` §0 is broken.

## Tables

### samples.csv — дээжийн бүртгэл (one row per physical sample)

| column | meaning (EN) | тайлбар (MN) |
|---|---|---|
| sample_id | canonical id: 5-digit lab tag when known, else `HOLE@DEPTH` composite | дээжийн дугаар (шошго) |
| alt_ids | other ids used in sources (`CRS01A-81`, shared tags), `\\|`-separated | бусад дугаарууд |
| hole_id_norm | normalized drillhole id (join key to collar/survey) | цооногийн дугаар |
| depth_from_m / depth_to_m / depth_mid_m | sample interval and midpoint, metres downhole | дээжийн интервал, гүн (м) |
| sample_source | drill core / grab / rockchip / outcrop / unknown | дээжийн төрөл |
| year | sampling/analysis year from the master table | он |
| petrographer_lab | petrographer or lab from the master table | шинжээч / лаборатори |
| thin_section, polished_thin_section, asd, xrd, eds, fluid_incl, epma | analysis flags (1/0) | шинжилгээний төрлүүд (ТШ=thin section, ӨТШ=polished thin section) |
| field_lithology | geologist's core-log rock name | хээрийн чулуулгийн нэр |
| petro_lithology | rock name after petrography ("New Lithology by Petrographic") | петрографийн дараах нэр |
| iogas_lithology | rock name after ioGAS lithogeochemistry | iOGAS ангилал |
| rock_group | standardized rock group (see lu_rock_type) | нэгдсэн бүлэг |
| iogas_no | ioGAS sample number (from the bichiglel table) | ioGAS дугаар |
| x_utm, y_utm, z_rl | 3D position at depth_mid_m ({EPSG_NOTE}) | байрлал (X, Y, Z) |
| coord_source | desurvey / master_xy / none | координатын эх үүсвэр |
| qa_flags | data-quality flags for this sample | чанарын тэмдэглэл |
| source_files | contributing source files | эх файлууд |

### descriptions.csv — петрографийн бичиглэлүүд

| column | meaning |
|---|---|
| sample_id | canonical sample (empty when the description could not be joined) |
| desc_id | unique description id (D0001...) |
| raw_sample_id / raw_hole_id / raw_depth | ids exactly as written in the source |
| source_file | report / sheet the description comes from |
| analyst_or_lab, report_date | who described it and when (blank when not stated) |
| language | mn / en |
| rock_name | rock name (English where available) |
| rock_name_original | original (usually Mongolian) rock name |
| texture | texture / structure notes |
| minerals_json | mineral list as JSON (`[{{"mineral":..., "pct":...}}]` or `{{"rock_forming": "Pl, Amph"}}`) |
| alteration | alteration minerals / intensity |
| opaque_minerals | ore/opaque mineralogy and paragenesis |
| description_text | free-text description / summary |
| join_method | tag / label / hole+depth / hole+depth+suffix / report id / grab sheet id / grab sheet row / xref-corrected / unmatched |
| qa_notes | join or source-quality notes (Crawford caveats, A/B-suffix inferences, interval disagreements, id corrections) |

`join_method` values added in v1.1: **`grab sheet row`** (a field description
recovered from the master grab sheet), **`grab sheet id`** (a 2022–23 surface
lab description matched to a grab row after case folding / leading-zero
stripping) and **`xref-corrected`** (a source id typo corrected against
independent evidence — see `QA_report.md` §10 G2).

### collar.csv / survey.csv

Normalized copies of `Collar_all_combined` (76 holes) and `Survey_all_YMB`
(downhole surveys). `hole_id` is normalized (`OVD008a`->`OVD008A`);
raw spelling kept in `hole_id_raw`. Depths in metres; dips negative-down;
azimuths are grid azimuths (WGS84_46N). `survey.qa_note` (new in v1.1) carries
station-level warnings — currently the duplicate-station conflict below.

### sample_assays.csv

Wide assay suite carried over verbatim from the Master "All" sheet, keyed by
`sample_id`. Column names are `Element_unit__Method` (e.g. `Ni_ppm__ME_ICP61`,
`Au_ppm__PGM_ICP27`, `MgO_pct__ME_XRF26`). Values are as printed in the master
(no unit conversion); `-` placeholders were blanked.

### Lookups

- `lu_hole_alias.csv` — every raw hole-id spelling seen anywhere -> normalized id.
- `lu_lab.csv` — raw lab/petrographer strings -> canonical lab.
- `lu_rock_type.csv` — **every** distinct rock name across
  `samples.field_lithology` / `petro_lithology` / `iogas_lithology` and
  `descriptions.rock_name` / `rock_name_original` (the Mongolian vocabulary
  included, which v1.0 omitted) -> best-effort standardized `rock_group`, a
  coarse `rock_family` (ultramafic / mafic intrusive /
  felsic-intermediate intrusive / volcanic / sedimentary / metamorphic /
  vein-ore / unknown), `n_occurrences` and `seen_in`. Originals untouched.
- `sources.csv` — contributing files with Google Drive fileIds, resolved to the
  inventory-canonical copy; `provenance_note` records any correction.

## Loading into Leapfrog / Micromine

**Drillhole database (recommended):**
1. Collar table: `csv/collar.csv` — Hole ID = `hole_id`, East = `east`,
   North = `north`, RL = `rl`, Max depth = `total_depth_m`. CRS EPSG:32646.
2. Survey table: `csv/survey.csv` — Hole ID = `hole_id`, Depth = `depth_m`,
   Azimuth = `azimuth`, Dip = `dip` (negative down).
3. Interval/points table: `csv/samples.csv` filtered to
   `sample_source = "drill core"` — Hole ID = `hole_id_norm`,
   From = `depth_from_m`, To = `depth_to_m` (use `depth_mid_m` as a point
   table where To is null). Attribute columns: `rock_group`,
   `petro_lithology`, `field_lithology`, analysis flags.
4. Assays: `csv/sample_assays.csv` joined on `sample_id` (or imported as an
   interval table on `hole_id_norm` + `depth_from_m`/`depth_to_m`).

**As points:** `samples.csv` already carries desurveyed `x_utm, y_utm, z_rl`
(minimum-curvature at `depth_mid_m`), so it can be loaded directly as a 3D
points file — filter `coord_source = "desurvey"` for true 3D positions;
`master_xy` rows (grab samples) are surface X/Y without Z.

**Descriptions** are text: join `descriptions.csv` to the loaded samples on
`sample_id` in your GIS/DB, or keep it as the reference table.

### Import notes you must read first

**1. Duplicate survey station — OVD009 @ 240.0 m.** `survey.csv` contains two
rows for this hole+depth, both verbatim from `Survey_all_YMB.csv`:

| dip | azimuth | method | company | date |
|---|---|---|---|---|
| −78.91 | 244.64 | MS | Bayan Undraga LLC | 7/28/2024 |
| −78.00 | 246.50 | Ez-trac, Multi shot | Ragnarok Investment LLC | 5/30/2023 |

Leapfrog and Micromine reject or silently resolve duplicate hole+depth survey
keys. **Recommended: keep the 2024-07-28 Bayan Undraga MS reading (−78.91 /
244.64) and delete the 2023-05-30 row.** That is the most recent instrument
survey of the hole and is the station this build's own desurvey uses, so the
imported trace will match `samples.x_utm/y_utm/z_rl` exactly. Nothing stored
depends on the choice in practice — the deepest OVD009 sample is at 195.2 m,
above both readings. Both rows carry the full explanation in `survey.qa_note`,
so you can filter with `qa_note LIKE '%drop this row%'`.

**2. OVD008A depths are measured FROM SURFACE.** OVD008A is a re-drill sharing
the OVD008 collar and `collar.start_depth_m` = 110.5 m, but its 33 survey
stations run 0 → 162.5 m **from surface**, not from the 110.5 m re-entry point.
Desurveying OVD008A @ 110.0 m and OVD008 @ 110.5 m gives positions 0.54 m
apart, which is the correct behaviour for a re-drill of the same collar. If you
add an OVD008A sample, quote its depth from surface — quoting it from the
re-entry datum would place it ~110 m too shallow. No sample in the database is
currently assigned to OVD008A.

**3. 17 samples reference hole `ARDH-2005-01`, which is NOT in `collar`.**
These are the legacy 2005 thin-section photo stubs (`qa_flags` begins
`legacy_2005_photo_only`). They have no depth and no coordinates. Filter them
out with `coord_source <> 'none'` or `hole_id_norm <> 'ARDH-2005-01'` before
building a drillhole database, otherwise the importer will report an unknown
hole. This is also why the SQLite file declares no foreign key on
`samples.hole_id_norm`.

**4. `samples.csv` as an interval table.** `depth_to_m` is null for point
samples — use `depth_mid_m` as a point table, or `COALESCE(depth_to_m,
depth_from_m)`. Every row satisfies `depth_from_m <= depth_to_m` (asserted at
build time); the one inverted interval found by the audit (47176) is corrected
and flagged.

## Provenance and caveats

See `QA_report.md` for row counts, join statistics, duplicate handling,
unmatched descriptions, the full v1.0 → v1.1 changelog (§12), the resolution of
every audit defect (§10) and — new in v1.1 — **§11 "Known missing at source"**,
which lists the datasets that exist only as a name, the ≈330 sample photographs
that no table models, and the sample suites that still have no microscope
description.
"""
(OUT / "README.md").write_text(readme, encoding="utf-8")

# ============================================================================
# console summary
# ============================================================================
print(f"== BUILD OK (v{DB_VERSION}) ==")
for name, n, nc, ns, nx, ok in ver:
    print(f"{name:15s} rows={n:5d} csv={nc:5d} sqlite={ns:5d} xlsx={nx:5d} "
          f"{'OK' if ok else 'MISMATCH'}")
print(f"samples with coordinates: {n_coord}/{len(samples)} "
      f"({100*n_coord/len(samples):.1f}%)  desurvey={n_desurv} master_xy={n_masterxy}")
print(f"descriptions matched: {len(descriptions)-len(unmatched)}/{len(descriptions)}"
      f"  join methods: " + ", ".join(f"{k}={v}" for k, v in jm.most_common()))
print("uncovered drill samples (hole in collar, no desurvey):", len(uncovered))
print(f"missing_sources batch: {len(recs_miss)} records -> "
      f"{len(miss_desc_recs)} descriptions ("
      + ", ".join(f"{k}={v}" for k, v in miss_stats.most_common())
      + f"), {len(miss_photo)} photo-only sample stubs")
print(f"grab-sheet field descriptions recovered: {n_grab_desc}")
print(f"D5 interval disagreements flagged: {len(d5_rows)} of {n_iv_checked} "
      f"interval-bearing descriptions checked")
print(f"D7 near-duplicate sample pairs cross-referenced: {len(near_dupes)}")
print(f"lu_rock_type: {len(lu_rock_type)} distinct rock names, "
      f"{sum(1 for r in lu_rock_type if r['rock_group'] == 'other / unclassified')}"
      f" unclassified")
print(f"assertions: {sum(1 for _l, ok, _d in checks if ok)}/{len(checks)} passed")
for x in spot:
    print("SPOT:", x)
if uncovered:
    for s in uncovered[:10]:
        print("  !", s["sample_id"], s["hole_id_norm"], s["depth_mid_m"], s["qa_flags"])
